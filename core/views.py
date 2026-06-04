import uuid
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.conf import settings
from datetime import timedelta, datetime

from .models import (
    Client, Project, Credential, Quote, QuoteItem, Invoice, InvoiceItem, Payment, CompanySettings,
    Expense, TeamMember, Task, TaskAttachment, TimeEntry, ActivityLog, Document,
    TaskComment, TaskIssue, TaskActivity,
    AMCContract, AMCPayment, CredentialRenewal,
    ProjectType, ProjectFeature, FeatureRequestLink,
    OpeningBalance, FYResetEvent,
    BankAccount, InternalTransfer,
    CompanyDocument, Partner, CapitalContribution, CompanyAsset,
)
from django.contrib.contenttypes.models import ContentType
from licensing.models import License, LicenseKey, LicenseActivation


# ============== Public Pages ==============

def interiodesk_guide(request):
    """Public getting started guide for InterioDesk desktop app."""
    return render(request, 'interiodesk/guide.html')


# ============== Authentication Views ==============

def login_view(request):
    if request.user.is_authenticated:
        # Check if user is a team member
        if hasattr(request.user, 'team_profile'):
            return redirect('team_dashboard')
        return redirect('dashboard')

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Redirect team members to their dashboard
            if hasattr(user, 'team_profile'):
                return redirect('team_dashboard')
            next_url = request.GET.get('next', 'dashboard')
            return redirect(next_url)
        else:
            messages.error(request, 'Invalid username or password.')

    return render(request, 'auth/login.html')


def logout_view(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully.')
    return redirect('login')


# ============== Dashboard ==============

@login_required
def dashboard(request):
    import json
    from dateutil.relativedelta import relativedelta

    # Get summary stats
    total_clients = Client.objects.filter(is_active=True).count()
    active_projects = Project.objects.exclude(status__in=['completed', 'cancelled']).count()

    # Pending invoices
    pending_invoices = Invoice.objects.exclude(status__in=['paid', 'cancelled'])
    pending_count = pending_invoices.count()
    pending_amount = pending_invoices.aggregate(
        total=Sum('total_amount') - Sum('amount_paid')
    )['total'] or 0

    # Revenue this month
    first_day_of_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    revenue_this_month = Payment.objects.filter(
        payment_date__gte=first_day_of_month
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Expenses this month
    expenses_this_month = Expense.objects.filter(
        date__gte=first_day_of_month.date()
    ).aggregate(total=Sum('amount'))['total'] or 0
    net_profit_this_month = float(revenue_this_month) - float(expenses_this_month)

    # Expiring credentials (next 30 days)
    expiring_soon = timezone.now().date() + timedelta(days=30)
    expiring_credentials = Credential.objects.filter(
        expiry_date__lte=expiring_soon,
        expiry_date__gte=timezone.now().date(),
        is_active=True
    ).select_related('project', 'project__client')[:5]

    # Expiring / expired company documents (badge in sidebar + tile on dashboard)
    today = timezone.now().date()
    expiring_documents_count = CompanyDocument.objects.filter(
        Q(expiry_date__lt=today) | Q(expiry_date__gte=today, expiry_date__lte=expiring_soon)
    ).count()
    expiring_documents = CompanyDocument.objects.filter(
        Q(expiry_date__lt=today) | Q(expiry_date__gte=today, expiry_date__lte=expiring_soon)
    ).order_by('expiry_date')[:5]

    # Capital invested (sum of all partner contributions)
    total_capital_invested = CapitalContribution.objects.aggregate(t=Sum('amount'))['t'] or 0
    partner_count = Partner.objects.filter(is_active=True).count()

    # License statistics
    from licensing.models import License
    all_licenses = License.objects.all()
    active_licenses = all_licenses.filter(status='active', valid_until__gte=timezone.now()).count()
    expiring_licenses = all_licenses.filter(
        status='active',
        valid_until__gte=timezone.now(),
        valid_until__lte=timezone.now() + timedelta(days=30)
    ).count()
    expired_licenses = all_licenses.filter(
        Q(status='expired') | Q(valid_until__lt=timezone.now())
    ).count()
    total_licenses = all_licenses.count()

    # Get licenses expiring soon for alerts
    licenses_expiring_soon = License.objects.filter(
        status='active',
        valid_until__gte=timezone.now(),
        valid_until__lte=timezone.now() + timedelta(days=30)
    ).select_related('client').order_by('valid_until')[:5]

    # Get expired licenses for alerts
    licenses_expired = License.objects.filter(
        Q(status='expired') | Q(valid_until__lt=timezone.now(), status='active')
    ).select_related('client').order_by('-valid_until')[:5]

    # Overdue invoices
    overdue_invoices = Invoice.objects.filter(
        due_date__lt=timezone.now().date(),
        status__in=['sent', 'viewed', 'partial']
    ).select_related('client')[:5]

    # Recent payments
    recent_payments = Payment.objects.select_related(
        'invoice', 'invoice__client'
    ).order_by('-payment_date')[:5]

    # Recent invoices
    recent_invoices = Invoice.objects.select_related('client').order_by('-created_at')[:5]

    # ============== Chart Data ==============

    # Monthly Revenue & Expenses (Last 6 months)
    monthly_revenue_labels = []
    monthly_revenue_data = []
    monthly_expenses_data = []
    today = timezone.now().date()

    for i in range(5, -1, -1):
        month_date = today - relativedelta(months=i)
        month_start = month_date.replace(day=1)
        if i > 0:
            month_end = (month_date + relativedelta(months=1)).replace(day=1) - timedelta(days=1)
        else:
            month_end = today

        month_revenue = Payment.objects.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        month_expense = Expense.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_revenue_labels.append(month_date.strftime('%b %Y'))
        monthly_revenue_data.append(float(month_revenue))
        monthly_expenses_data.append(float(month_expense))

    # Project Status Distribution
    project_status_data = {}
    for status_code, status_label in Project.STATUS_CHOICES:
        count = Project.objects.filter(status=status_code).count()
        if count > 0:
            project_status_data[status_label] = count

    # Invoice Status Distribution
    invoice_status_data = {}
    for status_code, status_label in Invoice.STATUS_CHOICES:
        count = Invoice.objects.filter(status=status_code).count()
        if count > 0:
            invoice_status_data[status_label] = count

    # Payment Method Distribution
    payment_method_data = {}
    for method_code, method_label in Payment.METHOD_CHOICES:
        total = Payment.objects.filter(payment_method=method_code).aggregate(
            total=Sum('amount')
        )['total'] or 0
        if total > 0:
            payment_method_data[method_label] = float(total)

    # Total revenue (all time)
    total_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0

    # ============== Dues & Renewals ==============
    # AMC dues
    overdue_amc = AMCContract.objects.filter(
        status='active', next_due_date__lt=today
    ).select_related('project', 'project__client')
    upcoming_amc = AMCContract.objects.filter(
        status='active', next_due_date__range=[today, today + timedelta(days=30)]
    ).select_related('project', 'project__client')
    total_amc_overdue = overdue_amc.aggregate(total=Sum('annual_amount'))['total'] or 0
    total_amc_upcoming = upcoming_amc.aggregate(total=Sum('annual_amount'))['total'] or 0

    # Credential renewals due
    expired_credentials = Credential.objects.filter(
        expiry_date__lt=today, is_active=True
    ).select_related('project', 'project__client')
    expiring_credentials_30 = Credential.objects.filter(
        expiry_date__range=[today, today + timedelta(days=30)], is_active=True
    ).select_related('project', 'project__client')
    total_credential_renewal_cost = (
        (expired_credentials.aggregate(total=Sum('renewal_cost'))['total'] or 0) +
        (expiring_credentials_30.aggregate(total=Sum('renewal_cost'))['total'] or 0)
    )

    # ============== Cash Position (per-account, see core/cash_position.py) ==============
    from .cash_position import cash_position, pending_transfers
    opening = OpeningBalance.current()
    cp = cash_position()
    account_balances = cp['accounts']
    total_assets = cp['total']
    total_other_assets = cp['other_assets']
    total_assets_combined = cp['total_with_assets']
    pending_transfer_qs = pending_transfers()
    # Back-compat aliases for any template/code still reading these names
    cash_card = next((r for r in account_balances if r['account'].is_cash), None)
    bank_card = next((r for r in account_balances if r['account'].is_primary_bank), None)
    current_cash_in_hand = cash_card['balance'] if cash_card else None
    current_cash_in_account = bank_card['balance'] if bank_card else None

    # Combined upcoming dues (sorted by date for timeline)
    upcoming_dues = []
    for amc in overdue_amc[:10]:
        upcoming_dues.append({
            'date': amc.next_due_date, 'type': 'AMC', 'name': amc.project.name,
            'client': amc.project.client.name, 'amount': amc.annual_amount, 'overdue': True,
            'url': f'/amc/{amc.pk}/',
        })
    for amc in upcoming_amc[:10]:
        upcoming_dues.append({
            'date': amc.next_due_date, 'type': 'AMC', 'name': amc.project.name,
            'client': amc.project.client.name, 'amount': amc.annual_amount, 'overdue': False,
            'url': f'/amc/{amc.pk}/',
        })
    for cred in expired_credentials[:10]:
        upcoming_dues.append({
            'date': cred.expiry_date, 'type': 'Credential', 'name': cred.name,
            'client': cred.project.client.name, 'amount': cred.renewal_cost or 0, 'overdue': True,
            'url': f'/credentials/{cred.pk}/',
        })
    for cred in expiring_credentials_30[:10]:
        upcoming_dues.append({
            'date': cred.expiry_date, 'type': 'Credential', 'name': cred.name,
            'client': cred.project.client.name, 'amount': cred.renewal_cost or 0, 'overdue': False,
            'url': f'/credentials/{cred.pk}/',
        })
    upcoming_dues.sort(key=lambda x: x['date'])
    total_dues = float(total_amc_overdue + total_amc_upcoming + total_credential_renewal_cost)

    context = {
        'total_clients': total_clients,
        'active_projects': active_projects,
        'pending_count': pending_count,
        'pending_amount': pending_amount,
        'revenue_this_month': revenue_this_month,
        'total_revenue': total_revenue,
        'expenses_this_month': expenses_this_month,
        'net_profit_this_month': net_profit_this_month,
        'expiring_credentials': expiring_credentials,
        'expiring_documents': expiring_documents,
        'expiring_documents_count': expiring_documents_count,
        'total_capital_invested': total_capital_invested,
        'partner_count': partner_count,
        'overdue_invoices': overdue_invoices,
        'recent_payments': recent_payments,
        'recent_invoices': recent_invoices,
        # License data
        'active_licenses': active_licenses,
        'expiring_licenses': expiring_licenses,
        'expired_licenses': expired_licenses,
        'total_licenses': total_licenses,
        'licenses_expiring_soon': licenses_expiring_soon,
        'licenses_expired': licenses_expired,
        # Chart data as JSON
        'monthly_revenue_labels': json.dumps(monthly_revenue_labels),
        'monthly_revenue_data': json.dumps(monthly_revenue_data),
        'monthly_expenses_data': json.dumps(monthly_expenses_data),
        'project_status_labels': json.dumps(list(project_status_data.keys())),
        'project_status_data': json.dumps(list(project_status_data.values())),
        'invoice_status_labels': json.dumps(list(invoice_status_data.keys())),
        'invoice_status_data': json.dumps(list(invoice_status_data.values())),
        'payment_method_labels': json.dumps(list(payment_method_data.keys())),
        'payment_method_data': json.dumps(list(payment_method_data.values())),
        # Dues & Renewals
        'overdue_amc': overdue_amc[:5],
        'upcoming_amc': upcoming_amc[:5],
        'overdue_amc_count': overdue_amc.count(),
        'upcoming_amc_count': upcoming_amc.count(),
        'total_amc_overdue': total_amc_overdue,
        'total_amc_upcoming': total_amc_upcoming,
        'expired_credentials_list': expired_credentials[:5],
        'expired_credentials_count': expired_credentials.count(),
        'expiring_credentials_30': expiring_credentials_30[:5],
        'total_credential_renewal_cost': total_credential_renewal_cost,
        'upcoming_dues': upcoming_dues[:10],
        'total_dues': total_dues,
        # Cash position
        'opening_balance': opening,
        'current_cash_in_hand': current_cash_in_hand,
        'current_cash_in_account': current_cash_in_account,
        'account_balances': account_balances,
        'total_assets': total_assets,
        'total_other_assets': total_other_assets,
        'total_assets_combined': total_assets_combined,
        'pending_transfers': pending_transfer_qs,
        'pending_transfer_count': pending_transfer_qs.count(),
    }
    return render(request, 'dashboard/index.html', context)


# ============== Clients ==============

@login_required
def client_list(request):
    clients = Client.objects.all()

    # Search
    search = request.GET.get('search', '')
    if search:
        clients = clients.filter(
            Q(name__icontains=search) |
            Q(company_name__icontains=search) |
            Q(email__icontains=search)
        )

    # Filter by priority
    priority = request.GET.get('priority', '')
    if priority:
        clients = clients.filter(priority=priority)

    # Filter by status
    status = request.GET.get('status', '')
    if status == 'active':
        clients = clients.filter(is_active=True)
    elif status == 'inactive':
        clients = clients.filter(is_active=False)

    context = {
        'clients': clients,
        'search': search,
        'priority': priority,
        'status': status,
    }
    return render(request, 'clients/list.html', context)


@login_required
def client_detail(request, pk):
    from retailease.models import Backup, Business

    client = get_object_or_404(Client, pk=pk)
    projects = client.projects.all()
    invoices = client.invoices.all()
    quotes = client.quotes.all()
    licenses = client.licenses.all()  # Licenses linked to this client

    # Get payments through invoices
    payments = Payment.objects.filter(invoice__client=client).order_by('-payment_date')

    # Get backups through: Client -> Licenses -> Businesses -> Backups
    license_ids = licenses.values_list('id', flat=True)
    businesses = Business.objects.filter(license_id__in=license_ids)
    backups = Backup.objects.filter(business__in=businesses).select_related('business', 'counter').order_by('-created_at')

    context = {
        'client': client,
        'projects': projects,
        'invoices': invoices,
        'quotes': quotes,
        'licenses': licenses,
        'payments': payments,
        'backups': backups,
    }
    return render(request, 'clients/detail.html', context)


@login_required
def client_update_retailease(request, pk):
    """Update RetailEase App settings for a client"""
    client = get_object_or_404(Client, pk=pk)

    if request.method == 'POST':
        # Google OAuth Credentials
        client.google_client_id = request.POST.get('google_client_id', '')
        client.google_client_secret = request.POST.get('google_client_secret', '')
        client.google_client_id_ios = request.POST.get('google_client_id_ios', '')
        client.google_client_id_android = request.POST.get('google_client_id_android', '')
        client.google_reversed_client_id = request.POST.get('google_reversed_client_id', '')

        # Backup Features
        client.retailease_google_drive_enabled = 'retailease_google_drive_enabled' in request.POST
        client.retailease_server_backup_enabled = 'retailease_server_backup_enabled' in request.POST
        client.retailease_local_backup_enabled = 'retailease_local_backup_enabled' in request.POST

        # App Version Control
        client.retailease_min_version = request.POST.get('retailease_min_version', '1.0.0')
        client.retailease_latest_version = request.POST.get('retailease_latest_version', '1.0.0')
        client.retailease_update_url = request.POST.get('retailease_update_url', '')
        client.retailease_force_update = 'retailease_force_update' in request.POST

        # Maintenance Mode
        client.retailease_maintenance_mode = 'retailease_maintenance_mode' in request.POST
        client.retailease_maintenance_message = request.POST.get('retailease_maintenance_message', '')

        # Support Contact
        client.retailease_support_email = request.POST.get('retailease_support_email', '')
        client.retailease_support_phone = request.POST.get('retailease_support_phone', '')
        client.retailease_support_whatsapp = request.POST.get('retailease_support_whatsapp', '')

        client.save()
        messages.success(request, 'RetailEase App settings updated successfully.')

    return redirect('client_detail', pk=pk)


@login_required
def client_create(request):
    if request.method == 'POST':
        client = Client.objects.create(
            name=request.POST.get('name'),
            company_name=request.POST.get('company_name', ''),
            email=request.POST.get('email'),
            phone=request.POST.get('phone', ''),
            whatsapp=request.POST.get('whatsapp', ''),
            address=request.POST.get('address', ''),
            gst_number=request.POST.get('gst_number', ''),
            priority=request.POST.get('priority', 'medium'),
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'Client "{client}" created successfully.')
        return redirect('client_detail', pk=client.pk)

    return render(request, 'clients/form.html', {'form_title': 'Add New Client'})


@login_required
def client_update(request, pk):
    client = get_object_or_404(Client, pk=pk)

    if request.method == 'POST':
        client.name = request.POST.get('name')
        client.company_name = request.POST.get('company_name', '')
        client.email = request.POST.get('email')
        client.phone = request.POST.get('phone', '')
        client.whatsapp = request.POST.get('whatsapp', '')
        client.address = request.POST.get('address', '')
        client.gst_number = request.POST.get('gst_number', '')
        client.priority = request.POST.get('priority', 'medium')
        client.notes = request.POST.get('notes', '')
        client.is_active = request.POST.get('is_active') == 'on'
        client.save()

        messages.success(request, f'Client "{client}" updated successfully.')
        return redirect('client_detail', pk=client.pk)

    return render(request, 'clients/form.html', {
        'client': client,
        'form_title': 'Edit Client'
    })


# ============== Client Delete ==============

@login_required
def client_delete(request, pk):
    client = get_object_or_404(Client, pk=pk)

    if request.method == 'POST':
        client_name = str(client)
        # Check for related records
        project_count = client.projects.count()
        invoice_count = client.invoices.count()
        quote_count = client.quotes.count()

        if project_count > 0 or invoice_count > 0 or quote_count > 0:
            messages.error(
                request,
                f'Cannot delete "{client_name}". It has {project_count} projects, '
                f'{invoice_count} invoices, and {quote_count} quotes associated with it.'
            )
            return redirect('client_detail', pk=pk)

        client.delete()
        messages.success(request, f'Client "{client_name}" deleted successfully.')
        return redirect('client_list')

    return render(request, 'clients/delete.html', {'client': client})


# ============== Client Portal Account Management ==============

@login_required
def client_portal_create_account(request, pk):
    """Create a portal login account for a client"""
    from django.contrib.auth.models import User
    from django.utils.crypto import get_random_string

    client = get_object_or_404(Client, pk=pk)

    if client.user:
        messages.warning(request, f'Client already has a portal account (username: {client.user.username}).')
        return redirect('client_detail', pk=pk)

    if request.method == 'POST':
        username = request.POST.get('username', '').strip() or client.email or f'client_{client.id.hex[:8]}'
        password = request.POST.get('password', '').strip() or get_random_string(12)

        if User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" already exists. Choose a different one.')
            return redirect('client_detail', pk=pk)

        user = User.objects.create_user(
            username=username,
            email=client.email,
            password=password,
            first_name=client.name.split()[0] if client.name else '',
            last_name=' '.join(client.name.split()[1:]) if client.name and len(client.name.split()) > 1 else '',
        )
        client.user = user
        client.save(update_fields=['user'])

        messages.success(
            request,
            f'Portal account created! Username: {username} | Password: {password} — '
            f'Share these credentials with the client. They can login at /portal/login/'
        )

    return redirect('client_detail', pk=pk)


@login_required
def client_portal_reset_password(request, pk):
    """Reset a client's portal password"""
    from django.utils.crypto import get_random_string

    client = get_object_or_404(Client, pk=pk)

    if not client.user:
        messages.error(request, 'Client has no portal account.')
        return redirect('client_detail', pk=pk)

    if request.method == 'POST':
        new_password = get_random_string(12)
        client.user.set_password(new_password)
        client.user.save()
        messages.success(
            request,
            f'Password reset! New password: {new_password} — Share this with the client.'
        )

    return redirect('client_detail', pk=pk)


@login_required
def client_portal_toggle(request, pk):
    """Activate/deactivate a client's portal access"""
    client = get_object_or_404(Client, pk=pk)

    if not client.user:
        messages.error(request, 'Client has no portal account.')
        return redirect('client_detail', pk=pk)

    if request.method == 'POST':
        client.user.is_active = not client.user.is_active
        client.user.save(update_fields=['is_active'])
        state = 'activated' if client.user.is_active else 'deactivated'
        messages.success(request, f'Portal access {state} for {client.name}.')

    return redirect('client_detail', pk=pk)


# ============== Projects ==============

@login_required
def project_post_update(request, pk):
    """Post a project update visible to clients"""
    from client_portal.models import ProjectUpdate, ClientNotification

    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        update_type = request.POST.get('update_type', 'progress')
        progress_raw = request.POST.get('progress_percentage', '').strip()
        progress = int(progress_raw) if progress_raw else None
        is_visible = 'is_visible_to_client' in request.POST

        if title and description:
            update = ProjectUpdate.objects.create(
                project=project,
                author=request.user,
                title=title,
                description=description,
                update_type=update_type,
                progress_percentage=progress,
                is_visible_to_client=is_visible,
            )

            # Auto-notify client
            if is_visible and project.client.user:
                ClientNotification.objects.create(
                    client=project.client,
                    title=f'Project Update: {title}',
                    message=description[:200],
                    notification_type='project_update',
                    project=project,
                )

            messages.success(request, 'Update posted successfully.')
        else:
            messages.error(request, 'Title and description are required.')

    return redirect('project_detail', pk=pk)


@login_required
def project_delete_update(request, pk, update_pk):
    """Delete a project update"""
    from client_portal.models import ProjectUpdate

    project = get_object_or_404(Project, pk=pk)
    update = get_object_or_404(ProjectUpdate, pk=update_pk, project=project)

    if request.method == 'POST':
        update.delete()
        messages.success(request, 'Update deleted.')

    return redirect('project_detail', pk=pk)


@login_required
def project_reply_comment(request, pk):
    """Reply to client comments from admin"""
    from client_portal.models import ProjectComment

    project = get_object_or_404(Project, pk=pk)
    if request.method == 'POST':
        message = request.POST.get('message', '').strip()
        if message:
            ProjectComment.objects.create(
                project=project,
                author=request.user,
                message=message,
                is_internal=False,
            )
            messages.success(request, 'Reply posted.')

    return redirect('project_detail', pk=pk)


@login_required
def project_list(request):
    projects = Project.objects.select_related('client').all()

    # Search
    search = request.GET.get('search', '')
    if search:
        projects = projects.filter(
            Q(name__icontains=search) |
            Q(client__name__icontains=search) |
            Q(client__company_name__icontains=search)
        )

    # Filter by status
    status = request.GET.get('status', '')
    if status:
        projects = projects.filter(status=status)

    # Filter by type
    project_type = request.GET.get('type', '')
    if project_type:
        projects = projects.filter(project_type=project_type)

    context = {
        'projects': projects,
        'search': search,
        'status': status,
        'project_type': project_type,
        'status_choices': Project.STATUS_CHOICES,
        'type_choices': Project.TYPE_CHOICES,
    }
    return render(request, 'projects/list.html', context)


@login_required
def project_detail(request, pk):
    project = get_object_or_404(Project.objects.select_related('client'), pk=pk)
    credentials = project.credentials.all()
    invoices = project.invoices.all()
    quotes = project.quotes.all()

    # Get all payments for invoices related to this project
    payments = Payment.objects.filter(invoice__project=project).select_related('invoice').order_by('-payment_date')

    # Calculate financial stats
    from django.db.models import Sum
    from decimal import Decimal

    # Total project cost (use final_amount if set, otherwise estimated_budget)
    total_project_cost = project.final_amount or project.estimated_budget or Decimal('0')

    # Total invoiced amount for this project
    total_invoiced = invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

    # Total amount received (sum of all payments)
    amount_received = payments.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # Pending amount = Total Project Cost - Amount Received
    pending_amount = total_project_cost - amount_received

    # Client portal data
    from client_portal.models import ProjectUpdate, ProjectComment
    project_updates = ProjectUpdate.objects.filter(project=project).select_related('author')
    client_comments = ProjectComment.objects.filter(
        project=project, is_internal=False, parent__isnull=True
    ).select_related('author').order_by('-created_at')

    # AMC contracts
    amc_contracts = project.amc_contracts.all()
    amc_payments = AMCPayment.objects.filter(amc__project=project).order_by('-payment_date')

    context = {
        'project': project,
        'credentials': credentials,
        'invoices': invoices,
        'quotes': quotes,
        'payments': payments,
        'total_project_cost': total_project_cost,
        'total_invoiced': total_invoiced,
        'amount_received': amount_received,
        'pending_amount': pending_amount,
        'project_updates': project_updates,
        'client_comments': client_comments,
        'amc_contracts': amc_contracts,
        'amc_payments': amc_payments,
    }
    return render(request, 'projects/detail.html', context)


@login_required
def project_create(request):
    clients = Client.objects.filter(is_active=True)
    team_members = TeamMember.objects.filter(is_active=True)

    if request.method == 'POST':
        project = Project.objects.create(
            client_id=request.POST.get('client'),
            name=request.POST.get('name'),
            project_type=request.POST.get('project_type', 'web_app'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'lead'),
            estimated_budget=request.POST.get('estimated_budget') or None,
            start_date=request.POST.get('start_date') or None,
            deadline=request.POST.get('deadline') or None,
            tech_stack=request.POST.get('tech_stack', ''),
            notes=request.POST.get('notes', ''),
        )
        # Assign team members
        selected_members = request.POST.getlist('team_members')
        if selected_members:
            project.team_members.set(selected_members)
        messages.success(request, f'Project "{project.name}" created successfully.')
        return redirect('project_detail', pk=project.pk)

    return render(request, 'projects/form.html', {
        'clients': clients,
        'team_members': team_members,
        'form_title': 'Add New Project',
        'status_choices': Project.STATUS_CHOICES,
        'type_choices': Project.TYPE_CHOICES,
    })


@login_required
def project_update(request, pk):
    project = get_object_or_404(Project, pk=pk)
    clients = Client.objects.filter(is_active=True)
    team_members = TeamMember.objects.filter(is_active=True)

    if request.method == 'POST':
        project.client_id = request.POST.get('client')
        project.name = request.POST.get('name')
        project.project_type = request.POST.get('project_type', 'web_app')
        project.description = request.POST.get('description', '')
        project.status = request.POST.get('status', 'lead')
        project.estimated_budget = request.POST.get('estimated_budget') or None
        project.final_amount = request.POST.get('final_amount') or None
        project.start_date = request.POST.get('start_date') or None
        project.deadline = request.POST.get('deadline') or None
        project.completed_date = request.POST.get('completed_date') or None
        project.tech_stack = request.POST.get('tech_stack', '')
        project.github_repo = request.POST.get('github_repo', '')
        project.live_url = request.POST.get('live_url', '')
        project.notes = request.POST.get('notes', '')
        # Completion & AMC fields
        project.warranty_period = request.POST.get('warranty_period') or None
        project.completion_notes = request.POST.get('completion_notes', '')
        project.deliverables = request.POST.get('deliverables', '')
        project.amc_amount = request.POST.get('amc_amount') or None
        project.amc_billing_cycle = request.POST.get('amc_billing_cycle', '')

        # Auto-set completed_date when status changes to completed
        old_status = Project.objects.filter(pk=pk).values_list('status', flat=True).first()
        if project.status == 'completed' and old_status != 'completed' and not project.completed_date:
            project.completed_date = timezone.now().date()

        project.save()

        # Auto-create AMC contract when project is completed with AMC amount
        if project.status == 'completed' and project.amc_amount and not project.amc_contracts.exists():
            from dateutil.relativedelta import relativedelta
            start = project.completed_date or timezone.now().date()
            cycle = project.amc_billing_cycle or 'yearly'
            cycle_map = {
                'monthly': relativedelta(months=1),
                'quarterly': relativedelta(months=3),
                'half_yearly': relativedelta(months=6),
                'yearly': relativedelta(years=1),
            }
            AMCContract.objects.create(
                project=project,
                annual_amount=project.amc_amount,
                billing_cycle=cycle,
                start_date=start,
                end_date=start + relativedelta(years=1),
                next_due_date=start + cycle_map[cycle],
                status='active',
            )
            messages.info(request, 'AMC contract created automatically.')

        # Update team members
        selected_members = request.POST.getlist('team_members')
        project.team_members.set(selected_members)

        messages.success(request, f'Project "{project.name}" updated successfully.')
        return redirect('project_detail', pk=project.pk)

    return render(request, 'projects/form.html', {
        'project': project,
        'clients': clients,
        'team_members': team_members,
        'form_title': 'Edit Project',
        'status_choices': Project.STATUS_CHOICES,
        'type_choices': Project.TYPE_CHOICES,
    })


# ============== Project Delete ==============

@login_required
def project_delete(request, pk):
    project = get_object_or_404(Project, pk=pk)

    if request.method == 'POST':
        project_name = project.name
        client_pk = project.client.pk

        # Check for related records
        credential_count = project.credentials.count()
        invoice_count = project.invoices.count()
        quote_count = project.quotes.count()

        if credential_count > 0 or invoice_count > 0 or quote_count > 0:
            messages.error(
                request,
                f'Cannot delete "{project_name}". It has {credential_count} credentials, '
                f'{invoice_count} invoices, and {quote_count} quotes associated with it.'
            )
            return redirect('project_detail', pk=pk)

        project.delete()
        messages.success(request, f'Project "{project_name}" deleted successfully.')
        return redirect('client_detail', pk=client_pk)

    return render(request, 'projects/delete.html', {'project': project})


# ============== Credentials ==============

@login_required
def credential_list(request):
    credentials = Credential.objects.select_related('project', 'project__client').all()

    # Search
    search = request.GET.get('search', '')
    if search:
        credentials = credentials.filter(
            Q(name__icontains=search) |
            Q(provider__icontains=search) |
            Q(project__name__icontains=search)
        )

    # Filter by type
    cred_type = request.GET.get('type', '')
    if cred_type:
        credentials = credentials.filter(credential_type=cred_type)

    # Filter by expiry status
    expiry = request.GET.get('expiry', '')
    today = timezone.now().date()
    if expiry == 'expired':
        credentials = credentials.filter(expiry_date__lt=today)
    elif expiry == 'expiring':
        credentials = credentials.filter(
            expiry_date__gte=today,
            expiry_date__lte=today + timedelta(days=30)
        )

    context = {
        'credentials': credentials,
        'search': search,
        'cred_type': cred_type,
        'expiry': expiry,
        'type_choices': Credential.TYPE_CHOICES,
    }
    return render(request, 'credentials/list.html', context)


@login_required
def credential_detail(request, pk):
    credential = get_object_or_404(
        Credential.objects.select_related('project', 'project__client'),
        pk=pk
    )
    return render(request, 'credentials/detail.html', {'credential': credential})


@login_required
def credential_create(request):
    projects = Project.objects.select_related('client').all()

    if request.method == 'POST':
        credential = Credential.objects.create(
            project_id=request.POST.get('project'),
            name=request.POST.get('name'),
            credential_type=request.POST.get('credential_type', 'hosting'),
            provider=request.POST.get('provider', ''),
            username=request.POST.get('username', ''),
            password=request.POST.get('password', ''),
            url=request.POST.get('url', ''),
            expiry_date=request.POST.get('expiry_date') or None,
            notes=request.POST.get('notes', ''),
            client_visible=request.POST.get('client_visible') == 'on',
        )
        messages.success(request, f'Credential "{credential.name}" created successfully.')
        return redirect('credential_detail', pk=credential.pk)

    return render(request, 'credentials/form.html', {
        'projects': projects,
        'form_title': 'Add New Credential',
        'type_choices': Credential.TYPE_CHOICES,
    })


@login_required
def credential_update(request, pk):
    credential = get_object_or_404(Credential, pk=pk)
    projects = Project.objects.select_related('client').all()

    if request.method == 'POST':
        credential.project_id = request.POST.get('project')
        credential.name = request.POST.get('name')
        credential.credential_type = request.POST.get('credential_type', 'hosting')
        credential.provider = request.POST.get('provider', '')
        credential.username = request.POST.get('username', '')
        credential.password = request.POST.get('password', '')
        credential.url = request.POST.get('url', '')
        credential.expiry_date = request.POST.get('expiry_date') or None
        credential.notes = request.POST.get('notes', '')
        credential.is_active = request.POST.get('is_active') == 'on'
        credential.client_visible = request.POST.get('client_visible') == 'on'
        credential.save()

        messages.success(request, f'Credential "{credential.name}" updated successfully.')
        return redirect('credential_detail', pk=credential.pk)

    return render(request, 'credentials/form.html', {
        'credential': credential,
        'projects': projects,
        'form_title': 'Edit Credential',
        'type_choices': Credential.TYPE_CHOICES,
    })


@login_required
def credential_expiry(request):
    today = timezone.now().date()

    # Expired
    expired = Credential.objects.filter(
        expiry_date__lt=today, is_active=True
    ).select_related('project', 'project__client')

    # Expiring this week
    this_week = Credential.objects.filter(
        expiry_date__gte=today,
        expiry_date__lte=today + timedelta(days=7),
        is_active=True
    ).select_related('project', 'project__client')

    # Expiring this month
    this_month = Credential.objects.filter(
        expiry_date__gt=today + timedelta(days=7),
        expiry_date__lte=today + timedelta(days=30),
        is_active=True
    ).select_related('project', 'project__client')

    context = {
        'expired': expired,
        'this_week': this_week,
        'this_month': this_month,
    }
    return render(request, 'credentials/expiry.html', context)


# ============== Credential Delete ==============

@login_required
def credential_delete(request, pk):
    credential = get_object_or_404(Credential, pk=pk)

    if request.method == 'POST':
        credential_name = credential.name
        project_pk = credential.project.pk
        credential.delete()
        messages.success(request, f'Credential "{credential_name}" deleted successfully.')
        return redirect('project_detail', pk=project_pk)

    return render(request, 'credentials/delete.html', {'credential': credential})


# ============== Credential Renewal ==============

@login_required
def credential_renew(request, pk):
    credential = get_object_or_404(Credential, pk=pk)
    if request.method == 'POST':
        new_expiry = request.POST.get('new_expiry_date')
        cost = request.POST.get('cost') or None
        notes = request.POST.get('notes', '')
        if new_expiry:
            CredentialRenewal.objects.create(
                credential=credential,
                old_expiry=credential.expiry_date,
                new_expiry=new_expiry,
                cost=cost,
                notes=notes,
            )
            credential.expiry_date = new_expiry
            credential.last_renewed_date = timezone.now().date()
            credential.save()
            messages.success(request, f'Credential "{credential.name}" renewed successfully.')
        else:
            messages.error(request, 'New expiry date is required.')
        return redirect('project_detail', pk=credential.project.pk)
    return redirect('project_detail', pk=credential.project.pk)


# ============== AMC Contracts ==============

@login_required
def amc_list(request):
    contracts = AMCContract.objects.select_related('project', 'project__client').all()

    status_filter = request.GET.get('status', '')
    if status_filter:
        contracts = contracts.filter(status=status_filter)

    type_filter = request.GET.get('type', '')
    if type_filter:
        contracts = contracts.filter(contract_type=type_filter)

    search = request.GET.get('search', '')
    if search:
        contracts = contracts.filter(
            Q(project__name__icontains=search) |
            Q(project__client__name__icontains=search)
        )

    today = timezone.now().date()
    overdue_count = AMCContract.objects.filter(status='active', next_due_date__lt=today).count()
    due_soon_count = AMCContract.objects.filter(status='active', next_due_date__range=[today, today + timedelta(days=30)]).count()
    active_count = AMCContract.objects.filter(status='active').count()
    total_annual = AMCContract.objects.filter(status='active').aggregate(total=Sum('annual_amount'))['total'] or 0

    return render(request, 'amc/list.html', {
        'contracts': contracts,
        'status_filter': status_filter,
        'type_filter': type_filter,
        'search': search,
        'overdue_count': overdue_count,
        'due_soon_count': due_soon_count,
        'active_count': active_count,
        'total_annual': total_annual,
        'status_choices': AMCContract.STATUS_CHOICES,
        'type_choices': AMCContract.CONTRACT_TYPE_CHOICES,
    })


@login_required
def amc_detail(request, pk):
    amc = get_object_or_404(AMCContract.objects.select_related('project', 'project__client'), pk=pk)
    payments = amc.payments.all()
    return render(request, 'amc/detail.html', {
        'amc': amc,
        'payments': payments,
    })


@login_required
def amc_create(request):
    if request.method == 'POST':
        from dateutil.relativedelta import relativedelta
        project_id = request.POST.get('project')
        project = get_object_or_404(Project, pk=project_id)
        contract_type = request.POST.get('contract_type', 'amc')
        annual_amount = request.POST.get('annual_amount')
        billing_cycle = request.POST.get('billing_cycle', 'yearly')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        notes = request.POST.get('notes', '')
        auto_renew = request.POST.get('auto_renew') == 'on'

        from datetime import datetime
        start = datetime.strptime(start_date, '%Y-%m-%d').date()
        end = datetime.strptime(end_date, '%Y-%m-%d').date()

        cycle_map = {
            'monthly': relativedelta(months=1),
            'quarterly': relativedelta(months=3),
            'half_yearly': relativedelta(months=6),
            'yearly': relativedelta(years=1),
        }
        next_due = start + cycle_map.get(billing_cycle, relativedelta(years=1))

        amc = AMCContract.objects.create(
            project=project,
            contract_type=contract_type,
            annual_amount=annual_amount,
            billing_cycle=billing_cycle,
            start_date=start,
            end_date=end,
            next_due_date=next_due,
            auto_renew=auto_renew,
            notes=notes,
        )
        messages.success(request, f'{amc.get_contract_type_display()} contract created for "{project.name}".')
        return redirect('amc_detail', pk=amc.pk)

    projects = Project.objects.select_related('client').all()
    preselected_project = request.GET.get('project', '')
    return render(request, 'amc/form.html', {
        'projects': projects,
        'preselected_project': preselected_project,
        'form_title': 'Create Recurring Contract',
        'billing_choices': AMCContract.BILLING_CYCLE_CHOICES,
        'type_choices': AMCContract.CONTRACT_TYPE_CHOICES,
    })


@login_required
def amc_update(request, pk):
    amc = get_object_or_404(AMCContract.objects.select_related('project'), pk=pk)
    if request.method == 'POST':
        amc.contract_type = request.POST.get('contract_type', 'amc')
        amc.annual_amount = request.POST.get('annual_amount')
        amc.billing_cycle = request.POST.get('billing_cycle', 'yearly')
        amc.start_date = request.POST.get('start_date')
        amc.end_date = request.POST.get('end_date')
        amc.next_due_date = request.POST.get('next_due_date')
        amc.status = request.POST.get('status', 'active')
        amc.auto_renew = request.POST.get('auto_renew') == 'on'
        amc.notes = request.POST.get('notes', '')
        amc.save()
        messages.success(request, 'Contract updated.')
        return redirect('amc_detail', pk=amc.pk)

    projects = Project.objects.select_related('client').all()
    return render(request, 'amc/form.html', {
        'amc': amc,
        'projects': projects,
        'preselected_project': str(amc.project_id),
        'form_title': 'Edit Contract',
        'billing_choices': AMCContract.BILLING_CYCLE_CHOICES,
        'status_choices': AMCContract.STATUS_CHOICES,
        'type_choices': AMCContract.CONTRACT_TYPE_CHOICES,
    })


@login_required
def amc_delete(request, pk):
    amc = get_object_or_404(AMCContract.objects.select_related('project'), pk=pk)
    if request.method == 'POST':
        project_pk = amc.project.pk
        amc.delete()
        messages.success(request, 'AMC contract deleted.')
        return redirect('project_detail', pk=project_pk)
    return render(request, 'amc/delete.html', {'amc': amc})


@login_required
def amc_record_payment(request, pk):
    amc = get_object_or_404(AMCContract, pk=pk)
    if request.method == 'POST':
        from dateutil.relativedelta import relativedelta
        payment = AMCPayment.objects.create(
            amc=amc,
            payment_date=request.POST.get('payment_date') or timezone.now().date(),
            amount=request.POST.get('amount'),
            period_start=request.POST.get('period_start'),
            period_end=request.POST.get('period_end'),
            payment_method=request.POST.get('payment_method', 'bank_transfer'),
            reference=request.POST.get('reference', ''),
            notes=request.POST.get('notes', ''),
        )
        amc.advance_due_date()
        messages.success(request, f'Payment of ₹{payment.amount} recorded. Next due date updated.')
        return redirect('amc_detail', pk=amc.pk)
    return redirect('amc_detail', pk=amc.pk)


# ============== Project Completion Certificate ==============

@login_required
def project_completion_certificate(request, pk):
    """Generate project completion certificate as HTML/PDF"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal

    project = get_object_or_404(
        Project.objects.select_related('client').prefetch_related('team_members', 'credentials'),
        pk=pk
    )

    company = CompanySettings.get_settings()
    download = request.GET.get('download', '0') == '1'

    # Financial summary
    invoices = project.invoices.all()
    total_invoiced = invoices.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    total_paid = Payment.objects.filter(invoice__project=project).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # AMC details
    amc = project.amc_contracts.first()

    # Credentials (names and types only, no secrets)
    credentials = project.credentials.filter(is_active=True)

    # Deliverables as list
    deliverables_list = [d.strip() for d in project.deliverables.split('\n') if d.strip()] if project.deliverables else []

    context = {
        'project': project,
        'company': company,
        'total_invoiced': total_invoiced,
        'total_paid': total_paid,
        'balance_due': total_invoiced - total_paid,
        'amc': amc,
        'credentials': credentials,
        'deliverables_list': deliverables_list,
        'team_members': project.team_members.filter(is_active=True),
    }

    if download:
        try:
            from weasyprint import HTML

            html_string = render_to_string('projects/completion_certificate.html', context)
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf = html.write_pdf()

            safe_name = project.name.replace(' ', '_')[:50]
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="completion_certificate_{safe_name}.pdf"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'projects/completion_certificate.html', context)


# ============== Quotes ==============

@login_required
def quote_list(request):
    quotes = Quote.objects.select_related('client', 'project').all()

    # Get stats counts
    all_quotes = Quote.objects.all()
    draft_count = all_quotes.filter(status='draft').count()
    sent_count = all_quotes.filter(status='sent').count()
    accepted_count = all_quotes.filter(status='accepted').count()
    rejected_count = all_quotes.filter(status='rejected').count()
    # Count quotes expiring within 7 days
    from datetime import timedelta
    expiring_date = timezone.now().date() + timedelta(days=7)
    expiring_count = all_quotes.filter(
        status__in=['sent', 'viewed'],
        valid_until__lte=expiring_date,
        valid_until__gte=timezone.now().date()
    ).count()

    search = request.GET.get('search', '')
    if search:
        quotes = quotes.filter(
            Q(quote_number__icontains=search) |
            Q(title__icontains=search) |
            Q(client__name__icontains=search)
        )

    status = request.GET.get('status', '')
    if status:
        quotes = quotes.filter(status=status)

    client_filter = request.GET.get('client', '')
    if client_filter:
        quotes = quotes.filter(client_id=client_filter)

    clients = Client.objects.filter(is_active=True)

    context = {
        'quotes': quotes,
        'search': search,
        'status': status,
        'status_choices': Quote.STATUS_CHOICES,
        'clients': clients,
        'selected_client': client_filter,
        'draft_count': draft_count,
        'sent_count': sent_count,
        'accepted_count': accepted_count,
        'rejected_count': rejected_count,
        'expiring_count': expiring_count,
    }
    return render(request, 'quotes/list.html', context)


@login_required
def quote_detail(request, pk):
    quote = get_object_or_404(
        Quote.objects.select_related('client', 'project').prefetch_related('items'),
        pk=pk
    )
    return render(request, 'quotes/detail.html', {'quote': quote})


@login_required
def quote_create(request):
    from crm.models import Lead
    clients = Client.objects.filter(is_active=True)
    projects = Project.objects.select_related('client').all()

    # Quote can be raised directly for an unconverted CRM lead.
    lead_id = request.POST.get('lead') or request.GET.get('lead')
    lead = None
    if lead_id:
        lead = Lead.objects.filter(pk=lead_id).first()

    if request.method == 'POST':
        from decimal import Decimal, InvalidOperation

        if not request.POST.get('client') and not lead:
            messages.error(request, 'Select a client, or create the quote from a lead.')
            return redirect('quote_create')

        # Safe tax_rate conversion - default to 0 if empty or invalid
        try:
            tax_rate_val = request.POST.get('tax_rate', '0')
            tax_rate = Decimal(tax_rate_val) if tax_rate_val else Decimal('0')
        except (InvalidOperation, ValueError):
            tax_rate = Decimal('0')

        # Safe discount conversion
        try:
            discount_val = request.POST.get('discount', '0')
            discount = Decimal(discount_val) if discount_val else Decimal('0')
        except (InvalidOperation, ValueError):
            discount = Decimal('0')

        # Parse start_date
        start_date_val = request.POST.get('start_date', '')
        start_date = start_date_val if start_date_val else None

        quote = Quote.objects.create(
            client_id=request.POST.get('client') or None,
            lead=lead if lead else None,
            project_id=request.POST.get('project') or None,
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            issue_date=request.POST.get('issue_date') or timezone.now().date(),
            valid_until=request.POST.get('valid_until') or None,
            status=request.POST.get('status', 'draft'),
            discount=discount,
            tax_rate=tax_rate,
            notes=request.POST.get('notes', ''),
            client_notes=request.POST.get('client_notes', ''),
            terms=request.POST.get('terms', ''),
            # Timeline & Deliverables
            duration=request.POST.get('duration', ''),
            start_date=start_date,
            deliverables=request.POST.get('deliverables', ''),
            payment_terms=request.POST.get('payment_terms', '50-50'),
        )

        # Process line items
        item_count = int(request.POST.get('item_count', 0))
        for i in range(1, item_count + 10):  # Check extra indices for dynamically added items
            description = request.POST.get(f'item_description_{i}')
            if description:
                quantity = Decimal(request.POST.get(f'item_quantity_{i}', 1) or 1)
                unit_price = Decimal(request.POST.get(f'item_price_{i}', 0) or 0)
                QuoteItem.objects.create(
                    quote=quote,
                    description=description,
                    quantity=quantity,
                    unit_price=unit_price,
                    amount=quantity * unit_price
                )

        # Recalculate totals
        quote.calculate_totals()

        messages.success(request, f'Quote "{quote.quote_number}" created successfully.')
        return redirect('quote_detail', pk=quote.pk)

    # Get company settings for defaults
    company = CompanySettings.get_settings()

    # Default dates using settings
    from datetime import timedelta
    today = timezone.now().date()
    validity_days = company.default_quote_validity_days or 30
    valid_until_default = today + timedelta(days=validity_days)

    return render(request, 'quotes/form.html', {
        'clients': clients,
        'projects': projects,
        'lead': lead,
        'form_title': 'Create New Quote',
        'status_choices': Quote.STATUS_CHOICES,
        'today': today.strftime('%Y-%m-%d'),
        'valid_until_default': valid_until_default.strftime('%Y-%m-%d'),
        'company': company,
    })


@login_required
def quote_update(request, pk):
    import traceback
    import sys
    try:
        quote = get_object_or_404(Quote.objects.prefetch_related('items'), pk=pk)
        clients = Client.objects.filter(is_active=True)
        projects = Project.objects.select_related('client').all()

        if request.method == 'POST':
            from decimal import Decimal, InvalidOperation

            quote.client_id = request.POST.get('client')
            quote.project_id = request.POST.get('project') or None
            quote.title = request.POST.get('title')
            quote.description = request.POST.get('description', '')
            quote.issue_date = request.POST.get('issue_date')
            quote.valid_until = request.POST.get('valid_until') or None
            quote.status = request.POST.get('status', 'draft')

            # Safe decimal conversion - default to 0 if empty
            try:
                discount_val = request.POST.get('discount', '0')
                quote.discount = Decimal(discount_val) if discount_val else Decimal('0')
            except (InvalidOperation, ValueError):
                quote.discount = Decimal('0')

            try:
                tax_rate_val = request.POST.get('tax_rate', '0')
                quote.tax_rate = Decimal(tax_rate_val) if tax_rate_val else Decimal('0')
            except (InvalidOperation, ValueError):
                quote.tax_rate = Decimal('0')

            quote.notes = request.POST.get('notes', '')
            quote.client_notes = request.POST.get('client_notes', '')
            quote.terms = request.POST.get('terms', '')

            # Timeline & Deliverables
            quote.duration = request.POST.get('duration', '')
            start_date_val = request.POST.get('start_date', '')
            quote.start_date = start_date_val if start_date_val else None
            quote.deliverables = request.POST.get('deliverables', '')
            quote.payment_terms = request.POST.get('payment_terms', '50-50')

            quote.save()

            # Delete existing items and recreate
            quote.items.all().delete()

            # Process line items
            item_count = int(request.POST.get('item_count', 0))
            for i in range(1, item_count + 10):  # Check extra indices for dynamically added items
                description = request.POST.get(f'item_description_{i}')
                if description:
                    quantity = Decimal(request.POST.get(f'item_quantity_{i}', 1) or 1)
                    unit_price = Decimal(request.POST.get(f'item_price_{i}', 0) or 0)
                    QuoteItem.objects.create(
                        quote=quote,
                        description=description,
                        quantity=quantity,
                        unit_price=unit_price,
                        amount=quantity * unit_price
                    )

            # Recalculate totals
            quote.calculate_totals()

            messages.success(request, f'Quote "{quote.quote_number}" updated successfully.')
            return redirect('quote_detail', pk=quote.pk)

        # Get company settings for defaults
        company = CompanySettings.get_settings()

        # Default dates using settings
        from datetime import timedelta
        today = timezone.now().date()
        validity_days = company.default_quote_validity_days or 30
        valid_until_default = today + timedelta(days=validity_days)

        return render(request, 'quotes/form.html', {
            'quote': quote,
            'clients': clients,
            'projects': projects,
            'form_title': 'Edit Quote',
            'status_choices': Quote.STATUS_CHOICES,
            'today': today.strftime('%Y-%m-%d'),
            'valid_until_default': valid_until_default.strftime('%Y-%m-%d'),
            'company': company,
        })
    except Exception as e:
        print(f"ERROR in quote_update: {e}", file=sys.stderr)
        print(traceback.format_exc(), file=sys.stderr)
        raise


@login_required
def quote_pdf(request, pk):
    """Generate PDF for a quote"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal

    quote = get_object_or_404(
        Quote.objects.select_related('client', 'project').prefetch_related('items'),
        pk=pk
    )

    # Get company settings
    company = CompanySettings.get_settings()

    # Check if GST should be included
    with_gst = request.GET.get('gst', '0') == '1'
    download = request.GET.get('download', '0') == '1'

    # Calculate amounts
    taxable_amount = quote.subtotal - (quote.discount or Decimal('0'))
    # Use quote's tax_rate (0 means no tax), only default to 0 if None
    tax_rate = Decimal(str(quote.tax_rate)) if quote.tax_rate is not None else Decimal('0')

    cgst_amount = Decimal('0')
    sgst_amount = Decimal('0')
    tax_amount = Decimal('0')
    total = taxable_amount

    if with_gst:
        cgst_rate = tax_rate / 2
        sgst_rate = tax_rate / 2
        cgst_amount = taxable_amount * (cgst_rate / 100)
        sgst_amount = taxable_amount * (sgst_rate / 100)
        tax_amount = cgst_amount + sgst_amount
        total = taxable_amount + tax_amount

    context = {
        'quote': quote,
        'company': company,
        'with_gst': with_gst,
        'taxable_amount': taxable_amount,
        'tax_rate': tax_rate,
        'cgst_rate': tax_rate / 2 if with_gst else 0,
        'sgst_rate': tax_rate / 2 if with_gst else 0,
        'cgst_amount': cgst_amount,
        'sgst_amount': sgst_amount,
        'tax_amount': tax_amount,
        'total_with_gst': total,
    }

    # If download requested, generate PDF
    if download:
        try:
            from weasyprint import HTML, CSS
            from django.conf import settings
            import os

            html_string = render_to_string('quotes/pdf.html', context)

            # Create PDF
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf = html.write_pdf()

            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="quote_{quote.quote_number}.pdf"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'quotes/pdf.html', context)


# ============== Quote Delete ==============

@login_required
def quote_delete(request, pk):
    quote = get_object_or_404(Quote, pk=pk)

    if request.method == 'POST':
        quote_number = quote.quote_number
        # Check if quote has been converted to invoice
        if Invoice.objects.filter(quote=quote).exists():
            messages.error(
                request,
                f'Cannot delete "{quote_number}". It has been converted to an invoice.'
            )
            return redirect('quote_detail', pk=pk)

        quote.delete()
        messages.success(request, f'Quote "{quote_number}" deleted successfully.')
        return redirect('quote_list')

    return render(request, 'quotes/delete.html', {'quote': quote})


# ============== Quote Clone ==============

@login_required
def quote_clone(request, pk):
    """Clone an existing quote"""
    original_quote = get_object_or_404(
        Quote.objects.prefetch_related('items'),
        pk=pk
    )

    # Create new quote with copied data
    from decimal import Decimal

    # Get company settings for validity
    company = CompanySettings.get_settings()
    validity_days = company.default_quote_validity_days or 30

    today = timezone.now().date()
    valid_until = today + timedelta(days=validity_days)

    new_quote = Quote.objects.create(
        client=original_quote.client,
        lead=original_quote.lead,
        project=original_quote.project,
        title=f"Copy of {original_quote.title}",
        description=original_quote.description,
        status='draft',
        subtotal=original_quote.subtotal,
        discount=original_quote.discount,
        tax_rate=original_quote.tax_rate,
        tax_amount=original_quote.tax_amount,
        total_amount=original_quote.total_amount,
        issue_date=today,
        valid_until=valid_until,
        terms=original_quote.terms,
        client_notes=original_quote.client_notes,
        notes=original_quote.notes,
        # Copy timeline & deliverables
        duration=original_quote.duration,
        deliverables=original_quote.deliverables,
        payment_terms=original_quote.payment_terms,
    )

    # Clone all items
    for item in original_quote.items.all():
        QuoteItem.objects.create(
            quote=new_quote,
            description=item.description,
            details=item.details,
            quantity=item.quantity,
            unit_price=item.unit_price,
            amount=item.amount,
            order=item.order,
        )

    messages.success(request, f'Quote cloned successfully. New quote: {new_quote.quote_number}')
    return redirect('quote_update', pk=new_quote.pk)


# ============== Quote to Invoice Conversion ==============

@login_required
def quote_convert(request, pk):
    """Convert a quote to an invoice"""
    quote = get_object_or_404(
        Quote.objects.select_related('client', 'project').prefetch_related('items'),
        pk=pk
    )

    # Check if already converted
    if Invoice.objects.filter(quote=quote).exists():
        existing_invoice = Invoice.objects.get(quote=quote)
        messages.warning(request, f'This quote has already been converted to invoice {existing_invoice.invoice_number}.')
        return redirect('invoice_detail', pk=existing_invoice.pk)

    # A lead quote has no client yet; convert the lead first.
    if not quote.client_id:
        messages.error(request, 'This quote is for a lead. Convert the lead to a client before raising an invoice.')
        return redirect('quote_detail', pk=quote.pk)

    from decimal import Decimal

    today = timezone.now().date()
    due_date = today + timedelta(days=15)

    # Create the invoice
    invoice = Invoice.objects.create(
        client=quote.client,
        project=quote.project,
        quote=quote,
        title=quote.title,
        description=quote.description,
        status='draft',
        subtotal=quote.subtotal,
        discount=quote.discount,
        tax_rate=quote.tax_rate,
        tax_amount=quote.tax_amount,
        total_amount=quote.total_amount,
        issue_date=today,
        due_date=due_date,
        terms=quote.terms,
        client_notes=quote.client_notes,
        notes=quote.notes,
    )

    # Copy all items from quote to invoice
    for item in quote.items.all():
        InvoiceItem.objects.create(
            invoice=invoice,
            description=item.description,
            details=item.details,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total=item.amount,
            order=item.order,
        )

    # Update quote status to accepted if not already
    if quote.status not in ['accepted', 'rejected', 'expired']:
        quote.status = 'accepted'
        quote.save()

    messages.success(request, f'Quote converted to invoice {invoice.invoice_number} successfully.')
    return redirect('invoice_detail', pk=invoice.pk)


# ============== Invoices ==============

@login_required
def invoice_list(request):
    from datetime import date
    from calendar import monthrange
    from django.db.models import Sum

    invoices = Invoice.objects.select_related('client', 'project').all()

    search = request.GET.get('search', '')
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(title__icontains=search) |
            Q(client__name__icontains=search)
        )

    status = request.GET.get('status', '')
    if status == 'paid_partial':
        invoices = invoices.filter(status__in=['paid', 'partial'])
    elif status:
        invoices = invoices.filter(status=status)

    gst_filing = request.GET.get('gst_filing', '')
    if gst_filing:
        invoices = invoices.filter(gst_filing_status=gst_filing)

    # Date range filter on issue_date (used for GST-period filtering)
    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')

    # Quick preset: ?period=prev_month | this_month fills from/to if empty
    period = request.GET.get('period', '')
    if period in ('prev_month', 'this_month') and not from_date and not to_date:
        today = date.today()
        if period == 'prev_month':
            year = today.year if today.month > 1 else today.year - 1
            month = today.month - 1 if today.month > 1 else 12
        else:
            year, month = today.year, today.month
        from_date = date(year, month, 1).isoformat()
        to_date = date(year, month, monthrange(year, month)[1]).isoformat()

    if from_date:
        invoices = invoices.filter(issue_date__gte=from_date)
    if to_date:
        invoices = invoices.filter(issue_date__lte=to_date)

    invoices = invoices.order_by('-issue_date', '-created_at')

    totals = invoices.aggregate(
        subtotal=Sum('subtotal'),
        tax=Sum('tax_amount'),
        total=Sum('total_amount'),
        paid=Sum('amount_paid'),
    )

    context = {
        'invoices': invoices,
        'search': search,
        'status': status,
        'status_choices': Invoice.STATUS_CHOICES,
        'gst_filing': gst_filing,
        'gst_filing_choices': Invoice.GST_FILING_STATUS_CHOICES,
        'from_date': from_date,
        'to_date': to_date,
        'period': period,
        'totals': totals,
        'invoice_count': invoices.count(),
    }
    return render(request, 'invoices/list.html', context)


def _filter_invoices_from_post(request):
    """Re-apply invoice_list filters from POST data and return the queryset."""
    from datetime import date
    from calendar import monthrange

    invoices = Invoice.objects.exclude(status='cancelled')

    search = request.POST.get('search', '')
    if search:
        invoices = invoices.filter(
            Q(invoice_number__icontains=search) |
            Q(title__icontains=search) |
            Q(client__name__icontains=search)
        )

    status = request.POST.get('status', '')
    if status == 'paid_partial':
        invoices = invoices.filter(status__in=['paid', 'partial'])
    elif status:
        invoices = invoices.filter(status=status)

    gst_filing = request.POST.get('gst_filing', '')
    if gst_filing:
        invoices = invoices.filter(gst_filing_status=gst_filing)

    from_date = request.POST.get('from_date', '')
    to_date = request.POST.get('to_date', '')
    period = request.POST.get('period', '')
    if period in ('prev_month', 'this_month') and not from_date and not to_date:
        today = date.today()
        if period == 'prev_month':
            year = today.year if today.month > 1 else today.year - 1
            month = today.month - 1 if today.month > 1 else 12
        else:
            year, month = today.year, today.month
        from_date = date(year, month, 1).isoformat()
        to_date = date(year, month, monthrange(year, month)[1]).isoformat()
    if from_date:
        invoices = invoices.filter(issue_date__gte=from_date)
    if to_date:
        invoices = invoices.filter(issue_date__lte=to_date)

    return invoices


@login_required
def invoices_mark_gst_filed(request):
    """Mark every invoice matching the currently-applied filters as GST-filed.

    Reads the same query params as invoice_list, applies them, and updates the
    matching rows. Intended to be hit after filtering down to a specific
    GST return period (e.g. issue_date in May 2026, status != cancelled).
    """
    if request.method != 'POST':
        return redirect('invoice_list')

    invoices = _filter_invoices_from_post(request)
    candidates = invoices.exclude(gst_filing_status='not_applicable').filter(tax_amount__gt=0)
    updated = candidates.update(gst_filing_status='filed', gst_filed_at=timezone.now())

    if updated:
        messages.success(request, f'Marked {updated} invoice(s) as GST-filed.')
    else:
        messages.info(request, 'No matching invoices to mark as filed.')

    qs = request.POST.urlencode()
    return redirect(f'{reverse("invoice_list")}?{qs}')


@login_required
def invoices_mark_gst_pending(request):
    """Reverse of invoices_mark_gst_filed: flip filtered invoices back to pending."""
    if request.method != 'POST':
        return redirect('invoice_list')

    invoices = _filter_invoices_from_post(request)
    candidates = invoices.exclude(gst_filing_status='not_applicable').filter(tax_amount__gt=0)
    updated = candidates.update(gst_filing_status='pending', gst_filed_at=None)

    if updated:
        messages.success(request, f'Marked {updated} invoice(s) as GST pending.')
    else:
        messages.info(request, 'No matching invoices to revert.')

    qs = request.POST.urlencode()
    return redirect(f'{reverse("invoice_list")}?{qs}')


@login_required
def invoice_set_gst_status(request, pk):
    """Per-invoice GST filing status update (POST only)."""
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method != 'POST':
        return redirect('invoice_detail', pk=pk)

    new_status = request.POST.get('gst_filing_status', '')
    valid = {key for key, _ in Invoice.GST_FILING_STATUS_CHOICES}
    if new_status not in valid:
        messages.error(request, 'Invalid GST status.')
        return redirect('invoice_detail', pk=pk)

    invoice.gst_filing_status = new_status
    invoice.gst_filed_at = timezone.now() if new_status == 'filed' else None
    invoice.save(update_fields=['gst_filing_status', 'gst_filed_at'])
    messages.success(request, f'GST status set to "{invoice.get_gst_filing_status_display()}".')
    return redirect('invoice_detail', pk=pk)


@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(
        Invoice.objects.select_related('client', 'project').prefetch_related('items', 'payments'),
        pk=pk
    )

    # Always recalculate totals to ensure they match current items and tax rate
    if invoice.items.exists():
        invoice.calculate_totals()

    return render(request, 'invoices/detail.html', {'invoice': invoice})


@login_required
def invoice_create(request):
    clients = Client.objects.filter(is_active=True)
    projects = Project.objects.select_related('client').all()
    quotes = Quote.objects.filter(status='accepted')

    if request.method == 'POST':
        from decimal import Decimal, InvalidOperation

        # Safe tax_rate conversion - default to 0 if empty or invalid
        try:
            tax_rate_val = request.POST.get('tax_rate', '0')
            tax_rate = Decimal(tax_rate_val) if tax_rate_val else Decimal('0')
        except (InvalidOperation, ValueError):
            tax_rate = Decimal('0')

        # Safe discount conversion
        try:
            discount_val = request.POST.get('discount', '0')
            discount = Decimal(discount_val) if discount_val else Decimal('0')
        except (InvalidOperation, ValueError):
            discount = Decimal('0')

        create_kwargs = dict(
            client_id=request.POST.get('client'),
            project_id=request.POST.get('project') or None,
            quote_id=request.POST.get('quote') or None,
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            issue_date=request.POST.get('issue_date') or timezone.now().date(),
            due_date=request.POST.get('due_date') or None,
            status=request.POST.get('status', 'draft'),
            discount=discount,
            tax_rate=tax_rate,
            notes=request.POST.get('notes', ''),
            client_notes=request.POST.get('client_notes', ''),
            terms=request.POST.get('terms', ''),
        )

        # Allow manual invoice number override
        manual_invoice_number = request.POST.get('invoice_number', '').strip()
        if manual_invoice_number:
            if Invoice.objects.filter(invoice_number=manual_invoice_number).exists():
                messages.error(request, f'Invoice number "{manual_invoice_number}" is already in use. Please choose a different number or leave blank to auto-generate.')
                return redirect('invoice_create')
            create_kwargs['invoice_number'] = manual_invoice_number

        invoice = Invoice.objects.create(**create_kwargs)

        # Process line items
        item_count = int(request.POST.get('item_count', 0))
        for i in range(1, item_count + 10):  # Check a few extra indices in case of gaps
            description = request.POST.get(f'item_description_{i}')
            if description:
                quantity = Decimal(request.POST.get(f'item_quantity_{i}', 1) or 1)
                unit_price = Decimal(request.POST.get(f'item_price_{i}', 0) or 0)
                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=description,
                    quantity=quantity,
                    unit_price=unit_price
                )

        # Recalculate totals
        invoice.calculate_totals()

        messages.success(request, f'Invoice "{invoice.invoice_number}" created successfully.')
        return redirect('invoice_detail', pk=invoice.pk)

    # Get today's date and default due date (15 days from now)
    today = timezone.now().date()
    due_date_default = today + timezone.timedelta(days=15)

    # Get company settings for default terms
    company = CompanySettings.get_settings()

    return render(request, 'invoices/form.html', {
        'clients': clients,
        'projects': projects,
        'quotes': quotes,
        'company': company,
        'form_title': 'Create New Invoice',
        'status_choices': Invoice.STATUS_CHOICES,
        'today': today.isoformat(),
        'due_date_default': due_date_default.isoformat(),
    })


@login_required
def invoice_update(request, pk):
    invoice = get_object_or_404(Invoice.objects.prefetch_related('items'), pk=pk)
    clients = Client.objects.filter(is_active=True)
    projects = Project.objects.select_related('client').all()
    quotes = Quote.objects.filter(status='accepted')

    if request.method == 'POST':
        from decimal import Decimal, InvalidOperation

        invoice.client_id = request.POST.get('client')
        invoice.project_id = request.POST.get('project') or None
        invoice.quote_id = request.POST.get('quote') or None
        invoice.title = request.POST.get('title')
        invoice.description = request.POST.get('description', '')
        invoice.issue_date = request.POST.get('issue_date') or timezone.now().date()
        invoice.due_date = request.POST.get('due_date') or None
        invoice.status = request.POST.get('status', 'draft')

        # Safe decimal conversion - default to 0 if empty
        try:
            discount_val = request.POST.get('discount', '0')
            invoice.discount = Decimal(discount_val) if discount_val else Decimal('0')
        except (InvalidOperation, ValueError):
            invoice.discount = Decimal('0')

        try:
            tax_rate_val = request.POST.get('tax_rate', '0')
            invoice.tax_rate = Decimal(tax_rate_val) if tax_rate_val else Decimal('0')
        except (InvalidOperation, ValueError):
            invoice.tax_rate = Decimal('0')

        invoice.notes = request.POST.get('notes', '')
        invoice.client_notes = request.POST.get('client_notes', '')
        invoice.terms = request.POST.get('terms', '')

        # Allow manual invoice number override
        manual_invoice_number = request.POST.get('invoice_number', '').strip()
        if manual_invoice_number and manual_invoice_number != invoice.invoice_number:
            if Invoice.objects.filter(invoice_number=manual_invoice_number).exclude(pk=invoice.pk).exists():
                messages.error(request, f'Invoice number "{manual_invoice_number}" is already in use. Please choose a different number.')
                return redirect('invoice_update', pk=invoice.pk)
            invoice.invoice_number = manual_invoice_number

        invoice.save()

        # Delete existing items and recreate
        invoice.items.all().delete()

        # Process line items
        try:
            item_count = int(request.POST.get('item_count', 0) or 0)
        except ValueError:
            item_count = 0

        for i in range(1, item_count + 10):  # Check a few extra indices in case of gaps
            description = request.POST.get(f'item_description_{i}')
            if description:
                try:
                    qty_val = request.POST.get(f'item_quantity_{i}', '1') or '1'
                    quantity = Decimal(qty_val)
                except (InvalidOperation, ValueError):
                    quantity = Decimal('1')

                try:
                    price_val = request.POST.get(f'item_price_{i}', '0') or '0'
                    unit_price = Decimal(price_val)
                except (InvalidOperation, ValueError):
                    unit_price = Decimal('0')

                InvoiceItem.objects.create(
                    invoice=invoice,
                    description=description,
                    quantity=quantity,
                    unit_price=unit_price
                )

        # Recalculate totals
        invoice.calculate_totals()

        messages.success(request, f'Invoice "{invoice.invoice_number}" updated successfully.')
        return redirect('invoice_detail', pk=invoice.pk)

    # Get today's date and default due date (15 days from now) for form defaults
    today = timezone.now().date()
    due_date_default = today + timezone.timedelta(days=15)

    # Get company settings for default terms
    company = CompanySettings.get_settings()

    return render(request, 'invoices/form.html', {
        'invoice': invoice,
        'clients': clients,
        'projects': projects,
        'quotes': quotes,
        'company': company,
        'form_title': 'Edit Invoice',
        'status_choices': Invoice.STATUS_CHOICES,
        'today': today.isoformat(),
        'due_date_default': due_date_default.isoformat(),
    })


@login_required
def invoice_update_number(request, pk):
    """Quick inline update for invoice number from detail page."""
    invoice = get_object_or_404(Invoice, pk=pk)
    if request.method == 'POST':
        new_number = request.POST.get('invoice_number', '').strip()
        if new_number and new_number != invoice.invoice_number:
            # Check uniqueness
            if Invoice.objects.filter(invoice_number=new_number).exclude(pk=pk).exists():
                messages.error(request, f'Invoice number "{new_number}" is already in use.')
            else:
                invoice.invoice_number = new_number
                invoice.save()
                messages.success(request, f'Invoice number updated to "{new_number}".')
    return redirect('invoice_detail', pk=pk)


@login_required
def invoice_pdf(request, pk):
    """Generate PDF for an invoice"""
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal

    invoice = get_object_or_404(
        Invoice.objects.select_related('client', 'project').prefetch_related('items', 'payments'),
        pk=pk
    )

    # Get company settings
    company = CompanySettings.get_settings()

    # Check if GST should be included
    with_gst = request.GET.get('gst', '0') == '1'
    download = request.GET.get('download', '0') == '1'

    # Calculate amounts
    taxable_amount = invoice.subtotal - (invoice.discount or Decimal('0'))
    # Use invoice's tax_rate (0 means no tax), only default to 0 if None
    tax_rate = Decimal(str(invoice.tax_rate)) if invoice.tax_rate is not None else Decimal('0')

    cgst_amount = Decimal('0')
    sgst_amount = Decimal('0')
    tax_amount = Decimal('0')
    total = taxable_amount

    if with_gst:
        cgst_rate = tax_rate / 2
        sgst_rate = tax_rate / 2
        cgst_amount = taxable_amount * (cgst_rate / 100)
        sgst_amount = taxable_amount * (sgst_rate / 100)
        tax_amount = cgst_amount + sgst_amount
        total = taxable_amount + tax_amount

    # Calculate balance due
    balance_due = total - (invoice.amount_paid or Decimal('0'))

    context = {
        'invoice': invoice,
        'company': company,
        'with_gst': with_gst,
        'taxable_amount': taxable_amount,
        'tax_rate': tax_rate,
        'cgst_rate': tax_rate / 2 if with_gst else 0,
        'sgst_rate': tax_rate / 2 if with_gst else 0,
        'cgst_amount': cgst_amount,
        'sgst_amount': sgst_amount,
        'tax_amount': tax_amount,
        'total_with_gst': total,
        'balance_due': balance_due,
    }

    # If download requested, generate PDF
    if download:
        try:
            from weasyprint import HTML, CSS

            html_string = render_to_string('invoices/pdf.html', context)

            # Create PDF
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf = html.write_pdf()

            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'invoices/pdf.html', context)


@login_required
def invoices_backup_pdf(request):
    """Render every invoice in the system into a single archival PDF.

    Intended as an offline backup before a Start-New-FY reset that
    wipes invoices/payments/expenses.
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal

    invoices = (
        Invoice.objects
        .select_related('client', 'project')
        .prefetch_related('items', 'payments')
        .order_by('issue_date', 'invoice_number')
    )

    agg = invoices.aggregate(
        total_billed_pre_tax=Sum('subtotal') - Sum('discount'),
        total_tax=Sum('tax_amount'),
        total_with_tax=Sum('total_amount'),
        total_paid=Sum('amount_paid'),
    )

    total_with_tax = agg['total_with_tax'] or Decimal('0')
    total_paid = agg['total_paid'] or Decimal('0')

    if invoices.exists():
        first_date = invoices.order_by('issue_date').first().issue_date
        last_date = invoices.order_by('-issue_date').first().issue_date
    else:
        first_date = last_date = None

    stats = {
        'count': invoices.count(),
        'first_date': first_date,
        'last_date': last_date,
        'total_billed': agg['total_billed_pre_tax'] or Decimal('0'),
        'total_tax': agg['total_tax'] or Decimal('0'),
        'total_with_tax': total_with_tax,
        'total_paid': total_paid,
        'total_outstanding': total_with_tax - total_paid,
    }

    context = {
        'invoices': invoices,
        'company': CompanySettings.get_settings(),
        'stats': stats,
        'generated_on': timezone.now(),
    }

    download = request.GET.get('download', '0') == '1'
    if download:
        try:
            from weasyprint import HTML
            html_string = render_to_string('invoices/backup_pdf.html', context)
            pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
            filename = f"invoice_archive_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'invoices/backup_pdf.html', context)


@login_required
def expenses_backup_pdf(request):
    """Render every expense in the system into a single archival PDF.

    Intended as an offline backup before a Start-New-FY reset that
    wipes expenses (along with invoices/payments).
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string
    from decimal import Decimal

    expenses = (
        Expense.objects
        .select_related('project', 'project__client')
        .order_by('date', 'created_at')
    )

    agg = expenses.aggregate(total=Sum('amount'))
    total_amount = agg['total'] or Decimal('0')

    by_category_qs = (
        expenses.values('category')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    cat_labels = dict(Expense.CATEGORY_CHOICES)
    by_category = [
        {
            'label': cat_labels.get(row['category'], row['category']),
            'count': row['count'],
            'total': row['total'] or Decimal('0'),
        }
        for row in by_category_qs
    ]

    by_method_qs = (
        expenses.values('payment_method')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    method_labels = dict(Expense.PAYMENT_METHOD_CHOICES)
    by_method = [
        {
            'label': method_labels.get(row['payment_method'], row['payment_method']),
            'count': row['count'],
            'total': row['total'] or Decimal('0'),
        }
        for row in by_method_qs
    ]

    if expenses.exists():
        first_date = expenses.order_by('date').first().date
        last_date = expenses.order_by('-date').first().date
    else:
        first_date = last_date = None

    stats = {
        'count': expenses.count(),
        'first_date': first_date,
        'last_date': last_date,
        'total_amount': total_amount,
    }

    context = {
        'expenses': expenses,
        'company': CompanySettings.get_settings(),
        'stats': stats,
        'by_category': by_category,
        'by_method': by_method,
        'generated_on': timezone.now(),
    }

    download = request.GET.get('download', '0') == '1'
    if download:
        try:
            from weasyprint import HTML
            html_string = render_to_string('expenses/backup_pdf.html', context)
            pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
            filename = f"expense_archive_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'expenses/backup_pdf.html', context)


@login_required
def clients_credentials_backup_pdf(request):
    """Render every client and their project credentials (including plaintext passwords)
    into a single offline-readable PDF.

    Sensitive: this PDF contains plaintext vault passwords. Login-only; admin access required.
    """
    from django.http import HttpResponse
    from django.template.loader import render_to_string

    clients = (
        Client.objects
        .prefetch_related('projects', 'projects__credentials')
        .order_by('company_name', 'name')
    )

    total_credentials = Credential.objects.count()
    total_active_credentials = Credential.objects.filter(is_active=True).count()

    stats = {
        'client_count': clients.count(),
        'project_count': Project.objects.count(),
        'credential_count': total_credentials,
        'active_credentials': total_active_credentials,
    }

    context = {
        'clients': clients,
        'company': CompanySettings.get_settings(),
        'stats': stats,
        'generated_on': timezone.now(),
    }

    download = request.GET.get('download', '0') == '1'
    if download:
        try:
            from weasyprint import HTML
            html_string = render_to_string('clients/backup_pdf.html', context)
            pdf = HTML(string=html_string, base_url=request.build_absolute_uri('/')).write_pdf()
            filename = f"clients_credentials_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"
            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'clients/backup_pdf.html', context)


@login_required
def clients_credentials_backup_xlsx(request):
    """Excel workbook with two sheets: Clients (every field) and Credentials
    (every field plus joined client/project context). Includes plaintext passwords.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from django.http import HttpResponse

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')

    # Sheet 1: Clients (all model fields)
    ws_clients = wb.active
    ws_clients.title = 'Clients'
    client_headers = [
        'ID', 'Name', 'Company Name', 'Email', 'Phone', 'WhatsApp', 'Address',
        'GST Number', 'Priority', 'Active', 'Notes', 'Created At', 'Updated At',
        'Google Client ID (Desktop)', 'Google Client ID (iOS)', 'Google Client ID (Android)',
        'Google Reversed Client ID', 'Google Client Secret',
        'RetailEase Drive Backup', 'RetailEase Server Backup', 'RetailEase Local Backup',
        'RetailEase Min Version', 'RetailEase Latest Version', 'RetailEase Update URL',
        'RetailEase Force Update', 'RetailEase Maintenance Mode', 'RetailEase Maintenance Message',
        'Support Email', 'Support Phone', 'Support WhatsApp',
    ]
    ws_clients.append(client_headers)
    for cell in ws_clients[1]:
        cell.font = header_font
        cell.fill = header_fill

    for c in Client.objects.all().order_by('company_name', 'name'):
        ws_clients.append([
            str(c.id), c.name, c.company_name, c.email, c.phone, c.whatsapp, c.address,
            c.gst_number, c.get_priority_display(), 'Yes' if c.is_active else 'No', c.notes,
            c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            c.updated_at.strftime('%Y-%m-%d %H:%M') if c.updated_at else '',
            c.google_client_id, c.google_client_id_ios, c.google_client_id_android,
            c.google_reversed_client_id, c.google_client_secret,
            'Yes' if c.retailease_google_drive_enabled else 'No',
            'Yes' if c.retailease_server_backup_enabled else 'No',
            'Yes' if c.retailease_local_backup_enabled else 'No',
            c.retailease_min_version, c.retailease_latest_version, c.retailease_update_url,
            'Yes' if c.retailease_force_update else 'No',
            'Yes' if c.retailease_maintenance_mode else 'No',
            c.retailease_maintenance_message,
            c.retailease_support_email, c.retailease_support_phone, c.retailease_support_whatsapp,
        ])

    # Sheet 2: Credentials (vault, includes plaintext passwords)
    ws_cred = wb.create_sheet('Credentials')
    cred_headers = [
        'ID', 'Client', 'Project', 'Type', 'Name', 'Provider', 'URL', 'IP Address',
        'Username', 'Password', 'SSH Key', 'Port',
        'Purchase Date', 'Expiry Date', 'Last Renewed', 'Auto Renew', 'Renewal Cost',
        'Active', 'Client Visible', 'Notes', 'Created At', 'Updated At',
    ]
    ws_cred.append(cred_headers)
    for cell in ws_cred[1]:
        cell.font = header_font
        cell.fill = header_fill

    creds = (
        Credential.objects
        .select_related('project', 'project__client')
        .order_by('project__client__company_name', 'project__name', 'name')
    )
    for cr in creds:
        client = cr.project.client if cr.project else None
        ws_cred.append([
            str(cr.id),
            (client.company_name or client.name) if client else '',
            cr.project.name if cr.project else '',
            cr.get_credential_type_display(),
            cr.name, cr.provider, cr.url, str(cr.ip_address) if cr.ip_address else '',
            cr.username, cr.password, cr.ssh_key, cr.port if cr.port else '',
            cr.purchase_date.strftime('%Y-%m-%d') if cr.purchase_date else '',
            cr.expiry_date.strftime('%Y-%m-%d') if cr.expiry_date else '',
            cr.last_renewed_date.strftime('%Y-%m-%d') if cr.last_renewed_date else '',
            'Yes' if cr.auto_renew else 'No',
            float(cr.renewal_cost) if cr.renewal_cost is not None else '',
            'Yes' if cr.is_active else 'No',
            'Yes' if cr.client_visible else 'No',
            cr.notes,
            cr.created_at.strftime('%Y-%m-%d %H:%M') if cr.created_at else '',
            cr.updated_at.strftime('%Y-%m-%d %H:%M') if cr.updated_at else '',
        ])

    for ws in (ws_clients, ws_cred):
        for col_cells in ws.columns:
            length = max((len(str(cell.value or '')) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 50)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = (
        f'attachment; filename="clients_credentials_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    )
    wb.save(response)
    return response


# ============== Invoice Delete ==============

@login_required
def invoice_delete(request, pk):
    invoice = get_object_or_404(Invoice, pk=pk)

    if request.method == 'POST':
        invoice_number = invoice.invoice_number
        # Check if invoice has payments
        payment_count = invoice.payments.count()

        if payment_count > 0:
            messages.error(
                request,
                f'Cannot delete "{invoice_number}". It has {payment_count} payment(s) recorded against it.'
            )
            return redirect('invoice_detail', pk=pk)

        invoice.delete()
        messages.success(request, f'Invoice "{invoice_number}" deleted successfully.')
        return redirect('invoice_list')

    return render(request, 'invoices/delete.html', {'invoice': invoice})


# ============== Invoice Clone ==============

@login_required
def invoice_clone(request, pk):
    """Clone an existing invoice"""
    original_invoice = get_object_or_404(
        Invoice.objects.prefetch_related('items'),
        pk=pk
    )

    from decimal import Decimal

    today = timezone.now().date()
    due_date = today + timedelta(days=15)

    new_invoice = Invoice.objects.create(
        client=original_invoice.client,
        project=original_invoice.project,
        title=f"Copy of {original_invoice.title}",
        description=original_invoice.description,
        status='draft',
        subtotal=original_invoice.subtotal,
        discount=original_invoice.discount,
        tax_rate=original_invoice.tax_rate,
        tax_amount=original_invoice.tax_amount,
        total_amount=original_invoice.total_amount,
        issue_date=today,
        due_date=due_date,
        terms=original_invoice.terms,
        client_notes=original_invoice.client_notes,
        notes=original_invoice.notes,
    )

    # Clone all items
    for item in original_invoice.items.all():
        InvoiceItem.objects.create(
            invoice=new_invoice,
            description=item.description,
            details=item.details,
            quantity=item.quantity,
            unit_price=item.unit_price,
            total=item.total,
            order=item.order,
        )

    messages.success(request, f'Invoice cloned successfully. New invoice: {new_invoice.invoice_number}')
    return redirect('invoice_update', pk=new_invoice.pk)


@login_required
def invoice_split_to_milestones(request, pk):
    """Split an existing invoice into N milestone invoices, each a draft with its own number."""
    if request.method != 'POST':
        return redirect('invoice_detail', pk=pk)

    from decimal import Decimal, InvalidOperation
    source = get_object_or_404(Invoice, pk=pk)

    labels = request.POST.getlist('milestone_label')
    percents = request.POST.getlist('milestone_percent')
    cancel_source = request.POST.get('cancel_source') == 'on'

    rows = [(l.strip() or f'Milestone {i+1}', p.strip()) for i, (l, p) in enumerate(zip(labels, percents)) if p.strip()]
    if len(rows) < 2:
        messages.error(request, 'Add at least 2 milestones to split.')
        return redirect('invoice_detail', pk=pk)

    parsed = []
    total_pct = Decimal('0')
    for label, raw_pct in rows:
        try:
            pct = Decimal(raw_pct)
        except (InvalidOperation, ValueError):
            messages.error(request, f'Invalid percentage "{raw_pct}".')
            return redirect('invoice_detail', pk=pk)
        if pct <= 0:
            messages.error(request, 'Each milestone percentage must be greater than 0.')
            return redirect('invoice_detail', pk=pk)
        parsed.append((label, pct))
        total_pct += pct

    if total_pct != Decimal('100'):
        messages.error(request, f'Milestone percentages must sum to 100% (currently {total_pct}%).')
        return redirect('invoice_detail', pk=pk)

    today = timezone.now().date()
    base_amount = source.subtotal or (source.total_amount - (source.tax_amount or Decimal('0')))
    created = []
    for label, pct in parsed:
        milestone_amount = (base_amount * pct / Decimal('100')).quantize(Decimal('0.01'))
        new_invoice = Invoice.objects.create(
            client=source.client,
            project=source.project,
            quote=source.quote,
            title=f'{source.title} — {label}',
            description=source.description,
            status='draft',
            tax_rate=source.tax_rate,
            issue_date=today,
            due_date=today + timedelta(days=15),
            terms=source.terms,
            client_notes=source.client_notes,
        )
        InvoiceItem.objects.create(
            invoice=new_invoice,
            description=label,
            quantity=Decimal('1'),
            unit_price=milestone_amount,
        )
        new_invoice.calculate_totals()
        created.append(new_invoice.invoice_number)

    if cancel_source:
        source.status = 'cancelled'
        source.save(update_fields=['status'])

    cancel_msg = ' Source invoice was cancelled.' if cancel_source else ''
    messages.success(request, f'Created {len(created)} milestone invoices: {", ".join(created)}.{cancel_msg}')
    return redirect('invoice_list')


# ============== Payments ==============

@login_required
def payment_list(request):
    import json
    from dateutil.relativedelta import relativedelta

    payments = Payment.objects.select_related('invoice', 'invoice__client').all()

    search = request.GET.get('search', '')
    if search:
        payments = payments.filter(
            Q(invoice__invoice_number__icontains=search) |
            Q(transaction_id__icontains=search) |
            Q(invoice__client__name__icontains=search)
        )

    method = request.GET.get('method', '')
    if method:
        payments = payments.filter(payment_method=method)

    # ============== Stats & Chart Data ==============
    today = timezone.now().date()

    # Total payments
    total_payments = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0

    # This month
    first_day = today.replace(day=1)
    this_month = Payment.objects.filter(
        payment_date__gte=first_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Payment count
    payment_count = Payment.objects.count()

    # Monthly payments (last 6 months)
    monthly_labels = []
    monthly_data = []
    for i in range(5, -1, -1):
        month_date = today - relativedelta(months=i)
        month_start = month_date.replace(day=1)
        if i > 0:
            month_end = (month_date + relativedelta(months=1)).replace(day=1) - timedelta(days=1)
        else:
            month_end = today

        month_total = Payment.objects.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_labels.append(month_date.strftime('%b'))
        monthly_data.append(float(month_total))

    # Payment method distribution
    method_data = {}
    for method_code, method_label in Payment.METHOD_CHOICES:
        total = Payment.objects.filter(payment_method=method_code).aggregate(
            total=Sum('amount')
        )['total'] or 0
        if total > 0:
            method_data[method_label] = float(total)

    context = {
        'payments': payments.order_by('-payment_date'),
        'search': search,
        'method': method,
        'method_choices': Payment.METHOD_CHOICES,
        'total_payments': total_payments,
        'this_month': this_month,
        'payment_count': payment_count,
        'monthly_labels': json.dumps(monthly_labels),
        'monthly_data': json.dumps(monthly_data),
        'method_labels': json.dumps(list(method_data.keys())),
        'method_data': json.dumps(list(method_data.values())),
    }
    return render(request, 'payments/list.html', context)


@login_required
def payment_create(request):
    invoices = Invoice.objects.exclude(status__in=['paid', 'cancelled']).select_related('client')

    if request.method == 'POST':
        payment = Payment.objects.create(
            invoice_id=request.POST.get('invoice'),
            amount=request.POST.get('amount'),
            payment_date=request.POST.get('payment_date') or timezone.now().date(),
            payment_method=request.POST.get('payment_method', 'bank_transfer'),
            transaction_id=request.POST.get('transaction_id', ''),
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'Payment of ₹{payment.amount} recorded successfully.')
        return redirect('invoice_detail', pk=payment.invoice.pk)

    return render(request, 'payments/form.html', {
        'invoices': invoices,
        'form_title': 'Record New Payment',
        'method_choices': Payment.METHOD_CHOICES,
    })


@login_required
def payment_edit(request, pk):
    """Edit an existing payment. Recomputes invoice.amount_paid via Payment.save()."""
    payment = get_object_or_404(Payment.objects.select_related('invoice', 'invoice__client'), pk=pk)

    if request.method == 'POST':
        try:
            payment.amount = request.POST.get('amount') or payment.amount
            payment.payment_date = request.POST.get('payment_date') or payment.payment_date
            payment.payment_method = request.POST.get('payment_method', payment.payment_method)
            payment.transaction_id = request.POST.get('transaction_id', '')
            payment.notes = request.POST.get('notes', '')
            payment.save()
            messages.success(request, f'Payment of ₹{payment.amount} updated.')
            return redirect('invoice_detail', pk=payment.invoice.pk)
        except Exception as e:
            messages.error(request, f'Could not update payment: {e}')

    # For edit, restrict the invoice dropdown to the one this payment belongs to
    # so the user can't accidentally re-link it (which would break amount_paid on
    # both invoices). Reassigning a payment to a different invoice is out of scope.
    invoices = Invoice.objects.filter(pk=payment.invoice.pk).select_related('client')

    return render(request, 'payments/form.html', {
        'invoices': invoices,
        'payment': payment,
        'form_title': f'Edit Payment for {payment.invoice.invoice_number}',
        'method_choices': Payment.METHOD_CHOICES,
        'is_edit': True,
        'selected_invoice': payment.invoice.pk,
    })


@login_required
def payment_delete(request, pk):
    """Delete a payment. Recomputes invoice.amount_paid via Payment.delete()."""
    payment = get_object_or_404(Payment.objects.select_related('invoice'), pk=pk)
    invoice_pk = payment.invoice.pk
    invoice_number = payment.invoice.invoice_number
    amount = payment.amount

    if request.method == 'POST':
        payment.delete()
        messages.success(
            request,
            f'Payment of ₹{amount} on invoice {invoice_number} deleted. '
            'Invoice balance and status updated.'
        )
        return redirect('invoice_detail', pk=invoice_pk)

    return render(request, 'payments/delete.html', {'payment': payment})


@login_required
def payment_receipt(request, pk):
    """Generate receipt for a payment"""
    payment = get_object_or_404(
        Payment.objects.select_related('invoice', 'invoice__client'),
        pk=pk
    )

    # Get company settings
    company = CompanySettings.get_settings()

    # Generate receipt number based on payment
    receipt_number = f"REC{payment.payment_date.strftime('%Y%m%d')}{str(payment.pk)[:8].upper()}"

    context = {
        'payment': payment,
        'invoice': payment.invoice,
        'client': payment.invoice.client,
        'company': company,
        'receipt_number': receipt_number,
    }

    download = request.GET.get('download', '0') == '1'

    # If download requested, generate PDF
    if download:
        try:
            from weasyprint import HTML
            from django.template.loader import render_to_string

            html_string = render_to_string('payments/receipt.html', context)
            html = HTML(string=html_string, base_url=request.build_absolute_uri('/'))
            pdf = html.write_pdf()

            response = HttpResponse(pdf, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="receipt_{receipt_number}.pdf"'
            return response
        except ImportError:
            messages.warning(request, 'PDF generation requires WeasyPrint. Showing printable view instead.')

    return render(request, 'payments/receipt.html', context)


# ============== Settings & Reports ==============

@login_required
def settings_view(request):
    company = CompanySettings.get_settings()

    if request.method == 'POST':
        company.company_name = request.POST.get('company_name', '')
        company.tagline = request.POST.get('tagline', '')
        company.email = request.POST.get('email', '')
        company.phone = request.POST.get('phone', '')
        company.address = request.POST.get('address', '')
        company.gst_number = request.POST.get('gst_number', '')
        company.pan_number = request.POST.get('pan_number', '')
        company.hsn_code = request.POST.get('hsn_code', '')
        company.bank_name = request.POST.get('bank_name', '')
        company.bank_account_number = request.POST.get('bank_account_number', '')
        company.bank_ifsc = request.POST.get('bank_ifsc', '')
        company.bank_branch = request.POST.get('bank_branch', '')
        company.upi_id = request.POST.get('upi_id', '')

        # Safe default_tax_rate conversion - default to 0 if empty
        from decimal import Decimal, InvalidOperation
        try:
            tax_rate_val = request.POST.get('default_tax_rate', '0')
            company.default_tax_rate = Decimal(tax_rate_val) if tax_rate_val else Decimal('0')
        except (InvalidOperation, ValueError):
            company.default_tax_rate = Decimal('0')

        # Quote validity days
        try:
            validity_val = request.POST.get('default_quote_validity_days', '30')
            company.default_quote_validity_days = int(validity_val) if validity_val else 30
        except (ValueError, TypeError):
            company.default_quote_validity_days = 30

        company.invoice_terms = request.POST.get('invoice_terms', '')
        company.quote_terms = request.POST.get('quote_terms', '')
        company.default_payment_terms = request.POST.get('default_payment_terms', '50-50')
        company.invoice_prefix = request.POST.get('invoice_prefix', 'INVRT')
        company.quote_prefix = request.POST.get('quote_prefix', 'QT')

        # Invoice starting number
        try:
            starting_num = request.POST.get('invoice_starting_number', '201')
            company.invoice_starting_number = int(starting_num) if starting_num else 201
        except (ValueError, TypeError):
            company.invoice_starting_number = 201

        # Quote starting number
        try:
            q_starting_num = request.POST.get('quote_starting_number', '1')
            company.quote_starting_number = int(q_starting_num) if q_starting_num else 1
        except (ValueError, TypeError):
            company.quote_starting_number = 1

        # InterioDesk App Settings
        company.interiodesk_min_version = request.POST.get('interiodesk_min_version', '1.0.0')
        company.interiodesk_latest_version = request.POST.get('interiodesk_latest_version', '1.0.0')
        company.interiodesk_update_url_macos = request.POST.get('interiodesk_update_url_macos', '')
        company.interiodesk_update_url_windows = request.POST.get('interiodesk_update_url_windows', '')
        if request.FILES.get('interiodesk_file_macos'):
            if company.interiodesk_file_macos:
                company.interiodesk_file_macos.delete(save=False)
            company.interiodesk_file_macos = request.FILES['interiodesk_file_macos']
        if request.FILES.get('interiodesk_file_windows'):
            if company.interiodesk_file_windows:
                company.interiodesk_file_windows.delete(save=False)
            company.interiodesk_file_windows = request.FILES['interiodesk_file_windows']
        company.interiodesk_force_update = 'interiodesk_force_update' in request.POST
        company.interiodesk_maintenance_mode = 'interiodesk_maintenance_mode' in request.POST
        company.interiodesk_maintenance_message = request.POST.get('interiodesk_maintenance_message', '')
        company.interiodesk_release_notes = request.POST.get('interiodesk_release_notes', '')

        if request.FILES.get('logo'):
            company.logo = request.FILES.get('logo')

        company.save()
        messages.success(request, 'Settings updated successfully.')
        return redirect('settings')

    opening_balance = OpeningBalance.current()
    opening_balance_history = OpeningBalance.objects.all()[:10]
    from django.db.models import F
    outstanding_receivables = Invoice.objects.exclude(
        status__in=['paid', 'cancelled']
    ).aggregate(total=Sum(F('total_amount') - F('amount_paid')))['total'] or 0
    fy_reset_history = FYResetEvent.objects.select_related('ran_by', 'opening_balance')[:10]
    return render(request, 'settings/index.html', {
        'company': company,
        'opening_balance': opening_balance,
        'opening_balance_history': opening_balance_history,
        'outstanding_receivables': outstanding_receivables,
        'fy_reset_history': fy_reset_history,
    })


@login_required
def opening_balance_save(request):
    """Create a new opening balance snapshot (preserves history)."""
    from decimal import Decimal, InvalidOperation

    if request.method != 'POST':
        return redirect('settings')

    label = (request.POST.get('label') or '').strip()
    as_of_raw = (request.POST.get('as_of_date') or '').strip()
    cash_hand_raw = (request.POST.get('cash_in_hand') or '0').strip()
    cash_acc_raw = (request.POST.get('cash_in_account') or '0').strip()
    receivable_raw = (request.POST.get('accounts_receivable') or '0').strip()
    notes = (request.POST.get('notes') or '').strip()

    if not label:
        messages.error(request, 'Label is required (e.g. "FY 2026-27").')
        return redirect('settings')
    if not as_of_raw:
        messages.error(request, 'As-of date is required.')
        return redirect('settings')

    try:
        as_of_date = datetime.strptime(as_of_raw, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, 'As-of date must be a valid date.')
        return redirect('settings')

    try:
        cash_hand = Decimal(cash_hand_raw or '0')
        cash_acc = Decimal(cash_acc_raw or '0')
        receivable = Decimal(receivable_raw or '0')
    except InvalidOperation:
        messages.error(request, 'Cash and receivable amounts must be valid numbers.')
        return redirect('settings')

    if cash_hand < 0 or cash_acc < 0 or receivable < 0:
        messages.error(request, 'Cash and receivable amounts cannot be negative.')
        return redirect('settings')

    OpeningBalance.objects.create(
        label=label,
        as_of_date=as_of_date,
        cash_in_hand=cash_hand,
        cash_in_account=cash_acc,
        accounts_receivable=receivable,
        notes=notes,
    )
    messages.success(request, f'Opening balance "{label}" saved.')
    return redirect('settings')


@login_required
def opening_balance_edit(request, pk):
    """Edit an existing OpeningBalance row (correct cash, receivables, label, date, notes)."""
    from decimal import Decimal, InvalidOperation

    ob = get_object_or_404(OpeningBalance, pk=pk)

    if request.method == 'POST':
        label = (request.POST.get('label') or '').strip()
        as_of_raw = (request.POST.get('as_of_date') or '').strip()
        cash_hand_raw = (request.POST.get('cash_in_hand') or '0').strip()
        cash_acc_raw = (request.POST.get('cash_in_account') or '0').strip()
        receivable_raw = (request.POST.get('accounts_receivable') or '0').strip()
        notes = (request.POST.get('notes') or '').strip()

        if not label:
            messages.error(request, 'Label is required.')
            return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})
        if not as_of_raw:
            messages.error(request, 'As-of date is required.')
            return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})

        try:
            as_of_date = datetime.strptime(as_of_raw, '%Y-%m-%d').date()
        except ValueError:
            messages.error(request, 'As-of date must be a valid date.')
            return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})

        try:
            cash_hand = Decimal(cash_hand_raw or '0')
            cash_acc = Decimal(cash_acc_raw or '0')
            receivable = Decimal(receivable_raw or '0')
        except InvalidOperation:
            messages.error(request, 'Cash and receivable amounts must be valid numbers.')
            return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})

        if cash_hand < 0 or cash_acc < 0 or receivable < 0:
            messages.error(request, 'Cash and receivable amounts cannot be negative.')
            return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})

        ob.label = label
        ob.as_of_date = as_of_date
        ob.cash_in_hand = cash_hand
        ob.cash_in_account = cash_acc
        ob.accounts_receivable = receivable
        ob.notes = notes
        ob.save()
        messages.success(request, f'Opening balance "{ob.label}" updated.')
        return redirect('settings')

    return render(request, 'settings/opening_balance_edit.html', {'opening_balance': ob})


@login_required
def opening_balance_delete(request, pk):
    """Delete an OpeningBalance row. POST only, confirmation handled on Settings page."""
    ob = get_object_or_404(OpeningBalance, pk=pk)
    if request.method != 'POST':
        return redirect('settings')
    label = ob.label
    ob.delete()
    messages.success(request, f'Opening balance "{label}" deleted.')
    return redirect('settings')


@login_required
def fy_wizard(request):
    """Guided financial-year rollover: backups -> opening balance -> numbering -> wipe.

    Single-page checklist. Each step shows derived status (done / pending / action
    needed) plus the right action button. Backup downloads are session-tracked
    because the system can't observe the actual download.
    """
    from django.db.models import F
    from crm.models import Lead

    if request.method == 'POST':
        flag = request.POST.get('mark_done')
        if flag in ('invoice_backup', 'expense_backup'):
            request.session[f'fy_wizard.{flag}_done'] = True
            messages.success(request, 'Step acknowledged.')
        elif flag == 'reset_backups':
            request.session.pop('fy_wizard.invoice_backup_done', None)
            request.session.pop('fy_wizard.expense_backup_done', None)
            messages.info(request, 'Backup acknowledgements cleared.')
        return redirect('fy_wizard')

    company = CompanySettings.get_settings()
    opening = OpeningBalance.current()
    last_reset = FYResetEvent.objects.first()

    outstanding_receivables = Invoice.objects.exclude(
        status__in=['paid', 'cancelled']
    ).aggregate(total=Sum(F('total_amount') - F('amount_paid')))['total'] or 0

    invoice_count = Invoice.objects.count()
    expense_count = Expense.objects.count()
    quote_count = Quote.objects.count()
    payment_count = Payment.objects.count()
    lead_count = Lead.objects.count()

    invoice_prefix_used = Invoice.objects.filter(invoice_number__startswith=company.invoice_prefix).exists()
    quote_prefix_used = Quote.objects.filter(quote_number__startswith=company.quote_prefix).exists()

    receivables_covered = bool(opening and opening.accounts_receivable >= outstanding_receivables)
    opening_after_last_reset = bool(opening and (not last_reset or opening.created_at > last_reset.ran_at))

    steps = [
        {
            'key': 'invoice_backup',
            'title': '1. Download invoice archive',
            'description': f'Single PDF of all {invoice_count} invoices. Save somewhere safe.',
            'done': request.session.get('fy_wizard.invoice_backup_done', False),
            'action_url': reverse('invoices_backup_pdf') + '?download=1',
            'action_label': 'Download Invoice PDF',
            'preview_url': reverse('invoices_backup_pdf'),
            'mark_done_flag': 'invoice_backup',
            'mandatory': invoice_count > 0,
        },
        {
            'key': 'expense_backup',
            'title': '2. Download expense archive',
            'description': f'Single PDF of all {expense_count} expenses, grouped by category. Save somewhere safe.',
            'done': request.session.get('fy_wizard.expense_backup_done', False),
            'action_url': reverse('expenses_backup_pdf') + '?download=1',
            'action_label': 'Download Expense PDF',
            'preview_url': reverse('expenses_backup_pdf'),
            'mark_done_flag': 'expense_backup',
            'mandatory': expense_count > 0,
        },
        {
            'key': 'opening_balance',
            'title': '3. Record opening balance for the new FY',
            'description': (
                f'Capture cash + receivables (currently &#8377;{outstanding_receivables:.2f} outstanding) '
                'so year-end position is preserved.'
            ),
            'done': opening_after_last_reset and receivables_covered,
            'action_url': reverse('settings') + '#financials-panel',
            'action_label': 'Open Opening Balance form',
            'mandatory': True,
        },
        {
            'key': 'invoice_numbering',
            'title': '4. Start new invoice numbering',
            'description': (
                f'Current prefix: <strong>{company.invoice_prefix}</strong>. '
                'Switch to a fresh prefix for the new FY (e.g. RT2627-).'
            ),
            'done': not invoice_prefix_used,
            'action_url': reverse('settings') + '#invoice-panel',
            'action_label': 'Open Invoice Settings',
            'mandatory': True,
        },
        {
            'key': 'quote_numbering',
            'title': '5. Start new quote numbering',
            'description': (
                f'Current prefix: <strong>{company.quote_prefix}</strong>. '
                'Switch to a fresh quote prefix so new quotes restart at 1.'
            ),
            'done': not quote_prefix_used,
            'action_url': reverse('settings') + '#invoice-panel',
            'action_label': 'Open Invoice Settings',
            'mandatory': False,
        },
        {
            'key': 'reset',
            'title': '6. Wipe last FY data',
            'description': (
                f'Delete {invoice_count} invoices, {payment_count} payments, '
                f'{expense_count} expenses, {lead_count} leads. Cannot be undone.'
            ),
            'done': False,
            'action_url': reverse('fy_reset'),
            'action_label': 'Go to Reset confirmation',
            'mandatory': True,
        },
    ]

    prior_steps_done = all(s['done'] for s in steps[:-1] if s['mandatory'])

    context = {
        'steps': steps,
        'opening': opening,
        'last_reset': last_reset,
        'outstanding_receivables': outstanding_receivables,
        'receivables_covered': receivables_covered,
        'company': company,
        'invoice_count': invoice_count,
        'expense_count': expense_count,
        'quote_count': quote_count,
        'prior_steps_done': prior_steps_done,
    }
    return render(request, 'settings/fy_wizard.html', context)


@login_required
def fy_reset(request):
    """Wipe all financial data so a new financial year can be started from scratch.

    Deletes: Invoices (and cascading InvoiceItems + Payments), Expenses, CRM Leads.
    Keeps: Clients, Projects, Credentials, AMC contracts, Licenses, HR/Team, Users.

    GET shows counts and confirmation form. POST executes after confirmation
    (typed "RESET" + checkbox acknowledging the backup was downloaded).
    """
    from django.db import transaction
    from django.db.models import F
    from crm.models import Lead

    outstanding_receivables = Invoice.objects.exclude(
        status__in=['paid', 'cancelled']
    ).aggregate(total=Sum(F('total_amount') - F('amount_paid')))['total'] or 0
    opening = OpeningBalance.current()
    receivables_captured = opening.accounts_receivable if opening else 0

    counts = {
        'invoices': Invoice.objects.count(),
        'invoice_items': InvoiceItem.objects.count(),
        'payments': Payment.objects.count(),
        'expenses': Expense.objects.count(),
        'leads': Lead.objects.count(),
        # Kept (shown for clarity)
        'clients': Client.objects.count(),
        'projects': Project.objects.count(),
        'quotes': Quote.objects.count(),
        'amc_contracts': AMCContract.objects.count(),
        'amc_payments': AMCPayment.objects.count(),
        'credentials': Credential.objects.count(),
        'team_members': TeamMember.objects.count(),
    }
    context = {
        'counts': counts,
        'outstanding_receivables': outstanding_receivables,
        'opening_balance': opening,
        'receivables_captured': receivables_captured,
    }

    if request.method == 'POST':
        typed = (request.POST.get('confirm_text') or '').strip()
        backed_up = request.POST.get('backup_downloaded') == 'on'

        if typed != 'RESET':
            messages.error(request, 'You must type RESET exactly to confirm.')
            return render(request, 'settings/fy_reset.html', context)
        if not backed_up:
            messages.error(request, 'Please confirm that you have downloaded the backup PDF before resetting.')
            return render(request, 'settings/fy_reset.html', context)

        with transaction.atomic():
            # Order: Payment first (it has a save() side-effect on Invoice; deleting
            # invoices first would cascade, but explicit deletion is safer and the
            # counts above are computed pre-wipe so they remain accurate.
            Payment.objects.all().delete()
            Invoice.objects.all().delete()  # cascades to InvoiceItem (and any leftover Payment)
            Expense.objects.all().delete()
            Lead.objects.all().delete()  # cascades to LeadNote/FollowUp/Activity/Demo

            FYResetEvent.objects.create(
                ran_by=request.user if request.user.is_authenticated else None,
                invoices_wiped=counts['invoices'],
                payments_wiped=counts['payments'],
                expenses_wiped=counts['expenses'],
                leads_wiped=counts['leads'],
                outstanding_receivables=outstanding_receivables,
                opening_balance=opening,
                invoice_prefix_after=CompanySettings.get_settings().invoice_prefix,
            )

        messages.success(
            request,
            f"Financial data reset complete. Removed {counts['invoices']} invoices, "
            f"{counts['payments']} payments, {counts['expenses']} expenses, "
            f"{counts['leads']} leads. Clients, projects, AMC contracts, and credentials kept."
        )
        return redirect('settings')

    return render(request, 'settings/fy_reset.html', context)


@login_required
def start_new_fy(request):
    """Reset invoice numbering for a new financial year by switching prefix and starting number."""
    if request.method != 'POST':
        return redirect('settings')

    company = CompanySettings.get_settings()
    new_prefix = request.POST.get('new_prefix', '').strip()
    new_start_raw = request.POST.get('new_starting_number', '1').strip()

    if not new_prefix:
        messages.error(request, 'New prefix is required.')
        return redirect('settings')

    max_prefix_len = CompanySettings._meta.get_field('invoice_prefix').max_length
    if len(new_prefix) > max_prefix_len:
        messages.error(request, f'Prefix is too long ({len(new_prefix)} chars). Max {max_prefix_len} characters allowed.')
        return redirect('settings')

    if new_prefix == company.invoice_prefix:
        messages.error(request, f'New prefix "{new_prefix}" is the same as the current prefix. Use a different prefix for the new financial year.')
        return redirect('settings')

    if Invoice.objects.filter(invoice_number__startswith=new_prefix).exists():
        messages.error(request, f'Prefix "{new_prefix}" is already used by existing invoices. Pick a prefix that has never been used.')
        return redirect('settings')

    try:
        new_start = int(new_start_raw) if new_start_raw else 1
        if new_start < 1:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, 'Starting number must be a positive integer.')
        return redirect('settings')

    company.invoice_prefix = new_prefix
    company.invoice_starting_number = new_start
    company.save(update_fields=['invoice_prefix', 'invoice_starting_number'])

    messages.success(request, f'New financial year started. Next invoice will be "{new_prefix}{new_start}". Previous invoices are unchanged.')
    return redirect('settings')


@login_required
def start_new_fy_quotes(request):
    """Reset quote numbering for a new financial year by switching prefix and starting number."""
    if request.method != 'POST':
        return redirect('settings')

    company = CompanySettings.get_settings()
    new_prefix = request.POST.get('new_prefix', '').strip()
    new_start_raw = request.POST.get('new_starting_number', '1').strip()

    if not new_prefix:
        messages.error(request, 'New quote prefix is required.')
        return redirect('settings')

    max_prefix_len = CompanySettings._meta.get_field('quote_prefix').max_length
    if len(new_prefix) > max_prefix_len:
        messages.error(request, f'Quote prefix is too long ({len(new_prefix)} chars). Max {max_prefix_len} characters allowed.')
        return redirect('settings')

    if new_prefix == company.quote_prefix:
        messages.error(request, f'New prefix "{new_prefix}" matches the current quote prefix. Pick a different one for the new financial year.')
        return redirect('settings')

    if Quote.objects.filter(quote_number__startswith=new_prefix).exists():
        messages.error(request, f'Prefix "{new_prefix}" is already used by existing quotes. Pick a prefix that has never been used.')
        return redirect('settings')

    try:
        new_start = int(new_start_raw) if new_start_raw else 1
        if new_start < 1:
            raise ValueError
    except (ValueError, TypeError):
        messages.error(request, 'Starting number must be a positive integer.')
        return redirect('settings')

    company.quote_prefix = new_prefix
    company.quote_starting_number = new_start
    company.save(update_fields=['quote_prefix', 'quote_starting_number'])

    messages.success(request, f'New financial year started for quotes. Next quote will be "{new_prefix}{new_start}". Previous quotes are unchanged.')
    return redirect('settings')


@login_required
def reports_view(request):
    import json
    from dateutil.relativedelta import relativedelta
    from collections import defaultdict

    # Revenue stats
    total_revenue = Payment.objects.aggregate(total=Sum('amount'))['total'] or 0

    # This month
    first_day = timezone.now().replace(day=1)
    this_month_revenue = Payment.objects.filter(
        payment_date__gte=first_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    # Expense stats
    total_expenses = Expense.objects.aggregate(total=Sum('amount'))['total'] or 0
    this_month_expenses = Expense.objects.filter(
        date__gte=first_day
    ).aggregate(total=Sum('amount'))['total'] or 0

    net_profit_total = float(total_revenue) - float(total_expenses)
    net_profit_month = float(this_month_revenue) - float(this_month_expenses)

    # Outstanding
    outstanding = Invoice.objects.exclude(
        status__in=['paid', 'cancelled']
    ).aggregate(
        total=Sum('total_amount') - Sum('amount_paid')
    )['total'] or 0

    # ============== Chart Data ==============
    today = timezone.now().date()

    # Monthly Revenue & Expenses (Last 12 months)
    monthly_revenue_labels = []
    monthly_revenue_data = []
    monthly_invoiced_data = []
    monthly_expenses_data = []

    for i in range(11, -1, -1):
        month_date = today - relativedelta(months=i)
        month_start = month_date.replace(day=1)
        if i > 0:
            month_end = (month_date + relativedelta(months=1)).replace(day=1) - timedelta(days=1)
        else:
            month_end = today

        # Payments received
        month_revenue = Payment.objects.filter(
            payment_date__gte=month_start,
            payment_date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        # Invoices issued
        month_invoiced = Invoice.objects.filter(
            issue_date__gte=month_start,
            issue_date__lte=month_end
        ).aggregate(total=Sum('total_amount'))['total'] or 0

        # Expenses incurred
        month_expense = Expense.objects.filter(
            date__gte=month_start,
            date__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        monthly_revenue_labels.append(month_date.strftime('%b %Y'))
        monthly_revenue_data.append(float(month_revenue))
        monthly_invoiced_data.append(float(month_invoiced))
        monthly_expenses_data.append(float(month_expense))

    # Revenue by Client (Top 5)
    client_revenue = defaultdict(float)
    for payment in Payment.objects.select_related('invoice__client').all():
        client_revenue[payment.invoice.client.name] += float(payment.amount)

    sorted_clients = sorted(client_revenue.items(), key=lambda x: x[1], reverse=True)[:5]
    client_labels = [c[0] for c in sorted_clients]
    client_data = [c[1] for c in sorted_clients]

    # Revenue by Project Type
    project_type_revenue = defaultdict(float)
    for payment in Payment.objects.select_related('invoice__project').all():
        if payment.invoice.project:
            project_type = payment.invoice.project.get_project_type_display()
        else:
            project_type = 'No Project'
        project_type_revenue[project_type] += float(payment.amount)

    project_type_labels = list(project_type_revenue.keys())
    project_type_data = list(project_type_revenue.values())

    # Expenses by Category
    category_map = dict(Expense.CATEGORY_CHOICES)
    expense_by_category_qs = Expense.objects.values('category').annotate(
        total=Sum('amount')
    ).order_by('-total')
    expense_category_labels = [
        category_map.get(row['category'], row['category']) for row in expense_by_category_qs
    ]
    expense_category_data = [float(row['total'] or 0) for row in expense_by_category_qs]

    # Quarterly Comparison
    quarterly_data = []
    quarterly_labels = []
    for q in range(3, -1, -1):
        quarter_start = today - relativedelta(months=q*3)
        quarter_end = quarter_start + relativedelta(months=3) - timedelta(days=1)
        q_start = quarter_start.replace(day=1)

        quarter_revenue = Payment.objects.filter(
            payment_date__gte=q_start,
            payment_date__lte=quarter_end
        ).aggregate(total=Sum('amount'))['total'] or 0

        quarterly_labels.append(f"Q{((quarter_start.month-1)//3)+1} {quarter_start.year}")
        quarterly_data.append(float(quarter_revenue))

    # Invoice collection rate
    total_invoiced = Invoice.objects.aggregate(total=Sum('total_amount'))['total'] or 0
    collection_rate = (float(total_revenue) / float(total_invoiced) * 100) if total_invoiced else 0

    # Counts for summary
    total_invoices = Invoice.objects.count()
    paid_invoices = Invoice.objects.filter(status='paid').count()
    total_projects = Project.objects.count()
    completed_projects = Project.objects.filter(status='completed').count()
    total_clients = Client.objects.filter(is_active=True).count()

    context = {
        'total_revenue': total_revenue,
        'this_month_revenue': this_month_revenue,
        'total_expenses': total_expenses,
        'this_month_expenses': this_month_expenses,
        'net_profit_total': net_profit_total,
        'net_profit_month': net_profit_month,
        'outstanding': outstanding,
        'collection_rate': collection_rate,
        'total_invoices': total_invoices,
        'paid_invoices': paid_invoices,
        'total_projects': total_projects,
        'completed_projects': completed_projects,
        'total_clients': total_clients,
        # Chart data
        'monthly_revenue_labels': json.dumps(monthly_revenue_labels),
        'monthly_revenue_data': json.dumps(monthly_revenue_data),
        'monthly_invoiced_data': json.dumps(monthly_invoiced_data),
        'monthly_expenses_data': json.dumps(monthly_expenses_data),
        'client_labels': json.dumps(client_labels),
        'client_data': json.dumps(client_data),
        'project_type_labels': json.dumps(project_type_labels),
        'project_type_data': json.dumps(project_type_data),
        'expense_category_labels': json.dumps(expense_category_labels),
        'expense_category_data': json.dumps(expense_category_data),
        'quarterly_labels': json.dumps(quarterly_labels),
        'quarterly_data': json.dumps(quarterly_data),
    }
    return render(request, 'reports/index.html', context)


@login_required
def monthly_report_view(request):
    """Monthly income & expense report for a single month."""
    import json
    from datetime import date
    from dateutil.relativedelta import relativedelta
    from collections import defaultdict

    # Resolve the target month from ?month=YYYY-MM (default current month)
    today = timezone.now().date()
    month_param = request.GET.get('month', '')
    try:
        year_str, mon_str = month_param.split('-')
        month_start = date(int(year_str), int(mon_str), 1)
    except (ValueError, AttributeError):
        month_start = today.replace(day=1)

    month_end = (month_start + relativedelta(months=1)) - timedelta(days=1)
    prev_month = (month_start - relativedelta(months=1)).strftime('%Y-%m')
    next_month = (month_start + relativedelta(months=1)).strftime('%Y-%m')

    # Income (payments)
    payments = Payment.objects.filter(
        payment_date__gte=month_start, payment_date__lte=month_end
    ).select_related('invoice', 'invoice__client').order_by('payment_date')
    total_income = payments.aggregate(t=Sum('amount'))['t'] or 0

    # Income by client
    client_income = defaultdict(float)
    for p in payments:
        client_income[p.invoice.client.name] += float(p.amount)
    income_by_client = sorted(client_income.items(), key=lambda x: x[1], reverse=True)

    # Income by payment method
    method_map = dict(Payment.METHOD_CHOICES)
    method_income_qs = payments.values('payment_method').annotate(
        total=Sum('amount')
    ).order_by('-total')
    income_by_method = [
        (method_map.get(row['payment_method'], row['payment_method']), float(row['total'] or 0))
        for row in method_income_qs
    ]

    # Expenses
    expenses = Expense.objects.filter(
        date__gte=month_start, date__lte=month_end
    ).select_related('project', 'project__client').order_by('date')
    total_expenses = expenses.aggregate(t=Sum('amount'))['t'] or 0

    # Expenses by category
    category_map = dict(Expense.CATEGORY_CHOICES)
    expense_by_category_qs = expenses.values('category').annotate(
        total=Sum('amount')
    ).order_by('-total')
    expense_by_category = [
        (category_map.get(row['category'], row['category']), float(row['total'] or 0))
        for row in expense_by_category_qs
    ]

    net_profit = float(total_income) - float(total_expenses)

    # Internal transfers in this month (reported separately, not part of P&L)
    transfers = InternalTransfer.objects.filter(
        date__gte=month_start, date__lte=month_end
    ).select_related('from_account', 'to_account').order_by('date')
    total_transfers = transfers.aggregate(t=Sum('amount'))['t'] or 0

    # Available months from earliest payment/expense to current month (for dropdown)
    earliest_payment = Payment.objects.order_by('payment_date').values_list('payment_date', flat=True).first()
    earliest_expense = Expense.objects.order_by('date').values_list('date', flat=True).first()
    candidates = [d for d in (earliest_payment, earliest_expense) if d]
    earliest = min(candidates) if candidates else today.replace(day=1)
    earliest = earliest.replace(day=1)
    available_months = []
    cursor = today.replace(day=1)
    while cursor >= earliest:
        available_months.append({
            'value': cursor.strftime('%Y-%m'),
            'label': cursor.strftime('%b %Y'),
        })
        cursor -= relativedelta(months=1)

    context = {
        'month_start': month_start,
        'month_end': month_end,
        'month_label': month_start.strftime('%B %Y'),
        'month_value': month_start.strftime('%Y-%m'),
        'prev_month': prev_month,
        'next_month': next_month,
        'available_months': available_months,
        'payments': payments,
        'expenses': expenses,
        'total_income': total_income,
        'total_expenses': total_expenses,
        'net_profit': net_profit,
        'income_by_client': income_by_client,
        'income_by_method': income_by_method,
        'expense_by_category': expense_by_category,
        'expense_category_labels': json.dumps([c[0] for c in expense_by_category]),
        'expense_category_data': json.dumps([c[1] for c in expense_by_category]),
        'income_method_labels': json.dumps([m[0] for m in income_by_method]),
        'income_method_data': json.dumps([m[1] for m in income_by_method]),
        'transfers': transfers,
        'total_transfers': total_transfers,
    }
    return render(request, 'reports/monthly.html', context)


# ============== Global Search ==============

@login_required
def global_search(request):
    from django.http import JsonResponse

    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'results': []})

    results = []

    # Search Clients
    clients = Client.objects.filter(
        Q(name__icontains=query) |
        Q(email__icontains=query) |
        Q(company_name__icontains=query)
    ).filter(is_active=True)[:5]

    for client in clients:
        results.append({
            'type': 'client',
            'icon': 'fa-user',
            'title': client.name,
            'subtitle': client.company_name or client.email or '',
            'url': f'/clients/{client.pk}/'
        })

    # Search Projects
    projects = Project.objects.filter(
        Q(name__icontains=query) |
        Q(description__icontains=query)
    ).select_related('client')[:5]

    for project in projects:
        results.append({
            'type': 'project',
            'icon': 'fa-folder-open',
            'title': project.name,
            'subtitle': project.client.name,
            'url': f'/projects/{project.pk}/'
        })

    # Search Invoices
    invoices = Invoice.objects.filter(
        Q(invoice_number__icontains=query) |
        Q(client__name__icontains=query)
    ).select_related('client')[:5]

    for invoice in invoices:
        results.append({
            'type': 'invoice',
            'icon': 'fa-file-invoice-dollar',
            'title': invoice.invoice_number,
            'subtitle': f'{invoice.client.name} - ₹{invoice.total_amount:,.0f}',
            'url': f'/invoices/{invoice.pk}/'
        })

    # Search Quotes
    quotes = Quote.objects.filter(
        Q(quote_number__icontains=query) |
        Q(client__name__icontains=query) |
        Q(title__icontains=query)
    ).select_related('client')[:5]

    for quote in quotes:
        results.append({
            'type': 'quote',
            'icon': 'fa-file-alt',
            'title': quote.quote_number,
            'subtitle': f'{quote.recipient_name} - {quote.title}',
            'url': f'/quotes/{quote.pk}/'
        })

    # Search Credentials
    credentials = Credential.objects.filter(
        Q(name__icontains=query) |
        Q(credential_type__icontains=query) |
        Q(project__name__icontains=query)
    ).select_related('project')[:5]

    for credential in credentials:
        results.append({
            'type': 'credential',
            'icon': 'fa-key',
            'title': credential.name,
            'subtitle': f'{credential.project.name} - {credential.get_credential_type_display()}',
            'url': f'/credentials/{credential.pk}/'
        })

    return JsonResponse({'results': results[:15]})


# ============== Excel Import ==============

@login_required
def client_import(request):
    """Import clients from Excel file"""
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            imported = 0
            skipped = 0
            errors = []

            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                name = str(row[0]).strip() if row[0] else ''
                email = str(row[1]).strip() if row[1] else ''
                phone = str(row[2]).strip() if row[2] else ''
                company_name = str(row[3]).strip() if len(row) > 3 and row[3] else ''
                address = str(row[4]).strip() if len(row) > 4 and row[4] else ''
                gst_number = str(row[5]).strip() if len(row) > 5 and row[5] else ''

                if not name:
                    errors.append(f'Row {row_num}: Name is required')
                    skipped += 1
                    continue

                # Check for duplicate email
                if email and Client.objects.filter(email=email).exists():
                    errors.append(f'Row {row_num}: Email {email} already exists')
                    skipped += 1
                    continue

                try:
                    Client.objects.create(
                        name=name,
                        email=email,
                        phone=phone,
                        company_name=company_name,
                        address=address,
                        gst_number=gst_number,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    skipped += 1

            if imported > 0:
                messages.success(request, f'Successfully imported {imported} client(s).')
            if skipped > 0:
                messages.warning(request, f'Skipped {skipped} row(s). Check errors below.')
            if errors:
                for error in errors[:5]:  # Show first 5 errors
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')

        return redirect('client_list')

    # GET request - show import form or download template
    if request.GET.get('template') == '1':
        # Generate sample template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Clients'

        # Header row
        headers = ['Name*', 'Email', 'Phone', 'Company Name', 'Address', 'GST Number']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Sample row
        sample = ['John Doe', 'john@example.com', '+91 9876543210', 'ABC Corp', '123 Main St, City', 'GSTIN123456']
        for col, value in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=value)

        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clients_template.xlsx'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

        return response

    return render(request, 'clients/import.html')


@login_required
def project_import(request):
    """Import projects from Excel file"""
    from django.http import HttpResponse
    import openpyxl
    from io import BytesIO

    clients = Client.objects.filter(is_active=True)

    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']

        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

            imported = 0
            skipped = 0
            errors = []

            # Skip header row
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue

                name = str(row[0]).strip() if row[0] else ''
                client_name = str(row[1]).strip() if row[1] else ''
                project_type = str(row[2]).strip().lower() if len(row) > 2 and row[2] else 'other'
                status = str(row[3]).strip().lower() if len(row) > 3 and row[3] else 'planning'
                description = str(row[4]).strip() if len(row) > 4 and row[4] else ''

                if not name:
                    errors.append(f'Row {row_num}: Project name is required')
                    skipped += 1
                    continue

                if not client_name:
                    errors.append(f'Row {row_num}: Client name is required')
                    skipped += 1
                    continue

                # Find client
                client = Client.objects.filter(
                    Q(name__iexact=client_name) | Q(email__iexact=client_name)
                ).first()

                if not client:
                    errors.append(f'Row {row_num}: Client "{client_name}" not found')
                    skipped += 1
                    continue

                # Validate project_type
                valid_types = ['website', 'mobile_app', 'webapp', 'ecommerce', 'maintenance', 'other']
                if project_type not in valid_types:
                    project_type = 'other'

                # Validate status
                valid_statuses = ['planning', 'in_progress', 'on_hold', 'completed', 'cancelled']
                if status not in valid_statuses:
                    status = 'planning'

                try:
                    Project.objects.create(
                        name=name,
                        client=client,
                        project_type=project_type,
                        status=status,
                        description=description,
                    )
                    imported += 1
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
                    skipped += 1

            if imported > 0:
                messages.success(request, f'Successfully imported {imported} project(s).')
            if skipped > 0:
                messages.warning(request, f'Skipped {skipped} row(s). Check errors below.')
            if errors:
                for error in errors[:5]:
                    messages.error(request, error)

        except Exception as e:
            messages.error(request, f'Error reading Excel file: {str(e)}')

        return redirect('project_list')

    # GET request - show import form or download template
    if request.GET.get('template') == '1':
        # Generate sample template
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Projects'

        # Header row
        headers = ['Project Name*', 'Client Name/Email*', 'Type', 'Status', 'Description']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # Sample row
        sample = ['Website Redesign', 'john@example.com', 'website', 'planning', 'Complete website redesign']
        for col, value in enumerate(sample, 1):
            ws.cell(row=2, column=col, value=value)

        # Add notes
        ws.cell(row=4, column=1, value='Notes:')
        ws.cell(row=5, column=1, value='Type: website, mobile_app, webapp, ecommerce, maintenance, other')
        ws.cell(row=6, column=1, value='Status: planning, in_progress, on_hold, completed, cancelled')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=projects_template.xlsx'

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response.write(buffer.read())

        return response

    return render(request, 'projects/import.html', {'clients': clients})


# ============== User Profile ==============

@login_required
def profile_view(request):
    """View and edit user profile"""
    user = request.user

    if request.method == 'POST':
        user.first_name = request.POST.get('first_name', '')
        user.last_name = request.POST.get('last_name', '')
        user.email = request.POST.get('email', '')
        user.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')

    return render(request, 'profile/index.html', {'profile_user': user})


@login_required
def change_password(request):
    """Change user password"""
    from django.contrib.auth import update_session_auth_hash

    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')

        if not request.user.check_password(current_password):
            messages.error(request, 'Current password is incorrect.')
            return redirect('profile')

        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match.')
            return redirect('profile')

        if len(new_password) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
            return redirect('profile')

        request.user.set_password(new_password)
        request.user.save()
        update_session_auth_hash(request, request.user)
        messages.success(request, 'Password changed successfully.')
        return redirect('profile')

    return redirect('profile')


# ============== Export to Excel ==============

@login_required
def export_clients(request):
    """Export clients to Excel"""
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Clients"

    # Header
    headers = ['Company Name', 'Contact Name', 'Email', 'Phone', 'GST Number', 'Address', 'Created Date']
    ws.append(headers)

    # Data
    for client in Client.objects.all().order_by('company_name'):
        ws.append([
            client.company_name or '',
            client.name,
            client.email,
            client.phone or '',
            client.gst_number or '',
            client.address or '',
            client.created_at.strftime('%Y-%m-%d') if client.created_at else ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="clients_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_projects(request):
    """Export projects to Excel"""
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"

    # Header
    headers = ['Project Name', 'Client', 'Status', 'Start Date', 'End Date', 'Budget', 'Description']
    ws.append(headers)

    # Data
    for project in Project.objects.select_related('client').all().order_by('-created_at'):
        ws.append([
            project.name,
            project.client.name if project.client else '',
            project.get_status_display(),
            project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
            project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
            float(project.budget) if project.budget else 0,
            project.description or ''
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="projects_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_invoices(request):
    """Export invoices to Excel, honoring invoice_list filters"""
    from openpyxl import Workbook
    from django.http import HttpResponse
    from datetime import date
    from calendar import monthrange

    qs = Invoice.objects.select_related('client', 'project').all()

    search = request.GET.get('search', '')
    if search:
        qs = qs.filter(
            Q(invoice_number__icontains=search) |
            Q(title__icontains=search) |
            Q(client__name__icontains=search)
        )

    status_filter = request.GET.get('status', '')
    if status_filter == 'paid_partial':
        qs = qs.filter(status__in=['paid', 'partial'])
    elif status_filter:
        qs = qs.filter(status=status_filter)

    from_date = request.GET.get('from_date', '')
    to_date = request.GET.get('to_date', '')
    period = request.GET.get('period', '')
    if period in ('prev_month', 'this_month') and not from_date and not to_date:
        today = date.today()
        if period == 'prev_month':
            year = today.year if today.month > 1 else today.year - 1
            month = today.month - 1 if today.month > 1 else 12
        else:
            year, month = today.year, today.month
        from_date = date(year, month, 1).isoformat()
        to_date = date(year, month, monthrange(year, month)[1]).isoformat()

    if from_date:
        qs = qs.filter(issue_date__gte=from_date)
    if to_date:
        qs = qs.filter(issue_date__lte=to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"

    # Header
    headers = ['Invoice Number', 'Client', 'Project', 'Title', 'Issue Date', 'Due Date', 'Subtotal', 'Tax', 'Total', 'Paid', 'Balance', 'Status']
    ws.append(headers)

    # Data
    for invoice in qs.order_by('-issue_date'):
        ws.append([
            invoice.invoice_number,
            invoice.client.name if invoice.client else '',
            invoice.project.name if invoice.project else '',
            invoice.title or '',
            invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else '',
            invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
            float(invoice.subtotal) if invoice.subtotal else 0,
            float(invoice.tax_amount) if invoice.tax_amount else 0,
            float(invoice.total_amount) if invoice.total_amount else 0,
            float(invoice.amount_paid) if invoice.amount_paid else 0,
            float(invoice.balance_due) if invoice.balance_due else 0,
            invoice.get_status_display()
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="invoices_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_quotes(request):
    """Export quotes to Excel"""
    from openpyxl import Workbook
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Quotes"

    # Header
    headers = ['Quote Number', 'Client', 'Project', 'Title', 'Issue Date', 'Valid Until', 'Subtotal', 'Tax', 'Total', 'Status']
    ws.append(headers)

    # Data
    for quote in Quote.objects.select_related('client', 'project').all().order_by('-issue_date'):
        ws.append([
            quote.quote_number,
            quote.client.name if quote.client else '',
            quote.project.name if quote.project else '',
            quote.title or '',
            quote.issue_date.strftime('%Y-%m-%d') if quote.issue_date else '',
            quote.valid_until.strftime('%Y-%m-%d') if quote.valid_until else '',
            float(quote.subtotal) if quote.subtotal else 0,
            float(quote.tax_amount) if quote.tax_amount else 0,
            float(quote.total_amount) if quote.total_amount else 0,
            quote.get_status_display()
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="quotes_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response


# ============== Backup & Restore ==============

@login_required
def backup_view(request):
    """Backup management page"""
    import os

    # List existing backups
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    backups = []

    if os.path.exists(backup_dir):
        for filename in sorted(os.listdir(backup_dir), reverse=True):
            if filename.endswith('.json'):
                filepath = os.path.join(backup_dir, filename)
                stat = os.stat(filepath)
                backups.append({
                    'filename': filename,
                    'size': stat.st_size,
                    'created': timezone.datetime.fromtimestamp(stat.st_mtime)
                })

    return render(request, 'backup/index.html', {'backups': backups[:10]})


@login_required
def backup_download(request):
    """Create and download database backup"""
    import json
    import os
    from django.http import HttpResponse
    from django.core import serializers

    # Create backup data
    backup_data = {
        'created_at': timezone.now().isoformat(),
        'version': '1.0',
        'data': {}
    }

    # Export all models
    models_to_backup = [
        ('clients', Client),
        ('projects', Project),
        ('credentials', Credential),
        ('quotes', Quote),
        ('quote_items', QuoteItem),
        ('invoices', Invoice),
        ('invoice_items', InvoiceItem),
        ('payments', Payment),
        ('company_settings', CompanySettings),
    ]

    for name, model in models_to_backup:
        backup_data['data'][name] = json.loads(serializers.serialize('json', model.objects.all()))

    # Create backup file
    backup_json = json.dumps(backup_data, indent=2, default=str)

    # Save to backups folder
    backup_dir = os.path.join(settings.BASE_DIR, 'backups')
    os.makedirs(backup_dir, exist_ok=True)

    filename = f"backup_{timezone.now().strftime('%Y%m%d_%H%M%S')}.json"
    filepath = os.path.join(backup_dir, filename)

    with open(filepath, 'w') as f:
        f.write(backup_json)

    # Return as download
    response = HttpResponse(backup_json, content_type='application/json')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    messages.success(request, f'Backup created: {filename}')
    return response


@login_required
def backup_restore(request):
    """Restore database from backup"""
    import json
    from django.core import serializers
    from django.db import transaction

    if request.method != 'POST':
        return redirect('backup')

    backup_file = request.FILES.get('backup_file')
    if not backup_file:
        messages.error(request, 'Please select a backup file.')
        return redirect('backup')

    try:
        backup_data = json.load(backup_file)

        if 'data' not in backup_data:
            messages.error(request, 'Invalid backup file format.')
            return redirect('backup')

        with transaction.atomic():
            # Restore in order (respecting foreign keys)
            restore_order = [
                ('clients', Client),
                ('projects', Project),
                ('credentials', Credential),
                ('quotes', Quote),
                ('quote_items', QuoteItem),
                ('invoices', Invoice),
                ('invoice_items', InvoiceItem),
                ('payments', Payment),
                ('company_settings', CompanySettings),
            ]

            for name, model in restore_order:
                if name in backup_data['data']:
                    # Clear existing data
                    model.objects.all().delete()

                    # Restore from backup
                    for obj_data in backup_data['data'][name]:
                        for obj in serializers.deserialize('json', json.dumps([obj_data])):
                            obj.save()

        messages.success(request, 'Backup restored successfully.')
    except json.JSONDecodeError:
        messages.error(request, 'Invalid JSON file.')
    except Exception as e:
        messages.error(request, f'Restore failed: {str(e)}')

    return redirect('backup')


# ============== Expense Views ==============

@login_required
def expense_list(request):
    """List all expenses with filtering"""
    expenses = Expense.objects.select_related('project', 'project__client').all()

    # Filters
    search = request.GET.get('search', '')
    category = request.GET.get('category', '')
    project_id = request.GET.get('project', '')

    if search:
        expenses = expenses.filter(
            Q(vendor__icontains=search) | Q(description__icontains=search)
        )
    if category:
        expenses = expenses.filter(category=category)
    if project_id:
        expenses = expenses.filter(project_id=project_id)

    # Calculate totals
    total_expenses = expenses.aggregate(total=Sum('amount'))['total'] or 0
    billable_total = expenses.filter(is_billable=True).aggregate(total=Sum('amount'))['total'] or 0

    context = {
        'expenses': expenses,
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'category_choices': Expense.CATEGORY_CHOICES,
        'search': search,
        'selected_category': category,
        'selected_project': project_id,
        'total_expenses': total_expenses,
        'billable_total': billable_total,
    }
    return render(request, 'expenses/list.html', context)


@login_required
def expense_create(request):
    """Create a new expense"""
    if request.method == 'POST':
        expense = Expense(
            category=request.POST.get('category'),
            amount=request.POST.get('amount'),
            date=request.POST.get('date') or timezone.now().date(),
            vendor=request.POST.get('vendor'),
            description=request.POST.get('description', ''),
            is_billable=request.POST.get('is_billable') == 'on',
            payment_method=request.POST.get('payment_method', 'bank_transfer'),
            notes=request.POST.get('notes', ''),
        )

        project_id = request.POST.get('project')
        if project_id:
            expense.project = Project.objects.get(pk=project_id)

        if request.FILES.get('receipt'):
            expense.receipt = request.FILES['receipt']

        expense.save()

        # Log activity
        log_activity(request, 'created', expense)

        messages.success(request, 'Expense created successfully.')
        return redirect('expense_list')

    context = {
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'category_choices': Expense.CATEGORY_CHOICES,
        'payment_method_choices': Expense.PAYMENT_METHOD_CHOICES,
    }
    return render(request, 'expenses/form.html', context)


@login_required
def expense_update(request, pk):
    """Update an expense"""
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        expense.category = request.POST.get('category')
        expense.amount = request.POST.get('amount')
        expense.date = request.POST.get('date')
        expense.vendor = request.POST.get('vendor')
        expense.description = request.POST.get('description', '')
        expense.is_billable = request.POST.get('is_billable') == 'on'
        expense.payment_method = request.POST.get('payment_method', 'bank_transfer')
        expense.notes = request.POST.get('notes', '')

        project_id = request.POST.get('project')
        expense.project = Project.objects.get(pk=project_id) if project_id else None

        if request.FILES.get('receipt'):
            expense.receipt = request.FILES['receipt']

        expense.save()

        log_activity(request, 'updated', expense)

        messages.success(request, 'Expense updated successfully.')
        return redirect('expense_list')

    context = {
        'expense': expense,
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'category_choices': Expense.CATEGORY_CHOICES,
        'payment_method_choices': Expense.PAYMENT_METHOD_CHOICES,
    }
    return render(request, 'expenses/form.html', context)


@login_required
def expense_delete(request, pk):
    """Delete an expense"""
    expense = get_object_or_404(Expense, pk=pk)

    if request.method == 'POST':
        log_activity(request, 'deleted', expense)
        expense.delete()
        messages.success(request, 'Expense deleted successfully.')
        return redirect('expense_list')

    return render(request, 'expenses/delete.html', {'expense': expense})


# ============== Team Member Views ==============

@login_required
def team_dashboard(request):
    """Dashboard for team members showing their tasks and time entries"""
    # Get the team member profile for the current user
    team_member = getattr(request.user, 'team_profile', None)

    if not team_member:
        # If user is not a team member (admin), redirect to main dashboard
        return redirect('dashboard')

    from datetime import date, timedelta
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # Get tasks assigned to this team member
    my_tasks = Task.objects.filter(assigned_to=team_member).select_related('project')
    tasks_todo = my_tasks.filter(status='todo').count()
    tasks_in_progress = my_tasks.filter(status='in_progress').count()
    tasks_review = my_tasks.filter(status='review').count()
    tasks_completed = my_tasks.filter(status='completed').count()

    # Recent tasks
    recent_tasks = my_tasks.exclude(status='completed').order_by('-updated_at')[:5]

    # Get time entries for this team member
    my_time_entries = TimeEntry.objects.filter(user=request.user)

    # Time logged this week
    week_entries = my_time_entries.filter(date__gte=week_start)
    hours_this_week = week_entries.aggregate(total=Sum('hours'))['total'] or 0

    # Time logged this month
    month_entries = my_time_entries.filter(date__gte=month_start)
    hours_this_month = month_entries.aggregate(total=Sum('hours'))['total'] or 0

    # Recent time entries
    recent_time_entries = my_time_entries.select_related('project', 'task').order_by('-date')[:5]

    # Projects assigned to me
    my_projects = team_member.assigned_projects.all()

    context = {
        'team_member': team_member,
        'tasks_todo': tasks_todo,
        'tasks_in_progress': tasks_in_progress,
        'tasks_review': tasks_review,
        'tasks_completed': tasks_completed,
        'recent_tasks': recent_tasks,
        'hours_this_week': hours_this_week,
        'hours_this_month': hours_this_month,
        'recent_time_entries': recent_time_entries,
        'my_projects': my_projects,
        'today': today,
    }
    return render(request, 'team/dashboard.html', context)


@login_required
def my_tasks(request):
    """View tasks assigned to the current team member"""
    team_member = getattr(request.user, 'team_profile', None)

    if not team_member:
        return redirect('task_list')

    tasks = Task.objects.filter(assigned_to=team_member).select_related('project')

    status = request.GET.get('status', '')
    priority = request.GET.get('priority', '')

    if status:
        tasks = tasks.filter(status=status)
    if priority:
        tasks = tasks.filter(priority=priority)

    context = {
        'tasks': tasks,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'selected_status': status,
        'selected_priority': priority,
        'is_my_tasks': True,
    }
    return render(request, 'team/my_tasks.html', context)


@login_required
def my_time(request):
    """View time entries for the current user"""
    team_member = getattr(request.user, 'team_profile', None)

    if not team_member:
        return redirect('timeentry_list')

    entries = TimeEntry.objects.filter(user=request.user).select_related('project', 'task')

    from datetime import date, timedelta
    today = date.today()

    # Filter by date range
    date_filter = request.GET.get('date_filter', 'week')
    if date_filter == 'today':
        entries = entries.filter(date=today)
    elif date_filter == 'week':
        week_start = today - timedelta(days=today.weekday())
        entries = entries.filter(date__gte=week_start)
    elif date_filter == 'month':
        month_start = today.replace(day=1)
        entries = entries.filter(date__gte=month_start)

    total_hours = entries.aggregate(total=Sum('hours'))['total'] or 0

    context = {
        'entries': entries,
        'total_hours': total_hours,
        'date_filter': date_filter,
        'is_my_time': True,
    }
    return render(request, 'team/my_time.html', context)


@login_required
def team_list(request):
    """List all team members"""
    members = TeamMember.objects.all()

    search = request.GET.get('search', '')
    role = request.GET.get('role', '')

    if search:
        members = members.filter(Q(name__icontains=search) | Q(email__icontains=search))
    if role:
        members = members.filter(role=role)

    context = {
        'members': members,
        'role_choices': TeamMember.ROLE_CHOICES,
        'search': search,
        'selected_role': role,
    }
    return render(request, 'team/list.html', context)


@login_required
def team_detail(request, pk):
    """View team member details including assigned projects"""
    member = get_object_or_404(TeamMember, pk=pk)

    # Get assigned projects
    assigned_projects = member.assigned_projects.all()

    # Get tasks assigned to this member
    tasks = Task.objects.filter(assigned_to=member).select_related('project')
    tasks_todo = tasks.filter(status='todo').count()
    tasks_in_progress = tasks.filter(status='in_progress').count()
    tasks_review = tasks.filter(status='review').count()
    tasks_completed = tasks.filter(status='completed').count()
    recent_tasks = tasks.exclude(status='completed').order_by('-updated_at')[:5]

    # Get time entries if freelancer
    time_entries = []
    total_hours = 0
    if member.is_freelancer and member.user:
        time_entries = TimeEntry.objects.filter(user=member.user).select_related('project', 'task').order_by('-date')[:10]
        total_hours = TimeEntry.objects.filter(user=member.user).aggregate(total=Sum('hours'))['total'] or 0

    # Get employee app data if linked
    employee = None
    leave_requests = []
    work_assignments = []
    recent_attendance = []
    if member.user:
        from employees.models import Employee, LeaveRequest, WorkAssignment, Attendance
        employee = Employee.objects.filter(user=member.user).first()
        if employee:
            leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')[:10]
            work_assignments = WorkAssignment.objects.filter(assigned_to=employee).order_by('-created_at')[:10]
            recent_attendance = Attendance.objects.filter(employee=employee).order_by('-date')[:10]

    context = {
        'member': member,
        'assigned_projects': assigned_projects,
        'tasks_todo': tasks_todo,
        'tasks_in_progress': tasks_in_progress,
        'tasks_review': tasks_review,
        'tasks_completed': tasks_completed,
        'recent_tasks': recent_tasks,
        'time_entries': time_entries,
        'total_hours': total_hours,
        'employee': employee,
        'leave_requests': leave_requests,
        'work_assignments': work_assignments,
        'recent_attendance': recent_attendance,
    }
    return render(request, 'team/detail.html', context)


@login_required
def team_create(request):
    """Create a new team member with optional login account"""
    if request.method == 'POST':
        employment_type = request.POST.get('employment_type', 'permanent')

        member = TeamMember(
            name=request.POST.get('name'),
            email=request.POST.get('email', ''),
            phone=request.POST.get('phone', ''),
            role=request.POST.get('role', 'developer'),
            employment_type=employment_type,
            is_active=request.POST.get('is_active') == 'true',
            notes=request.POST.get('notes', ''),
        )

        # Set salary or hourly rate based on employment type
        if employment_type == 'freelancer':
            hourly_rate = request.POST.get('hourly_rate')
            if hourly_rate:
                member.hourly_rate = hourly_rate
        else:
            monthly_salary = request.POST.get('monthly_salary')
            if monthly_salary:
                member.monthly_salary = monthly_salary

        # Create user account if requested
        create_account = request.POST.get('create_account') == 'on'
        if create_account:
            username = request.POST.get('username', '').strip()
            password = request.POST.get('password', '')

            if username and password:
                # Check if username exists
                if User.objects.filter(username=username).exists():
                    messages.error(request, f'Username "{username}" already exists.')
                    context = {
                        'role_choices': TeamMember.ROLE_CHOICES,
                        'employment_type_choices': TeamMember.EMPLOYMENT_TYPE_CHOICES,
                        'form_data': request.POST,
                    }
                    return render(request, 'team/form.html', context)

                # Create user
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    email=member.email,
                    first_name=member.name.split()[0] if member.name else '',
                    last_name=' '.join(member.name.split()[1:]) if member.name and len(member.name.split()) > 1 else '',
                )
                member.user = user

        member.save()
        log_activity(request, 'created', member)

        if member.user:
            messages.success(request, f'Team member "{member.name}" added with login account.')
        else:
            messages.success(request, 'Team member added successfully.')
        return redirect('team_list')

    context = {
        'role_choices': TeamMember.ROLE_CHOICES,
        'employment_type_choices': TeamMember.EMPLOYMENT_TYPE_CHOICES,
    }
    return render(request, 'team/form.html', context)


@login_required
def team_update(request, pk):
    """Update a team member"""
    member = get_object_or_404(TeamMember, pk=pk)

    if request.method == 'POST':
        employment_type = request.POST.get('employment_type', 'permanent')

        member.name = request.POST.get('name')
        member.email = request.POST.get('email', '')
        member.phone = request.POST.get('phone', '')
        member.role = request.POST.get('role', 'developer')
        member.employment_type = employment_type
        member.is_active = request.POST.get('is_active') == 'true'
        member.notes = request.POST.get('notes', '')

        # Set salary or hourly rate based on employment type
        if employment_type == 'freelancer':
            hourly_rate = request.POST.get('hourly_rate')
            member.hourly_rate = hourly_rate if hourly_rate else None
            member.monthly_salary = None
        else:
            monthly_salary = request.POST.get('monthly_salary')
            member.monthly_salary = monthly_salary if monthly_salary else None
            member.hourly_rate = None

        member.save()
        log_activity(request, 'updated', member)
        messages.success(request, 'Team member updated successfully.')
        return redirect('team_list')

    context = {
        'member': member,
        'role_choices': TeamMember.ROLE_CHOICES,
        'employment_type_choices': TeamMember.EMPLOYMENT_TYPE_CHOICES,
    }
    return render(request, 'team/form.html', context)


@login_required
def team_delete(request, pk):
    """Delete a team member"""
    member = get_object_or_404(TeamMember, pk=pk)

    if request.method == 'POST':
        # Check if member has tasks assigned
        if member.tasks.exists():
            messages.error(request, 'Cannot delete team member with assigned tasks. Reassign or delete tasks first.')
            return redirect('team_list')

        log_activity(request, 'deleted', member)
        member.delete()
        messages.success(request, 'Team member deleted successfully.')
        return redirect('team_list')

    return render(request, 'team/delete.html', {'member': member})


# ============== Task Views ==============

@login_required
def task_list(request):
    """List all tasks (filtered for team members)"""
    tasks = Task.objects.select_related('project', 'project__client', 'assigned_to').all()

    # Team members only see their assigned tasks
    team_member = getattr(request.user, 'team_profile', None)
    if team_member:
        tasks = tasks.filter(assigned_to=team_member)

    search = request.GET.get('search', '')
    status = request.GET.get('status', '')
    project_id = request.GET.get('project', '')
    priority = request.GET.get('priority', '')

    if search:
        tasks = tasks.filter(Q(title__icontains=search) | Q(description__icontains=search))
    if status:
        tasks = tasks.filter(status=status)
    if project_id:
        tasks = tasks.filter(project_id=project_id)
    if priority:
        tasks = tasks.filter(priority=priority)

    # Filter projects for team members
    if team_member:
        projects = Project.objects.filter(tasks__assigned_to=team_member).distinct()
    else:
        projects = Project.objects.filter(status__in=['in_progress', 'confirmed'])

    context = {
        'tasks': tasks,
        'projects': projects,
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'search': search,
        'selected_status': status,
        'selected_project': project_id,
        'selected_priority': priority,
        'is_team_member': team_member is not None,
    }
    return render(request, 'tasks/list.html', context)


@login_required
def task_board(request):
    """Kanban board view (filtered for team members)"""
    project_id = request.GET.get('project', '')

    tasks = Task.objects.select_related('project', 'assigned_to').all()

    # Team members only see their assigned tasks
    team_member = getattr(request.user, 'team_profile', None)
    if team_member:
        tasks = tasks.filter(assigned_to=team_member)

    if project_id:
        tasks = tasks.filter(project_id=project_id)

    # Filter projects for team members
    if team_member:
        projects = Project.objects.filter(tasks__assigned_to=team_member).distinct()
    else:
        projects = Project.objects.filter(status__in=['in_progress', 'confirmed'])

    context = {
        'todo_tasks': tasks.filter(status='todo'),
        'in_progress_tasks': tasks.filter(status='in_progress'),
        'review_tasks': tasks.filter(status='review'),
        'completed_tasks': tasks.filter(status='completed'),
        'projects': projects,
        'selected_project': project_id,
        'is_team_member': team_member is not None,
    }
    return render(request, 'tasks/board.html', context)


@login_required
def task_detail(request, pk):
    """Task detail view"""
    task = get_object_or_404(Task.objects.select_related('project', 'assigned_to'), pk=pk)
    time_entries = task.time_entries.all()
    comments = task.comments.filter(parent__isnull=True, is_deleted=False).select_related('author').prefetch_related('replies__author')
    issues = task.issues.select_related('reporter', 'assignee').all()
    activities = task.activities.select_related('actor')[:50]
    team_members = TeamMember.objects.filter(is_active=True) if hasattr(TeamMember, 'is_active') else TeamMember.objects.all()

    context = {
        'task': task,
        'time_entries': time_entries,
        'total_hours': time_entries.aggregate(total=Sum('hours'))['total'] or 0,
        'comments': comments,
        'issues': issues,
        'open_issue_count': issues.filter(status__in=['open', 'in_progress']).count(),
        'activities': activities,
        'team_members': team_members,
    }
    return render(request, 'tasks/detail.html', context)


# ============== Task Comments, Issues, Activity ==============

def _log_task_activity(task, actor, verb, from_value='', to_value='', message='', is_visible_to_client=False, related_comment=None, related_issue=None):
    return TaskActivity.objects.create(
        task=task, actor=actor, verb=verb,
        from_value=str(from_value)[:255], to_value=str(to_value)[:255],
        message=message[:500], is_visible_to_client=is_visible_to_client,
        related_comment=related_comment, related_issue=related_issue,
    )


@login_required
def task_comment_add(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method != 'POST':
        return redirect('task_detail', pk=pk)

    body = (request.POST.get('body') or '').strip()
    if not body:
        messages.error(request, 'Comment cannot be empty.')
        return redirect('task_detail', pk=pk)

    parent_id = request.POST.get('parent') or None
    parent = TaskComment.objects.filter(pk=parent_id, task=task).first() if parent_id else None
    is_visible = request.POST.get('is_visible_to_client') == 'on'
    attachment = request.FILES.get('attachment')

    comment = TaskComment.objects.create(
        task=task, author=request.user, body=body, parent=parent,
        attachment=attachment, is_visible_to_client=is_visible,
    )
    _log_task_activity(
        task, request.user, 'commented',
        message=body[:140], is_visible_to_client=is_visible, related_comment=comment,
    )
    messages.success(request, 'Comment posted.')
    return redirect(request.POST.get('next') or f"{reverse('task_detail', args=[pk])}#comments")


@login_required
def task_comment_delete(request, pk, comment_id):
    task = get_object_or_404(Task, pk=pk)
    comment = get_object_or_404(TaskComment, pk=comment_id, task=task)
    if comment.author_id != request.user.id and not request.user.is_superuser:
        messages.error(request, 'You cannot delete this comment.')
        return redirect('task_detail', pk=pk)
    comment.is_deleted = True
    comment.body = '[deleted]'
    comment.save(update_fields=['is_deleted', 'body', 'updated_at'])
    messages.success(request, 'Comment deleted.')
    return redirect(f"{reverse('task_detail', args=[pk])}#comments")


@login_required
def task_issue_create(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method != 'POST':
        return redirect('task_detail', pk=pk)

    title = (request.POST.get('title') or '').strip()
    if not title:
        messages.error(request, 'Issue title is required.')
        return redirect(f"{reverse('task_detail', args=[pk])}#issues")

    assignee_id = request.POST.get('assignee') or None
    assignee = TeamMember.objects.filter(pk=assignee_id).first() if assignee_id else None

    issue = TaskIssue.objects.create(
        task=task,
        reporter=request.user,
        assignee=assignee,
        title=title,
        description=request.POST.get('description', ''),
        severity=request.POST.get('severity', 'medium'),
        is_visible_to_client=request.POST.get('is_visible_to_client') == 'on',
    )
    _log_task_activity(
        task, request.user, 'issue_opened',
        to_value=issue.severity, message=issue.title[:140],
        is_visible_to_client=issue.is_visible_to_client, related_issue=issue,
    )
    messages.success(request, 'Issue reported.')
    return redirect(f"{reverse('task_detail', args=[pk])}#issues")


@login_required
def task_issue_update(request, pk, issue_id):
    task = get_object_or_404(Task, pk=pk)
    issue = get_object_or_404(TaskIssue, pk=issue_id, task=task)
    if request.method != 'POST':
        return redirect(f"{reverse('task_detail', args=[pk])}#issues")

    new_status = request.POST.get('status')
    resolution = request.POST.get('resolution', '').strip()
    changed = False
    if new_status and new_status in dict(TaskIssue.STATUS_CHOICES) and new_status != issue.status:
        old_status = issue.status
        issue.status = new_status
        if new_status in ('resolved', 'closed') and not issue.resolved_at:
            issue.resolved_at = timezone.now()
        if new_status in ('open', 'in_progress'):
            issue.resolved_at = None
        if resolution:
            issue.resolution = resolution
        issue.save()
        verb = 'issue_resolved' if new_status in ('resolved', 'closed') else 'issue_status_changed'
        _log_task_activity(
            task, request.user, verb,
            from_value=old_status, to_value=new_status, message=issue.title[:140],
            is_visible_to_client=issue.is_visible_to_client, related_issue=issue,
        )
        changed = True

    if changed:
        messages.success(request, 'Issue updated.')
    return redirect(f"{reverse('task_detail', args=[pk])}#issues")


@login_required
def task_create(request):
    """Create a new task"""
    if request.method == 'POST':
        task = Task(
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            status=request.POST.get('status', 'todo'),
            priority=request.POST.get('priority', 'medium'),
            notes=request.POST.get('notes', ''),
        )

        project_id = request.POST.get('project')
        if project_id:
            task.project = Project.objects.get(pk=project_id)

        assigned_to = request.POST.get('assigned_to')
        if assigned_to:
            task.assigned_to = TeamMember.objects.get(pk=assigned_to)

        due_date = request.POST.get('due_date')
        if due_date:
            task.due_date = due_date

        task.save()

        for f in request.FILES.getlist('attachments'):
            TaskAttachment.objects.create(task=task, file=f, name=f.name, uploaded_by=request.user)

        log_activity(request, 'created', task)

        messages.success(request, 'Task created successfully.')

        if request.GET.get('next') == 'board':
            return redirect('task_board')
        return redirect('task_list')

    context = {
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'team_members': TeamMember.objects.filter(is_active=True),
    }
    return render(request, 'tasks/form.html', context)


@login_required
def task_update(request, pk):
    """Update a task"""
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        task.title = request.POST.get('title')
        task.description = request.POST.get('description', '')
        task.status = request.POST.get('status', 'todo')
        task.priority = request.POST.get('priority', 'medium')
        task.notes = request.POST.get('notes', '')

        project_id = request.POST.get('project')
        task.project = Project.objects.get(pk=project_id) if project_id else None

        assigned_to = request.POST.get('assigned_to')
        task.assigned_to = TeamMember.objects.get(pk=assigned_to) if assigned_to else None

        due_date = request.POST.get('due_date')
        task.due_date = due_date if due_date else None

        # Set completed date if status changed to completed
        if task.status == 'completed' and not task.completed_date:
            task.completed_date = timezone.now().date()
        elif task.status != 'completed':
            task.completed_date = None

        task.save()

        # Delete removed attachments
        delete_ids = request.POST.getlist('delete_attachment')
        if delete_ids:
            TaskAttachment.objects.filter(pk__in=delete_ids, task=task).delete()

        # Add new attachments
        for f in request.FILES.getlist('attachments'):
            TaskAttachment.objects.create(task=task, file=f, name=f.name, uploaded_by=request.user)

        log_activity(request, 'updated', task)

        messages.success(request, 'Task updated successfully.')
        return redirect('task_detail', pk=pk)

    context = {
        'task': task,
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'status_choices': Task.STATUS_CHOICES,
        'priority_choices': Task.PRIORITY_CHOICES,
        'team_members': TeamMember.objects.filter(is_active=True),
    }
    return render(request, 'tasks/form.html', context)


@login_required
def task_delete(request, pk):
    """Delete a task"""
    task = get_object_or_404(Task, pk=pk)

    if request.method == 'POST':
        log_activity(request, 'deleted', task)
        task.delete()
        messages.success(request, 'Task deleted successfully.')
        return redirect('task_list')

    return render(request, 'tasks/delete.html', {'task': task})


@login_required
def task_status_update(request, pk):
    """Update task status via AJAX or form submission"""
    from django.http import JsonResponse

    if request.method == 'POST':
        task = get_object_or_404(Task, pk=pk)
        new_status = request.POST.get('status')

        if new_status in dict(Task.STATUS_CHOICES):
            old_status = task.status
            task.status = new_status
            if new_status == 'completed':
                task.completed_date = timezone.now().date()
            else:
                task.completed_date = None
            task.save()

            if old_status != new_status:
                _log_task_activity(
                    task, request.user, 'status_changed',
                    from_value=old_status, to_value=new_status, is_visible_to_client=True,
                )
            log_activity(request, 'updated', task)

            # Check if it's an AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': True})

            # Regular form submission - redirect back
            messages.success(request, f'Task status updated to "{task.get_status_display()}".')
            return redirect('task_detail', pk=task.pk)

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': False})

    return redirect('task_detail', pk=pk)


# ============== Time Entry Views ==============

@login_required
def timeentry_list(request):
    """List all time entries (filtered for team members)"""
    entries = TimeEntry.objects.select_related('project', 'task', 'user').all()

    # Team members only see their own time entries
    team_member = getattr(request.user, 'team_profile', None)
    if team_member:
        entries = entries.filter(user=request.user)

    search = request.GET.get('search', '')
    project_id = request.GET.get('project', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if search:
        entries = entries.filter(description__icontains=search)
    if project_id:
        entries = entries.filter(project_id=project_id)
    if date_from:
        entries = entries.filter(date__gte=date_from)
    if date_to:
        entries = entries.filter(date__lte=date_to)

    total_hours = entries.aggregate(total=Sum('hours'))['total'] or 0
    billable_hours = entries.filter(is_billable=True).aggregate(total=Sum('hours'))['total'] or 0

    # Filter projects for team members
    if team_member:
        projects = Project.objects.filter(tasks__assigned_to=team_member).distinct()
    else:
        projects = Project.objects.filter(status__in=['in_progress', 'confirmed'])

    context = {
        'entries': entries,
        'projects': projects,
        'search': search,
        'selected_project': project_id,
        'date_from': date_from,
        'date_to': date_to,
        'total_hours': total_hours,
        'billable_hours': billable_hours,
        'is_team_member': team_member is not None,
    }
    return render(request, 'time/list.html', context)


@login_required
def timeentry_create(request):
    """Create a time entry"""
    if request.method == 'POST':
        entry = TimeEntry(
            description=request.POST.get('description'),
            hours=request.POST.get('hours'),
            date=request.POST.get('date') or timezone.now().date(),
            is_billable=request.POST.get('is_billable') == 'on',
            notes=request.POST.get('notes', ''),
            user=request.user,
        )

        project_id = request.POST.get('project')
        if project_id:
            entry.project = Project.objects.get(pk=project_id)

        task_id = request.POST.get('task')
        if task_id:
            entry.task = Task.objects.get(pk=task_id)

        hourly_rate = request.POST.get('hourly_rate')
        if hourly_rate:
            entry.hourly_rate = hourly_rate

        entry.save()

        log_activity(request, 'created', entry)

        messages.success(request, 'Time entry created successfully.')
        return redirect('timeentry_list')

    # For team members, only show their assigned projects
    team_member = getattr(request.user, 'team_profile', None)
    if team_member:
        projects = team_member.assigned_projects.filter(status__in=['in_progress', 'confirmed'])
        tasks = Task.objects.filter(assigned_to=team_member).exclude(status='completed')
    else:
        projects = Project.objects.filter(status__in=['in_progress', 'confirmed'])
        tasks = Task.objects.exclude(status='completed')

    context = {
        'projects': projects,
        'tasks': tasks,
    }
    return render(request, 'time/form.html', context)


@login_required
def timeentry_update(request, pk):
    """Update a time entry"""
    entry = get_object_or_404(TimeEntry, pk=pk)

    if request.method == 'POST':
        entry.description = request.POST.get('description')
        entry.hours = request.POST.get('hours')
        entry.date = request.POST.get('date')
        entry.is_billable = request.POST.get('is_billable') == 'on'
        entry.notes = request.POST.get('notes', '')

        project_id = request.POST.get('project')
        entry.project = Project.objects.get(pk=project_id) if project_id else None

        task_id = request.POST.get('task')
        entry.task = Task.objects.get(pk=task_id) if task_id else None

        hourly_rate = request.POST.get('hourly_rate')
        entry.hourly_rate = hourly_rate if hourly_rate else None

        entry.save()

        log_activity(request, 'updated', entry)

        messages.success(request, 'Time entry updated successfully.')
        return redirect('timeentry_list')

    context = {
        'entry': entry,
        'projects': Project.objects.filter(status__in=['in_progress', 'confirmed']),
        'tasks': Task.objects.exclude(status='completed'),
    }
    return render(request, 'time/form.html', context)


@login_required
def timeentry_delete(request, pk):
    """Delete a time entry"""
    entry = get_object_or_404(TimeEntry, pk=pk)

    if request.method == 'POST':
        log_activity(request, 'deleted', entry)
        entry.delete()
        messages.success(request, 'Time entry deleted successfully.')
        return redirect('timeentry_list')

    return render(request, 'time/delete.html', {'entry': entry})


# ============== Activity Log Views ==============

@login_required
def activity_log(request):
    """View activity log"""
    logs = ActivityLog.objects.select_related('user').all()

    action = request.GET.get('action', '')
    model = request.GET.get('model', '')

    if action:
        logs = logs.filter(action=action)
    if model:
        logs = logs.filter(model_name=model)

    # Get unique model names for filter
    model_names = ActivityLog.objects.values_list('model_name', flat=True).distinct()

    context = {
        'logs': logs[:100],  # Limit to 100 most recent
        'action_choices': ActivityLog.ACTION_CHOICES,
        'model_names': model_names,
        'selected_action': action,
        'selected_model': model,
    }
    return render(request, 'activity/list.html', context)


# ============== Document Views ==============

@login_required
def document_upload(request):
    """Upload a document attachment"""
    if request.method == 'POST':
        file = request.FILES.get('file')
        if not file:
            messages.error(request, 'Please select a file.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        model_type = request.POST.get('model_type')
        object_id = request.POST.get('object_id')

        # Get content type
        model_map = {
            'client': Client,
            'project': Project,
            'invoice': Invoice,
            'quote': Quote,
        }

        if model_type not in model_map:
            messages.error(request, 'Invalid model type.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        model_class = model_map[model_type]
        content_type = ContentType.objects.get_for_model(model_class)

        document = Document(
            file=file,
            name=request.POST.get('name') or file.name,
            description=request.POST.get('description', ''),
            uploaded_by=request.user,
            content_type=content_type,
            object_id=object_id,
        )
        document.save()

        log_activity(request, 'created', document)

        messages.success(request, 'Document uploaded successfully.')
        return redirect(request.META.get('HTTP_REFERER', '/'))

    return redirect('/')


@login_required
def document_download(request, pk):
    """Download a document"""
    from django.http import FileResponse

    document = get_object_or_404(Document, pk=pk)

    try:
        response = FileResponse(document.file.open('rb'), as_attachment=True, filename=document.name)
        return response
    except FileNotFoundError:
        messages.error(request, 'File not found.')
        return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def document_delete(request, pk):
    """Delete a document"""
    document = get_object_or_404(Document, pk=pk)

    if request.method == 'POST':
        log_activity(request, 'deleted', document)
        document.file.delete()
        document.delete()
        messages.success(request, 'Document deleted successfully.')

    return redirect(request.META.get('HTTP_REFERER', '/'))


# ============== Email Views ==============

@login_required
def send_invoice_email(request, pk):
    """Send invoice via email"""
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string

    invoice = get_object_or_404(Invoice.objects.select_related('client'), pk=pk)
    company = CompanySettings.get_settings()

    if request.method == 'POST':
        to_email = request.POST.get('to_email', invoice.client.email)
        subject = request.POST.get('subject', f'Invoice {invoice.invoice_number} from {company.company_name}')
        message = request.POST.get('message', '')

        # Check if email settings are configured
        if not company.smtp_host or not company.smtp_user:
            messages.error(request, 'Email settings not configured. Please configure SMTP settings first.')
            return redirect('invoice_detail', pk=pk)

        try:
            # Configure email backend dynamically
            from django.core.mail import get_connection

            connection = get_connection(
                host=company.smtp_host,
                port=company.smtp_port,
                username=company.smtp_user,
                password=company.smtp_password,
                use_tls=company.smtp_use_tls,
            )

            # Generate PDF
            html_content = render_to_string('invoices/pdf.html', {
                'invoice': invoice,
                'company': company,
                'include_gst': True,
            })

            email = EmailMessage(
                subject=subject,
                body=message or f'Please find attached invoice {invoice.invoice_number}.',
                from_email=company.from_email or company.smtp_user,
                to=[to_email],
                connection=connection,
            )

            # Try to attach PDF if WeasyPrint is available
            try:
                from weasyprint import HTML
                pdf = HTML(string=html_content).write_pdf()
                email.attach(f'{invoice.invoice_number}.pdf', pdf, 'application/pdf')
            except ImportError:
                pass  # Send without attachment

            email.send()

            # Update invoice status
            if invoice.status == 'draft':
                invoice.status = 'sent'
                invoice.save()

            log_activity(request, 'sent', invoice)

            messages.success(request, f'Invoice sent to {to_email}')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')

        return redirect('invoice_detail', pk=pk)

    context = {
        'invoice': invoice,
        'company': company,
        'default_subject': f'Invoice {invoice.invoice_number} from {company.company_name}',
        'default_message': f'Dear {invoice.client.name},\n\nPlease find attached invoice {invoice.invoice_number} for {invoice.title}.\n\nAmount Due: ₹{invoice.balance_due}\nDue Date: {invoice.due_date.strftime("%d %b %Y") if invoice.due_date else "N/A"}\n\nThank you for your business.\n\nBest regards,\n{company.company_name}',
    }
    return render(request, 'emails/send_invoice.html', context)


@login_required
def send_quote_email(request, pk):
    """Send quote via email"""
    from django.core.mail import EmailMessage
    from django.template.loader import render_to_string

    quote = get_object_or_404(Quote.objects.select_related('client'), pk=pk)
    company = CompanySettings.get_settings()

    if request.method == 'POST':
        to_email = request.POST.get('to_email', quote.recipient_email)
        subject = request.POST.get('subject', f'Quote {quote.quote_number} from {company.company_name}')
        message = request.POST.get('message', '')

        if not company.smtp_host or not company.smtp_user:
            messages.error(request, 'Email settings not configured. Please configure SMTP settings first.')
            return redirect('quote_detail', pk=pk)

        try:
            from django.core.mail import get_connection

            connection = get_connection(
                host=company.smtp_host,
                port=company.smtp_port,
                username=company.smtp_user,
                password=company.smtp_password,
                use_tls=company.smtp_use_tls,
            )

            html_content = render_to_string('quotes/pdf.html', {
                'quote': quote,
                'company': company,
                'include_gst': True,
            })

            email = EmailMessage(
                subject=subject,
                body=message or f'Please find attached quote {quote.quote_number}.',
                from_email=company.from_email or company.smtp_user,
                to=[to_email],
                connection=connection,
            )

            try:
                from weasyprint import HTML
                pdf = HTML(string=html_content).write_pdf()
                email.attach(f'{quote.quote_number}.pdf', pdf, 'application/pdf')
            except ImportError:
                pass

            email.send()

            if quote.status == 'draft':
                quote.status = 'sent'
                quote.save()

            log_activity(request, 'sent', quote)

            messages.success(request, f'Quote sent to {to_email}')
        except Exception as e:
            messages.error(request, f'Failed to send email: {str(e)}')

        return redirect('quote_detail', pk=pk)

    context = {
        'quote': quote,
        'company': company,
        'default_subject': f'Quote {quote.quote_number} from {company.company_name}',
        'default_message': f'Dear {quote.recipient_name},\n\nPlease find attached quote {quote.quote_number} for {quote.title}.\n\nTotal Amount: ₹{quote.total_amount}\nValid Until: {quote.valid_until.strftime("%d %b %Y") if quote.valid_until else "N/A"}\n\nPlease let us know if you have any questions.\n\nBest regards,\n{company.company_name}',
    }
    return render(request, 'emails/send_quote.html', context)


# ============== Helper Functions ==============

def log_activity(request, action, instance):
    """Log an activity"""
    try:
        ActivityLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action=action,
            model_name=instance.__class__.__name__,
            object_id=str(instance.pk),
            object_repr=str(instance)[:255],
            ip_address=get_client_ip(request),
        )
    except Exception:
        pass  # Don't fail if logging fails


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        return x_forwarded_for.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR')


# ============== License Management Views ==============

@login_required
def license_list(request):
    """List all licenses"""
    licenses = License.objects.select_related('key_pair', 'client').order_by('-created_at')

    # Filters
    status = request.GET.get('status', '')
    search = request.GET.get('search', '')

    if status:
        licenses = licenses.filter(status=status)
    if search:
        licenses = licenses.filter(
            Q(license_code__icontains=search) |
            Q(customer_name__icontains=search) |
            Q(customer_email__icontains=search)
        )

    context = {
        'licenses': licenses,
        'status': status,
        'search': search,
        'total_count': License.objects.count(),
        'active_count': License.objects.filter(status='active').count(),
        'expired_count': License.objects.filter(status='expired').count(),
    }
    return render(request, 'licenses/list.html', context)


@login_required
def license_create(request):
    """Create a new license"""
    if request.method == 'POST':
        customer_name = request.POST.get('customer_name')
        customer_email = request.POST.get('customer_email', '')
        customer_company = request.POST.get('customer_company', '')
        valid_until = request.POST.get('valid_until')
        license_type = request.POST.get('license_type', 'basic')
        max_activations = int(request.POST.get('max_activations', 1))
        notes = request.POST.get('notes', '')
        client_id = request.POST.get('client', '')

        # Get key pair (selected or first active)
        key_pair_id = request.POST.get('key_pair', '')
        if key_pair_id:
            key_pair = LicenseKey.objects.filter(pk=key_pair_id, is_active=True).first()
        else:
            key_pair = LicenseKey.objects.filter(is_active=True).first()
        if not key_pair:
            messages.error(request, 'No active license key found. Please generate keys first.')
            return redirect('license_list')

        from datetime import datetime
        from django.utils import timezone
        valid_until_date = timezone.make_aware(datetime.strptime(valid_until, '%Y-%m-%d')) if valid_until else None

        # Get client if specified
        client = None
        if client_id:
            try:
                client = Client.objects.get(pk=client_id)
            except Client.DoesNotExist:
                pass

        license = License.objects.create(
            key_pair=key_pair,
            client=client,
            customer_name=customer_name,
            customer_email=customer_email or f"{customer_name.lower().replace(' ', '.')}@example.com",
            customer_company=customer_company,
            license_type=license_type,
            valid_until=valid_until_date,
            max_activations=max_activations,
            notes=notes,
        )

        messages.success(request, f'License created successfully!')
        return redirect('license_detail', pk=license.pk)

    from datetime import datetime, timedelta
    default_expiry = datetime.now() + timedelta(days=365)

    # Check if client is pre-selected from query param
    preselected_client = None
    client_id = request.GET.get('client', '')
    if client_id:
        try:
            preselected_client = Client.objects.get(pk=client_id)
        except Client.DoesNotExist:
            pass

    context = {
        'default_expiry': default_expiry.strftime('%Y-%m-%d'),
        'license_types': License.LICENSE_TYPE_CHOICES,
        'clients': Client.objects.filter(is_active=True).order_by('company_name', 'name'),
        'preselected_client': preselected_client,
        'key_pairs': LicenseKey.objects.filter(is_active=True).order_by('name'),
    }
    return render(request, 'licenses/form.html', context)


@login_required
def license_detail(request, pk):
    """View license details"""
    license = get_object_or_404(License, pk=pk)
    activations = license.activations.order_by('-activated_at')
    
    context = {
        'license': license,
        'activations': activations,
    }
    return render(request, 'licenses/detail.html', context)


@login_required
def license_deactivate_device(request, pk, activation_id):
    """Deactivate a specific device activation"""
    license = get_object_or_404(License, pk=pk)
    activation = get_object_or_404(LicenseActivation, pk=activation_id, license=license)

    if request.method == 'POST':
        # Deactivate the activation
        activation.is_active = False
        activation.save()

        # Update the license activation count
        license.current_activations = license.activations.filter(is_active=True).count()
        license.save(update_fields=['current_activations'])

        messages.success(request, f'Device "{activation.machine_name or "Unknown Device"}" has been deactivated.')
        return redirect('license_detail', pk=pk)

    return redirect('license_detail', pk=pk)


@login_required
def license_delete_activation(request, pk, activation_id):
    """Delete a device activation completely"""
    license = get_object_or_404(License, pk=pk)
    activation = get_object_or_404(LicenseActivation, pk=activation_id, license=license)

    if request.method == 'POST':
        device_name = activation.machine_name or "Unknown Device"
        activation.delete()

        # Update the license activation count
        license.current_activations = license.activations.filter(is_active=True).count()
        license.save(update_fields=['current_activations'])

        messages.success(request, f'Device "{device_name}" has been removed.')
        return redirect('license_detail', pk=pk)

    return redirect('license_detail', pk=pk)


@login_required
def license_update(request, pk):
    """Update license details including validity dates"""
    license = get_object_or_404(License, pk=pk)

    if request.method == 'POST':
        from datetime import datetime

        # Get form data
        valid_until_str = request.POST.get('valid_until', '')
        status = request.POST.get('status', license.status)
        billing_cycle = request.POST.get('billing_cycle', license.billing_cycle)
        max_activations = request.POST.get('max_activations', license.max_activations)
        notes = request.POST.get('notes', license.notes)

        # Track if validity was extended (for renewal tracking)
        old_valid_until = license.valid_until

        # Update validity date
        if valid_until_str:
            try:
                new_valid_until = datetime.strptime(valid_until_str, '%Y-%m-%d')
                from django.utils import timezone
                if timezone.is_naive(new_valid_until):
                    new_valid_until = timezone.make_aware(new_valid_until)

                # Check if this is a renewal (extension)
                if new_valid_until > old_valid_until:
                    license.renewal_count += 1
                    license.last_renewed_at = timezone.now()
                    # Add renewal note
                    renewal_note = f"\n[{timezone.now().strftime('%Y-%m-%d %H:%M')}] Renewed from {old_valid_until.strftime('%Y-%m-%d')} to {new_valid_until.strftime('%Y-%m-%d')}"
                    license.notes = (license.notes or '') + renewal_note

                license.valid_until = new_valid_until
            except ValueError:
                messages.error(request, 'Invalid date format.')
                return redirect('license_detail', pk=pk)

        # Update other fields
        license.status = status
        license.billing_cycle = billing_cycle
        try:
            license.max_activations = int(max_activations)
        except (ValueError, TypeError):
            pass

        # Regenerate license code if validity changed
        if license.valid_until != old_valid_until:
            license.license_code = license.generate_license_code()

        license.save()
        messages.success(request, 'License updated successfully!')
        return redirect('license_detail', pk=pk)

    return redirect('license_detail', pk=pk)


@login_required
def sync_licenses(request):
    """
    Sync all licenses - check validity and update expired status.
    Returns JSON with sync results for AJAX calls.
    """
    from licensing.models import License
    from django.http import JsonResponse

    if request.method == 'POST':
        now = timezone.now()
        updated_count = 0
        expired_count = 0
        active_count = 0
        expiring_soon_count = 0

        # Get all licenses
        licenses = License.objects.all()

        for license in licenses:
            original_status = license.status

            # Check if license should be expired
            if license.valid_until < now and license.status == 'active':
                license.status = 'expired'
                license.save(update_fields=['status'])
                updated_count += 1
                expired_count += 1
            elif license.valid_until >= now and license.status == 'expired':
                # License was renewed - reactivate it
                license.status = 'active'
                license.save(update_fields=['status'])
                updated_count += 1
                active_count += 1
            elif license.status == 'active':
                active_count += 1
                # Check if expiring soon (within 30 days)
                if license.valid_until <= now + timedelta(days=30):
                    expiring_soon_count += 1
            else:
                if license.status == 'expired':
                    expired_count += 1

        # Return JSON for AJAX
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': f'Sync complete! {updated_count} license(s) updated.',
                'stats': {
                    'total': licenses.count(),
                    'active': active_count,
                    'expired': expired_count,
                    'expiring_soon': expiring_soon_count,
                    'updated': updated_count,
                }
            })

        messages.success(request, f'License sync complete! {updated_count} license(s) updated.')
        return redirect('dashboard')

    return redirect('dashboard')


@login_required
def license_revoke(request, pk):
    """Revoke a license"""
    license = get_object_or_404(License, pk=pk)
    
    if request.method == 'POST':
        license.status = 'revoked'
        license.save()
        
        # Deactivate all activations
        license.activations.update(is_active=False)
        
        messages.success(request, f'License {license.license_code} has been revoked.')
        return redirect('license_list')
    
    return render(request, 'licenses/revoke.html', {'license': license})


@login_required
def license_generate_keys(request):
    """Generate new RSA key pair"""
    if request.method == 'POST':
        from cryptography.hazmat.primitives import serialization
        from cryptography.hazmat.primitives.asymmetric import rsa
        from cryptography.hazmat.backends import default_backend

        key_name = request.POST.get('key_name', '').strip() or 'Default'

        # Generate key pair
        private_key = rsa.generate_private_key(
            public_exponent=65537,
            key_size=4096,
            backend=default_backend()
        )

        # Serialize keys
        private_pem = private_key.private_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PrivateFormat.TraditionalOpenSSL,
            encryption_algorithm=serialization.NoEncryption()
        ).decode('utf-8')

        public_pem = private_key.public_key().public_bytes(
            encoding=serialization.Encoding.PEM,
            format=serialization.PublicFormat.SubjectPublicKeyInfo
        ).decode('utf-8')

        # Create new key (don't deactivate others — multiple products supported)
        key = LicenseKey.objects.create(
            name=key_name,
            private_key=private_pem,
            public_key=public_pem,
            is_active=True,
        )

        messages.success(request, f'New RSA key pair "{key_name}" generated successfully!')
        return redirect('license_keys')

    return redirect('license_keys')


@login_required
def license_keys(request):
    """View and manage license keys"""
    keys = LicenseKey.objects.order_by('-created_at')
    
    context = {
        'keys': keys,
    }
    return render(request, 'licenses/keys.html', context)


# ============================================================
# HR & Admin - Employee Management Web Views
# ============================================================

@login_required
def emp_employee_list(request):
    """List all employees"""
    from employees.models import Employee
    employees = Employee.objects.select_related('user').all()
    status_filter = request.GET.get('status', '')
    dept_filter = request.GET.get('department', '')
    if status_filter:
        employees = employees.filter(status=status_filter)
    if dept_filter:
        employees = employees.filter(department=dept_filter)

    # Count team members without employee profiles
    linked_user_ids = Employee.objects.values_list('user_id', flat=True)
    unlinked_count = TeamMember.objects.filter(user__isnull=False).exclude(user_id__in=linked_user_ids).count()

    context = {
        'employees': employees,
        'status_filter': status_filter,
        'dept_filter': dept_filter,
        'departments': Employee.DEPARTMENT_CHOICES,
        'statuses': Employee.STATUS_CHOICES,
        'unlinked_team_members': unlinked_count,
    }
    return render(request, 'hr/employee_list.html', context)


@login_required
def emp_employee_detail(request, pk):
    """Employee detail with attendance, leave, work"""
    from employees.models import Employee, Attendance, LeaveRequest, WorkAssignment
    from employees.utils import generate_face_encoding
    employee = get_object_or_404(Employee, pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'edit_info':
        # Update User fields
        employee.user.first_name = request.POST.get('first_name', employee.user.first_name)
        employee.user.last_name = request.POST.get('last_name', employee.user.last_name)
        employee.user.email = request.POST.get('email', employee.user.email)
        employee.user.save()

        # Update Employee fields
        employee.employment_type = request.POST.get('employment_type', employee.employment_type)
        employee.role = request.POST.get('role', employee.role)
        employee.department = request.POST.get('department', employee.department)
        employee.designation = request.POST.get('designation', employee.designation)
        employee.status = request.POST.get('status', employee.status)
        employee.phone = request.POST.get('phone', employee.phone)
        employee.emergency_contact = request.POST.get('emergency_contact', employee.emergency_contact)
        employee.address = request.POST.get('address', employee.address)
        employee.joining_date = request.POST.get('joining_date') or employee.joining_date
        employee.monthly_salary = request.POST.get('monthly_salary') or None
        employee.hourly_rate = request.POST.get('hourly_rate') or None
        employee.save()
        messages.success(request, 'Employee information updated.')
        return redirect('emp_employee_detail', pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'upload_profile_photo':
        profile_photo = request.FILES.get('profile_photo')
        if profile_photo:
            employee.profile_photo = profile_photo
            employee.save(update_fields=['profile_photo'])
            messages.success(request, 'Profile photo updated.')
        else:
            messages.error(request, 'Please select a photo.')
        return redirect('emp_employee_detail', pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'remove_profile_photo':
        employee.profile_photo = None
        employee.save(update_fields=['profile_photo'])
        messages.success(request, 'Profile photo removed.')
        return redirect('emp_employee_detail', pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'upload_face':
        face_photo = request.FILES.get('face_photo')
        if face_photo:
            employee.face_photo = face_photo
            employee.save()
            encoding = generate_face_encoding(employee.face_photo.path)
            if encoding:
                employee.face_encoding = encoding
                employee.save(update_fields=['face_encoding'])
                messages.success(request, 'Face photo uploaded and encoding generated successfully.')
            else:
                # Keep the photo but warn about encoding failure
                employee.face_encoding = None
                employee.save(update_fields=['face_encoding'])
                messages.warning(request,
                    'Face photo saved but no face was detected for encoding. '
                    'Try a different photo: front-facing, well-lit, clear face visible. '
                    f'Photo saved at: {employee.face_photo.name}'
                )
        else:
            messages.error(request, 'Please select a photo to upload.')
        return redirect('emp_employee_detail', pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'remove_face':
        employee.face_photo = None
        employee.face_encoding = None
        employee.save(update_fields=['face_photo', 'face_encoding'])
        messages.success(request, 'Face photo removed.')
        return redirect('emp_employee_detail', pk=pk)

    if request.method == 'POST' and request.POST.get('action') == 'mark_attendance':
        from datetime import datetime, time as dtime
        from django.utils.dateparse import parse_date
        att_date = parse_date(request.POST.get('att_date') or '') or timezone.localdate()
        att_status = request.POST.get('att_status', 'present')
        check_in_raw = request.POST.get('check_in', '').strip()
        check_out_raw = request.POST.get('check_out', '').strip()
        notes = request.POST.get('notes', '').strip()

        def _aware_dt(date_part, time_str):
            if not time_str:
                return None
            try:
                hh, mm = [int(x) for x in time_str.split(':')[:2]]
            except (ValueError, IndexError):
                return None
            naive = datetime.combine(date_part, dtime(hh, mm))
            return timezone.make_aware(naive, timezone.get_current_timezone())

        check_in_dt = _aware_dt(att_date, check_in_raw)
        check_out_dt = _aware_dt(att_date, check_out_raw)

        if check_out_dt and check_in_dt and check_out_dt < check_in_dt:
            messages.error(request, 'Check-out time cannot be earlier than check-in time.')
            return redirect('emp_employee_detail', pk=pk)

        worked_hours = None
        if check_in_dt and check_out_dt:
            worked_hours = round((check_out_dt - check_in_dt).total_seconds() / 3600, 2)

        defaults = {
            'status': att_status,
            'verification_method': 'manual',
            'check_in': check_in_dt,
            'check_out': check_out_dt,
            'worked_hours': worked_hours,
            'notes': notes,
        }
        att, created = Attendance.objects.update_or_create(
            employee=employee, date=att_date, defaults=defaults,
        )
        verb = 'recorded' if created else 'updated'
        messages.success(request, f'Attendance {verb} for {att_date:%d %b %Y}.')
        return redirect('emp_employee_detail', pk=pk)

    recent_attendance = Attendance.objects.filter(employee=employee).order_by('-date')[:15]
    leave_requests = LeaveRequest.objects.filter(employee=employee).order_by('-created_at')[:10]
    work_assignments = WorkAssignment.objects.filter(assigned_to=employee).order_by('-created_at')[:10]
    context = {
        'employee': employee,
        'recent_attendance': recent_attendance,
        'leave_requests': leave_requests,
        'work_assignments': work_assignments,
        'employment_types': Employee.EMPLOYMENT_TYPE_CHOICES,
        'roles': Employee.ROLE_CHOICES,
        'departments': Employee.DEPARTMENT_CHOICES,
        'statuses': Employee.STATUS_CHOICES,
        'attendance_statuses': Attendance.STATUS_CHOICES,
        'today': timezone.localdate(),
    }
    return render(request, 'hr/employee_detail.html', context)


@login_required
def emp_employee_delete(request, pk):
    """Delete an employee and their related data, keeping the User and TeamMember intact"""
    from employees.models import Employee, Attendance, LeaveRequest, WorkAssignment, WorkUpdate, Notification, DeviceToken
    if request.method != 'POST':
        return redirect('emp_employee_list')

    employee = get_object_or_404(Employee, pk=pk)
    name = employee.full_name

    # Delete employee-related data (keep User and TeamMember)
    DeviceToken.objects.filter(employee=employee).delete()
    Notification.objects.filter(employee=employee).delete()
    WorkUpdate.objects.filter(employee=employee).delete()
    Attendance.objects.filter(employee=employee).delete()
    LeaveRequest.objects.filter(employee=employee).delete()
    # Remove employee from assignments; delete assignments with no remaining assignees
    for wa in WorkAssignment.objects.filter(assigned_to=employee):
        wa.assigned_to.remove(employee)
        if not wa.assigned_to.exists():
            wa.delete()
    employee.delete()

    messages.success(request, f'Employee "{name}" and all related records deleted. User account and team member profile are preserved.')
    return redirect('emp_employee_list')


@login_required
def emp_employee_create(request):
    """Create a new employee"""
    from employees.models import Employee
    from django.contrib.auth.models import User

    if request.method == 'POST':
        # Create or get User
        user_id = request.POST.get('existing_user')
        if user_id:
            user = get_object_or_404(User, pk=user_id)
            if hasattr(user, 'employee_profile'):
                messages.error(request, 'This user already has an employee profile.')
                return redirect('emp_employee_create')
        else:
            username = request.POST.get('username', '').strip()
            if not username:
                messages.error(request, 'Username is required.')
                return redirect('emp_employee_create')
            if User.objects.filter(username=username).exists():
                messages.error(request, 'Username already exists.')
                return redirect('emp_employee_create')
            user = User.objects.create_user(
                username=username,
                password=request.POST.get('password', 'changeme123'),
                first_name=request.POST.get('first_name', ''),
                last_name=request.POST.get('last_name', ''),
                email=request.POST.get('email', ''),
            )

        # Generate employee ID
        last_emp = Employee.objects.order_by('-employee_id').first()
        if last_emp and last_emp.employee_id.startswith('EMP'):
            try:
                num = int(last_emp.employee_id[3:]) + 1
            except ValueError:
                num = 1
        else:
            num = 1
        employee_id = request.POST.get('employee_id') or f'EMP{num:03d}'

        Employee.objects.create(
            user=user,
            employee_id=employee_id,
            employment_type=request.POST.get('employment_type', 'fulltime'),
            role=request.POST.get('role', 'employee'),
            department=request.POST.get('department', 'engineering'),
            designation=request.POST.get('designation', ''),
            phone=request.POST.get('phone', ''),
            emergency_contact=request.POST.get('emergency_contact', ''),
            address=request.POST.get('address', ''),
            status='active',
            monthly_salary=request.POST.get('monthly_salary') or None,
            hourly_rate=request.POST.get('hourly_rate') or None,
        )
        messages.success(request, f'Employee created. Login: {user.username} / {request.POST.get("password", "changeme123")}')
        return redirect('emp_employee_list')

    # Get users who don't have employee profiles yet
    linked_user_ids = Employee.objects.values_list('user_id', flat=True)
    available_users = User.objects.exclude(pk__in=linked_user_ids).order_by('first_name', 'username')

    context = {
        'available_users': available_users,
        'employment_types': Employee.EMPLOYMENT_TYPE_CHOICES,
        'roles': Employee.ROLE_CHOICES,
        'departments': Employee.DEPARTMENT_CHOICES,
    }
    return render(request, 'hr/employee_create.html', context)


@login_required
def emp_employee_import(request):
    """Import all team members as employees"""
    from employees.models import Employee

    if request.method != 'POST':
        return redirect('emp_employee_list')

    linked_user_ids = Employee.objects.values_list('user_id', flat=True)
    unlinked = TeamMember.objects.filter(user__isnull=False).exclude(user_id__in=linked_user_ids)

    employment_type_map = {'permanent': 'fulltime', 'freelancer': 'parttime'}
    role_to_dept_map = {
        'developer': 'engineering', 'designer': 'design',
        'project_manager': 'operations', 'qa': 'engineering',
        'devops': 'engineering', 'other': 'other',
    }

    count = 0
    for tm in unlinked:
        last_emp = Employee.objects.order_by('-employee_id').first()
        if last_emp and last_emp.employee_id.startswith('EMP'):
            try:
                num = int(last_emp.employee_id[3:]) + 1
            except ValueError:
                num = 1
        else:
            num = 1

        Employee.objects.create(
            user=tm.user,
            employee_id=f'EMP{num:03d}',
            employment_type=employment_type_map.get(tm.employment_type, 'fulltime'),
            department=role_to_dept_map.get(tm.role, 'other'),
            designation=tm.get_role_display(),
            phone=tm.phone or '',
            monthly_salary=tm.monthly_salary,
            hourly_rate=tm.hourly_rate,
            status='active',
        )
        count += 1

    messages.success(request, f'{count} team member(s) imported as employees.')
    return redirect('emp_employee_list')


@login_required
def emp_leave_list(request):
    """List leave requests with filtering"""
    from employees.models import LeaveRequest
    leaves = LeaveRequest.objects.select_related('employee__user', 'leave_type').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        leaves = leaves.filter(status=status_filter)
    context = {
        'leaves': leaves,
        'status_filter': status_filter,
    }
    return render(request, 'hr/leave_list.html', context)


@login_required
def emp_leave_action(request, pk):
    """Approve or reject a leave request"""
    from employees.models import LeaveRequest, Notification
    from employees.utils import send_push_notification
    if request.method != 'POST':
        return redirect('emp_leave_list')
    leave = get_object_or_404(LeaveRequest, pk=pk)
    action = request.POST.get('action')
    if action in ('approve', 'reject'):
        leave.status = 'approved' if action == 'approve' else 'rejected'
        leave.reviewed_by = request.user
        leave.review_notes = request.POST.get('notes', '')
        leave.reviewed_at = timezone.now()
        leave.save()
        Notification.objects.create(
            employee=leave.employee,
            title=f'Leave {leave.get_status_display()}',
            body=f'Your leave from {leave.start_date} to {leave.end_date} has been {leave.get_status_display().lower()}.',
            notification_type='leave',
        )
        send_push_notification(
            leave.employee,
            f'Leave {leave.get_status_display()}',
            f'Your leave from {leave.start_date} to {leave.end_date} has been {leave.get_status_display().lower()}.',
        )
        messages.success(request, f'Leave request {leave.get_status_display().lower()}.')
    return redirect('emp_leave_list')


@login_required
def emp_attendance_list(request):
    """View attendance records"""
    from employees.models import Attendance, Employee, LateCheckInGrant
    from datetime import date
    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))
    employee_filter = request.GET.get('employee', '')
    records = Attendance.objects.select_related('employee__user').filter(
        date__month=month, date__year=year
    ).order_by('-date', '-check_in')
    if employee_filter:
        records = records.filter(employee__pk=employee_filter)
    employees = Employee.objects.filter(status='active').order_by('employee_id')
    pending_grants = LateCheckInGrant.objects.select_related('employee__user').filter(
        date=date.today(), consumed_at__isnull=True
    ).order_by('-created_at')
    context = {
        'records': records,
        'employees': employees,
        'month': month,
        'year': year,
        'employee_filter': employee_filter,
        'today': date.today(),
        'pending_grants': pending_grants,
    }
    return render(request, 'hr/attendance_list.html', context)


@login_required
def emp_late_checkin_grant(request):
    """Grant an employee permission to check in after the daily cutoff."""
    from employees.models import Employee, LateCheckInGrant, Notification
    from employees.utils import send_push_notification
    if request.method != 'POST':
        return redirect('emp_attendance_list')
    employee_id = request.POST.get('employee_id')
    grant_date = request.POST.get('date') or timezone.now().date().isoformat()
    reason = (request.POST.get('reason') or '').strip()
    if not employee_id or not reason:
        messages.error(request, 'Employee and reason are required.')
        return redirect('emp_attendance_list')
    employee = get_object_or_404(Employee, pk=employee_id)
    grant, created = LateCheckInGrant.objects.get_or_create(
        employee=employee, date=grant_date,
        defaults={'reason': reason, 'granted_by': request.user},
    )
    if not created:
        if grant.consumed_at:
            messages.warning(request, f'{employee.full_name} already used a late check-in on {grant_date}.')
            return redirect('emp_attendance_list')
        grant.reason = reason
        grant.granted_by = request.user
        grant.save(update_fields=['reason', 'granted_by'])
        messages.success(request, f'Updated late check-in for {employee.full_name} ({grant_date}).')
    else:
        messages.success(request, f'{employee.full_name} may now check in late on {grant_date}.')
    Notification.objects.create(
        employee=employee,
        title='Late check-in allowed',
        body=f'You may check in after the cutoff on {grant_date}. Reason: {reason}',
        notification_type='attendance',
    )
    send_push_notification(
        employee,
        'Late check-in allowed',
        f'You may check in after the cutoff on {grant_date}.',
    )
    return redirect('emp_attendance_list')


@login_required
def emp_late_checkin_revoke(request, pk):
    """Revoke an unused late check-in grant."""
    from employees.models import LateCheckInGrant
    if request.method != 'POST':
        return redirect('emp_attendance_list')
    grant = get_object_or_404(LateCheckInGrant, pk=pk)
    if grant.consumed_at:
        messages.warning(request, 'Grant already used; cannot revoke.')
    else:
        name = grant.employee.full_name
        grant.delete()
        messages.success(request, f'Revoked late check-in for {name}.')
    return redirect('emp_attendance_list')


@login_required
def emp_attendance_report(request):
    """Monthly attendance summary per employee, with Excel export."""
    from employees.models import Attendance, Employee
    from calendar import monthrange
    from datetime import date

    month = int(request.GET.get('month', timezone.now().month))
    year = int(request.GET.get('year', timezone.now().year))
    department = request.GET.get('department', '')

    employees = Employee.objects.select_related('user').filter(status='active')
    if department:
        employees = employees.filter(department=department)
    employees = employees.order_by('employee_id')

    records = Attendance.objects.filter(date__month=month, date__year=year)
    if department:
        records = records.filter(employee__department=department)

    # Bucket attendance by employee id for in-memory aggregation.
    by_employee = {}
    for rec in records:
        by_employee.setdefault(rec.employee_id, []).append(rec)

    working_days = monthrange(year, month)[1]

    rows = []
    totals = {'present': 0, 'late': 0, 'half_day': 0, 'work_from_home': 0,
              'absent': 0, 'hours': 0.0}
    for emp in employees:
        emp_records = by_employee.get(emp.id, [])
        counts = {'present': 0, 'late': 0, 'half_day': 0, 'work_from_home': 0, 'absent': 0}
        hours = 0.0
        for rec in emp_records:
            if rec.status in counts:
                counts[rec.status] += 1
            hours += float(rec.worked_hours or rec.working_hours or 0)
        present_equiv = counts['present'] + counts['late'] + counts['work_from_home'] + counts['half_day'] * 0.5
        rows.append({
            'employee': emp,
            'present': counts['present'],
            'late': counts['late'],
            'half_day': counts['half_day'],
            'work_from_home': counts['work_from_home'],
            'absent': counts['absent'],
            'days_present': round(present_equiv, 1),
            'hours': round(hours, 2),
        })
        for key in counts:
            totals[key] += counts[key]
        totals['hours'] += hours

    totals['hours'] = round(totals['hours'], 2)

    if request.GET.get('export') == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Attendance'
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='0D9488', end_color='0D9488', fill_type='solid')
        headers = ['Employee ID', 'Name', 'Department', 'Present', 'Late',
                   'Half Day', 'WFH', 'Absent', 'Days Present', 'Total Hours']
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
        for r in rows:
            emp = r['employee']
            ws.append([
                emp.employee_id, emp.full_name, emp.get_department_display(),
                r['present'], r['late'], r['half_day'], r['work_from_home'],
                r['absent'], r['days_present'], r['hours'],
            ])
        for col_cells in ws.columns:
            length = max((len(str(cell.value or '')) for cell in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 12), 40)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = (
            f'attachment; filename="attendance_report_{year}_{month:02d}.xlsx"'
        )
        wb.save(response)
        return response

    context = {
        'rows': rows,
        'totals': totals,
        'month': month,
        'year': year,
        'working_days': working_days,
        'department': department,
        'departments': Employee.DEPARTMENT_CHOICES,
        'months': [(i, date(2000, i, 1).strftime('%B')) for i in range(1, 13)],
        'years': range(timezone.now().year - 3, timezone.now().year + 1),
    }
    return render(request, 'hr/attendance_report.html', context)


@login_required
def emp_work_list(request):
    """List work assignments"""
    from employees.models import WorkAssignment, Employee
    assignments = WorkAssignment.objects.prefetch_related('assigned_to').select_related('assigned_by').order_by('-created_at')
    status_filter = request.GET.get('status', '')
    if status_filter:
        assignments = assignments.filter(status=status_filter)
    employees = Employee.objects.filter(status='active').order_by('employee_id')
    context = {
        'assignments': assignments,
        'employees': employees,
        'status_filter': status_filter,
        'status_choices': WorkAssignment.STATUS_CHOICES,
        'priority_choices': WorkAssignment.PRIORITY_CHOICES,
    }
    return render(request, 'hr/work_list.html', context)


@login_required
def emp_work_create(request):
    """Create a new work assignment"""
    from employees.models import WorkAssignment, Employee, Notification
    from employees.utils import send_push_notification
    if request.method == 'POST':
        employee_ids = request.POST.getlist('employees')
        employees = Employee.objects.filter(pk__in=employee_ids)
        if not employees.exists():
            messages.error(request, 'Please select at least one employee.')
            return redirect('emp_work_create')
        assignment = WorkAssignment.objects.create(
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            assigned_by=request.user,
            priority=request.POST.get('priority', 'medium'),
            due_date=request.POST.get('due_date') or None,
            attachment=request.FILES.get('attachment'),
            confidentiality_disclaimer=request.POST.get('confidentiality_disclaimer', ''),
        )
        assignment.assigned_to.set(employees)
        for employee in employees:
            Notification.objects.create(
                employee=employee,
                title='New Work Assignment',
                body=f'You have been assigned: {assignment.title}',
                notification_type='work',
            )
            send_push_notification(
                employee,
                'New Work Assignment',
                f'You have been assigned: {assignment.title}',
            )
        messages.success(request, 'Work assignment created and employees notified.')
        return redirect('emp_work_list')
    from employees.models import Employee
    employees = Employee.objects.filter(status='active').order_by('employee_id')
    projects = Project.objects.filter(status='active')
    context = {
        'employees': employees,
        'projects': projects,
        'priority_choices': WorkAssignment.PRIORITY_CHOICES,
    }
    return render(request, 'hr/work_create.html', context)


@login_required
def emp_work_detail(request, pk):
    """View and edit a work assignment"""
    from employees.models import WorkAssignment, Employee, WorkUpdate
    assignment = get_object_or_404(WorkAssignment, pk=pk)

    if request.method == 'POST':
        assignment.title = request.POST.get('title', assignment.title)
        assignment.description = request.POST.get('description', assignment.description)
        assignment.priority = request.POST.get('priority', assignment.priority)
        assignment.status = request.POST.get('status', assignment.status)
        assignment.confidentiality_disclaimer = request.POST.get('confidentiality_disclaimer', assignment.confidentiality_disclaimer)
        assignment.due_date = request.POST.get('due_date') or None

        employee_ids = request.POST.getlist('employees')
        if employee_ids:
            employees = Employee.objects.filter(pk__in=employee_ids)
            assignment.assigned_to.set(employees)

        if request.FILES.get('attachment'):
            assignment.attachment = request.FILES['attachment']
        elif request.POST.get('remove_attachment'):
            assignment.attachment = None

        if assignment.status == 'completed' and not assignment.completed_at:
            from django.utils import timezone
            assignment.completed_at = timezone.now()

        assignment.save()

        # Notify all assigned employees about the update
        from employees.models import Notification
        from employees.utils import send_push_notification
        for emp in assignment.assigned_to.all():
            Notification.objects.create(
                employee=emp,
                title='Work Assignment Updated',
                body=f'Your assignment "{assignment.title}" has been updated.',
                notification_type='work',
                data={'assignment_id': str(assignment.id)},
            )
            send_push_notification(
                emp,
                'Work Assignment Updated',
                f'Your assignment "{assignment.title}" has been updated.',
            )

        messages.success(request, 'Work assignment updated and employee notified.')
        return redirect('emp_work_detail', pk=pk)

    employees = Employee.objects.filter(status='active').order_by('employee_id')
    assigned_employees = list(assignment.assigned_to.all())
    updates = WorkUpdate.objects.filter(assignment=assignment).select_related('employee__user').order_by('-created_at')
    context = {
        'assignment': assignment,
        'employees': employees,
        'assigned_employees': assigned_employees,
        'updates': updates,
        'status_choices': WorkAssignment.STATUS_CHOICES,
        'priority_choices': WorkAssignment.PRIORITY_CHOICES,
    }
    return render(request, 'hr/work_detail.html', context)


@login_required
def emp_work_delete(request, pk):
    """Delete a work assignment"""
    from employees.models import WorkAssignment
    assignment = get_object_or_404(WorkAssignment, pk=pk)
    if request.method == 'POST':
        assignment.delete()
        messages.success(request, 'Work assignment deleted.')
    return redirect('emp_work_list')


@login_required
def emp_leave_types(request):
    """Manage leave types"""
    from employees.models import LeaveType
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create':
            LeaveType.objects.create(
                name=request.POST.get('name'),
                days_allowed=int(request.POST.get('days_allowed', 12)),
                is_paid=request.POST.get('is_paid') == 'on',
                is_active=True,
            )
            messages.success(request, 'Leave type created.')
        elif action == 'toggle':
            lt = get_object_or_404(LeaveType, pk=request.POST.get('pk'))
            lt.is_active = not lt.is_active
            lt.save()
            messages.success(request, f'Leave type {"activated" if lt.is_active else "deactivated"}.')
        elif action == 'delete':
            lt = get_object_or_404(LeaveType, pk=request.POST.get('pk'))
            lt.delete()
            messages.success(request, 'Leave type deleted.')
        return redirect('emp_leave_types')
    leave_types = LeaveType.objects.all().order_by('name')
    return render(request, 'hr/leave_types.html', {'leave_types': leave_types})


@login_required
def emp_office_qr(request):
    """Generate and manage office QR code sticker"""
    import qrcode
    import io
    import base64
    from employees.models import OfficeConfig

    config = OfficeConfig.objects.first()

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'generate':
            office_name = request.POST.get('office_name', 'Main Office')
            latitude = request.POST.get('latitude') or None
            longitude = request.POST.get('longitude') or None
            # Generate a unique QR code value
            qr_value = f"RALFIZ-OFFICE-{uuid.uuid4().hex[:12].upper()}"
            if config:
                config.qr_code = qr_value
                config.office_name = office_name
                config.latitude = latitude
                config.longitude = longitude
                config.save()
            else:
                config = OfficeConfig.objects.create(
                    qr_code=qr_value,
                    office_name=office_name,
                    latitude=latitude,
                    longitude=longitude,
                )
            messages.success(request, 'Office QR code generated! Print the QR code and place it in the office.')
            return redirect('emp_office_qr')
        elif action == 'delete' and config:
            config.delete()
            messages.success(request, 'Office QR code deleted.')
            return redirect('emp_office_qr')

    # Generate QR image for display
    qr_image_b64 = None
    if config:
        qr = qrcode.QRCode(version=1, box_size=10, border=4)
        qr.add_data(config.qr_code)
        qr.make(fit=True)
        img = qr.make_image(fill_color='black', back_color='white')
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        qr_image_b64 = base64.b64encode(buffer.getvalue()).decode()

    return render(request, 'hr/office_qr.html', {
        'config': config,
        'qr_image': qr_image_b64,
    })


@login_required
def emp_class_list(request):
    """List and manage scheduled classes for interns"""
    from employees.models import ScheduledClass, Employee, Notification
    from employees.utils import send_push_notification

    status_filter = request.GET.get('status', '')
    classes = ScheduledClass.objects.select_related('created_by').order_by('-date', '-start_time')
    if status_filter:
        classes = classes.filter(status=status_filter)

    context = {
        'classes': classes,
        'status_filter': status_filter,
        'status_choices': ScheduledClass.STATUS_CHOICES,
    }
    return render(request, 'hr/class_list.html', context)


@login_required
def emp_class_create(request):
    """Create a new scheduled class"""
    from employees.models import ScheduledClass, Employee, Notification
    from employees.utils import send_push_notification

    if request.method == 'POST':
        scheduled_class = ScheduledClass.objects.create(
            title=request.POST.get('title'),
            description=request.POST.get('description', ''),
            date=request.POST.get('date'),
            start_time=request.POST.get('start_time'),
            end_time=request.POST.get('end_time'),
            instructor=request.POST.get('instructor', ''),
            location=request.POST.get('location', ''),
            notes=request.POST.get('notes', ''),
            attachment=request.FILES.get('attachment'),
            created_by=request.user,
        )

        scheduled_class.refresh_from_db()

        intern_ids = request.POST.getlist('interns')
        if intern_ids:
            scheduled_class.interns.set(intern_ids)
            interns = Employee.objects.filter(pk__in=intern_ids)
        else:
            interns = Employee.objects.filter(employment_type='intern', status='active')

        for intern in interns:
            Notification.objects.create(
                employee=intern,
                title='New Class Scheduled',
                body=f'{scheduled_class.title} on {scheduled_class.date} at {scheduled_class.start_time.strftime("%I:%M %p")}',
                notification_type='general',
                data={'class_id': str(scheduled_class.id)},
            )
            send_push_notification(intern, 'New Class Scheduled', f'{scheduled_class.title} on {scheduled_class.date}')

        messages.success(request, 'Class scheduled and interns notified.')
        return redirect('emp_class_list')

    interns = Employee.objects.filter(employment_type='intern', status='active').order_by('employee_id')
    context = {
        'interns': interns,
    }
    return render(request, 'hr/class_create.html', context)


@login_required
def emp_class_detail(request, pk):
    """View and edit a scheduled class"""
    from employees.models import ScheduledClass, Employee, Notification
    from employees.utils import send_push_notification

    scheduled_class = get_object_or_404(ScheduledClass, pk=pk)

    if request.method == 'POST':
        scheduled_class.title = request.POST.get('title', scheduled_class.title)
        scheduled_class.description = request.POST.get('description', scheduled_class.description)
        scheduled_class.date = request.POST.get('date', scheduled_class.date)
        scheduled_class.start_time = request.POST.get('start_time', scheduled_class.start_time)
        scheduled_class.end_time = request.POST.get('end_time', scheduled_class.end_time)
        scheduled_class.instructor = request.POST.get('instructor', scheduled_class.instructor)
        scheduled_class.location = request.POST.get('location', scheduled_class.location)
        scheduled_class.status = request.POST.get('status', scheduled_class.status)
        scheduled_class.notes = request.POST.get('notes', scheduled_class.notes)

        if request.FILES.get('attachment'):
            scheduled_class.attachment = request.FILES['attachment']
        elif request.POST.get('remove_attachment'):
            scheduled_class.attachment = None

        scheduled_class.save()

        intern_ids = request.POST.getlist('interns')
        scheduled_class.interns.set(intern_ids)

        # Notify assigned interns about update
        if intern_ids:
            interns = Employee.objects.filter(pk__in=intern_ids)
        else:
            interns = Employee.objects.filter(employment_type='intern', status='active')

        for intern in interns:
            Notification.objects.create(
                employee=intern,
                title='Class Updated',
                body=f'"{scheduled_class.title}" on {scheduled_class.date} has been updated.',
                notification_type='general',
                data={'class_id': str(scheduled_class.id)},
            )
            send_push_notification(intern, 'Class Updated', f'"{scheduled_class.title}" has been updated.')

        messages.success(request, 'Class updated and interns notified.')
        return redirect('emp_class_detail', pk=pk)

    interns = Employee.objects.filter(employment_type='intern', status='active').order_by('employee_id')
    assigned_intern_ids = list(scheduled_class.interns.values_list('pk', flat=True))
    context = {
        'scheduled_class': scheduled_class,
        'interns': interns,
        'assigned_intern_ids': assigned_intern_ids,
        'status_choices': ScheduledClass.STATUS_CHOICES,
    }
    return render(request, 'hr/class_detail.html', context)


@login_required
def emp_class_delete(request, pk):
    """Delete a scheduled class"""
    from employees.models import ScheduledClass
    if request.method == 'POST':
        scheduled_class = get_object_or_404(ScheduledClass, pk=pk)
        scheduled_class.delete()
        messages.success(request, 'Class deleted.')
    return redirect('emp_class_list')


@login_required
def emp_payroll_list(request):
    """List payroll records, generate payroll"""
    from employees.models import Payroll, Employee, Attendance, LeaveRequest
    from employees.utils import send_push_notification
    import calendar

    now = timezone.now()
    month = int(request.GET.get('month', now.month))
    year = int(request.GET.get('year', now.year))

    # Generate payroll
    if request.method == 'POST' and request.POST.get('action') == 'generate':
        working_days = int(request.POST.get('working_days', 26))
        employees = Employee.objects.filter(status='active')
        count = 0

        for emp in employees:
            if Payroll.objects.filter(employee=emp, month=month, year=year).exists():
                continue

            base = emp.monthly_salary or 0
            attendance = Attendance.objects.filter(employee=emp, date__month=month, date__year=year)
            days_present = attendance.filter(status__in=['present', 'late', 'work_from_home']).count()
            half_days = attendance.filter(status='half_day').count()
            days_present += half_days * 0.5

            approved_leaves = LeaveRequest.objects.filter(
                employee=emp, status='approved',
                start_date__month=month, start_date__year=year,
            )
            paid_leave_days = 0
            unpaid_leave_days = 0
            for lr in approved_leaves:
                if lr.leave_type and lr.leave_type.is_paid:
                    paid_leave_days += lr.total_days
                else:
                    unpaid_leave_days += lr.total_days

            days_absent = max(0, working_days - int(days_present) - paid_leave_days - unpaid_leave_days)

            payroll = Payroll(
                employee=emp, month=month, year=year,
                base_salary=base, working_days=working_days,
                days_present=int(days_present), days_absent=days_absent,
                paid_leave_days=paid_leave_days, unpaid_leave_days=unpaid_leave_days,
                generated_by=request.user,
            )
            payroll.calculate()
            payroll.save()
            count += 1

        messages.success(request, f'Payroll generated for {count} employee(s).')
        return redirect(f'/hr/payroll/?month={month}&year={year}')

    # Update status
    if request.method == 'POST' and request.POST.get('action') == 'update_status':
        payroll_id = request.POST.get('payroll_id')
        new_status = request.POST.get('new_status')
        payroll = get_object_or_404(Payroll, pk=payroll_id)
        payroll.status = new_status
        payroll.save()
        if new_status in ['confirmed', 'paid']:
            from employees.models import Notification
            Notification.objects.create(
                employee=payroll.employee,
                title=f'Payslip {payroll.get_status_display()}',
                body=f'Your payslip for {calendar.month_name[payroll.month]} {payroll.year} is {payroll.get_status_display().lower()}. Net: {payroll.net_pay}',
                notification_type='general',
            )
            send_push_notification(payroll.employee, f'Payslip {payroll.get_status_display()}',
                                   f'Net pay for {calendar.month_name[payroll.month]}: {payroll.net_pay}')
        messages.success(request, f'Payroll status updated to {new_status}.')
        return redirect(f'/hr/payroll/?month={month}&year={year}')

    payrolls = Payroll.objects.filter(month=month, year=year).select_related('employee__user')
    total_net = sum(p.net_pay for p in payrolls)

    months = [(i, calendar.month_name[i]) for i in range(1, 13)]
    years = list(range(now.year - 1, now.year + 2))

    context = {
        'payrolls': payrolls,
        'month': month,
        'year': year,
        'month_name': calendar.month_name[month],
        'months': months,
        'years': years,
        'total_net': total_net,
    }
    return render(request, 'hr/payroll_list.html', context)


# ============================================================
# Certificates
# ============================================================

@login_required
def certificate_template_list(request):
    """List all certificate templates"""
    from employees.models import CertificateTemplate
    templates = CertificateTemplate.objects.all()
    return render(request, 'hr/certificate_template_list.html', {'templates': templates})


@login_required
def certificate_template_create(request):
    """Create a new certificate template"""
    from employees.models import CertificateTemplate
    if request.method == 'POST':
        CertificateTemplate.objects.create(
            name=request.POST.get('name'),
            certificate_type=request.POST.get('certificate_type', 'inter'),
            title=request.POST.get('title'),
            body_text=request.POST.get('body_text', ''),
            wish_text=request.POST.get('wish_text', ''),
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Template created.')
        return redirect('certificate_template_list')
    context = {
        'type_choices': CertificateTemplate.CERTIFICATE_TYPE_CHOICES,
    }
    return render(request, 'hr/certificate_template_form.html', context)


@login_required
def certificate_template_detail(request, pk):
    """Edit a certificate template"""
    from employees.models import CertificateTemplate
    template = get_object_or_404(CertificateTemplate, pk=pk)
    if request.method == 'POST':
        template.name = request.POST.get('name', template.name)
        template.certificate_type = request.POST.get('certificate_type', template.certificate_type)
        template.title = request.POST.get('title', template.title)
        template.body_text = request.POST.get('body_text', template.body_text)
        template.wish_text = request.POST.get('wish_text', template.wish_text)
        template.is_active = request.POST.get('is_active') == 'on'
        template.save()
        messages.success(request, 'Template updated.')
        return redirect('certificate_template_detail', pk=pk)
    context = {
        'template': template,
        'type_choices': CertificateTemplate.CERTIFICATE_TYPE_CHOICES,
    }
    return render(request, 'hr/certificate_template_form.html', context)


@login_required
def certificate_template_delete(request, pk):
    """Delete a certificate template"""
    from employees.models import CertificateTemplate
    if request.method == 'POST':
        template = get_object_or_404(CertificateTemplate, pk=pk)
        template.delete()
        messages.success(request, 'Template deleted.')
    return redirect('certificate_template_list')


@login_required
def certificate_list(request):
    """List all certificates"""
    from employees.models import Certificate

    search = request.GET.get('search', '')
    certificates = Certificate.objects.all().order_by('-created_at')
    if search:
        certificates = certificates.filter(
            models.Q(student_name__icontains=search) |
            models.Q(certificate_number__icontains=search) |
            models.Q(course_name__icontains=search) |
            models.Q(college_name__icontains=search)
        )

    context = {
        'certificates': certificates,
        'search': search,
    }
    return render(request, 'hr/certificate_list.html', context)


@login_required
def certificate_create(request):
    """Create a new certificate"""
    from employees.models import Certificate, CertificateTemplate

    if request.method == 'POST':
        skills_raw = request.POST.get('skills', '')
        skills = [s.strip() for s in skills_raw.split('\n') if s.strip()]

        template_id = request.POST.get('template') or None
        template = None
        if template_id:
            try:
                template = CertificateTemplate.objects.get(pk=template_id)
            except CertificateTemplate.DoesNotExist:
                pass

        cert = Certificate(
            template=template,
            certificate_type=request.POST.get('certificate_type', 'inter'),
            title=request.POST.get('title', 'INTERNSHIP CERTIFICATE'),
            salutation=request.POST.get('salutation', 'Mr.'),
            student_name=request.POST.get('student_name'),
            gender=request.POST.get('gender', 'male'),
            college_name=request.POST.get('college_name', ''),
            course_name=request.POST.get('course_name', ''),
            start_date=request.POST.get('start_date') or None,
            end_date=request.POST.get('end_date') or None,
            duration_days=int(request.POST.get('duration_days')) if request.POST.get('duration_days') else None,
            mode=request.POST.get('mode', 'offline'),
            skills=skills,
            body_text=request.POST.get('body_text', ''),
            wish_text=request.POST.get('wish_text',
                'We wish {pronoun} success in {possessive} future academic and professional endeavors.'),
            date_of_issuance=request.POST.get('date_of_issuance'),
            issued_by=request.user,
        )
        cert.save()
        messages.success(request, f'Certificate {cert.certificate_number} created for {cert.student_name}.')
        return redirect('certificate_detail', pk=cert.pk)

    templates = CertificateTemplate.objects.filter(is_active=True)
    templates_json = {
        str(t.pk): {'body': t.body_text, 'wish': t.wish_text}
        for t in templates
    }
    context = {
        'salutation_choices': Certificate.SALUTATION_CHOICES,
        'gender_choices': Certificate.GENDER_CHOICES,
        'mode_choices': Certificate.MODE_CHOICES,
        'type_choices': Certificate.CERTIFICATE_TYPE_CHOICES,
        'templates': templates,
        'templates_json': templates_json,
    }
    return render(request, 'hr/certificate_create.html', context)


@login_required
def certificate_detail(request, pk):
    """View and edit a certificate"""
    from employees.models import Certificate

    cert = get_object_or_404(Certificate, pk=pk)

    if request.method == 'POST':
        action = request.POST.get('action', '')

        # Handle status changes
        if action == 'publish':
            if cert.status == 'draft':
                cert.status = 'published'
                cert.save()
                messages.success(request, f'Certificate {cert.certificate_number} published.')
            return redirect('certificate_detail', pk=pk)

        if action == 'cancel':
            if cert.status in ('draft', 'published'):
                cert.status = 'cancelled'
                cert.save()
                messages.success(request, f'Certificate {cert.certificate_number} cancelled.')
            return redirect('certificate_detail', pk=pk)

        if action == 'revert_draft':
            if cert.status == 'published':
                cert.status = 'draft'
                cert.save()
                messages.success(request, 'Certificate reverted to draft.')
            return redirect('certificate_detail', pk=pk)

        # Block edits on published/cancelled
        if cert.status == 'published':
            messages.error(request, 'Published certificates cannot be edited. Revert to draft first.')
            return redirect('certificate_detail', pk=pk)
        if cert.status == 'cancelled':
            messages.error(request, 'Cancelled certificates cannot be edited.')
            return redirect('certificate_detail', pk=pk)

        skills_raw = request.POST.get('skills', '')
        skills = [s.strip() for s in skills_raw.split('\n') if s.strip()]

        cert.certificate_type = request.POST.get('certificate_type', cert.certificate_type)
        cert.title = request.POST.get('title', cert.title)
        cert.salutation = request.POST.get('salutation', cert.salutation)
        cert.student_name = request.POST.get('student_name', cert.student_name)
        cert.gender = request.POST.get('gender', cert.gender)
        cert.college_name = request.POST.get('college_name', '')
        cert.course_name = request.POST.get('course_name', cert.course_name)
        cert.start_date = request.POST.get('start_date') or None
        cert.end_date = request.POST.get('end_date') or None
        cert.duration_days = int(request.POST.get('duration_days')) if request.POST.get('duration_days') else None
        cert.mode = request.POST.get('mode', cert.mode)
        cert.skills = skills
        cert.body_text = request.POST.get('body_text', cert.body_text)
        cert.wish_text = request.POST.get('wish_text', cert.wish_text)
        cert.date_of_issuance = request.POST.get('date_of_issuance', cert.date_of_issuance)
        cert.save()
        messages.success(request, 'Certificate updated.')
        return redirect('certificate_detail', pk=pk)

    context = {
        'cert': cert,
        'skills_text': '\n'.join(cert.skills) if cert.skills else '',
        'salutation_choices': Certificate.SALUTATION_CHOICES,
        'gender_choices': Certificate.GENDER_CHOICES,
        'mode_choices': Certificate.MODE_CHOICES,
        'type_choices': Certificate.CERTIFICATE_TYPE_CHOICES,
    }
    return render(request, 'hr/certificate_detail.html', context)


@login_required
def certificate_delete(request, pk):
    """Delete a certificate"""
    from employees.models import Certificate
    if request.method == 'POST':
        cert = get_object_or_404(Certificate, pk=pk)
        cert.delete()
        messages.success(request, 'Certificate deleted.')
    return redirect('certificate_list')


@login_required
def certificate_pdf(request, pk):
    """Generate certificate PDF from the custom backend"""
    from employees.models import Certificate
    from django.template.loader import render_to_string
    import weasyprint
    import qrcode
    import base64
    from io import BytesIO

    cert = get_object_or_404(Certificate, pk=pk)

    # Generate QR code
    verify_url = request.build_absolute_uri(f'/api/employees/certificates/verify/{cert.verification_id}/')
    qr = qrcode.QRCode(version=1, box_size=10, border=1)
    qr.add_data(verify_url)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    qr_base64 = base64.b64encode(buffer.getvalue()).decode()

    # Asset paths
    from django.conf import settings as django_settings
    static_dir = django_settings.BASE_DIR / 'static' / 'certificates'
    header_logo = (static_dir / 'headerlogo.png').as_uri()
    signature = (static_dir / 'jobin_signature.png').as_uri()
    seal = (static_dir / 'seal.png').as_uri()
    footer_logo = (static_dir / 'footer_right_logo.png').as_uri()
    award_badge = (static_dir / 'award_badge.png').as_uri()
    bottom_graphics = (static_dir / 'bottom_graphics.png').as_uri()

    def format_date(d):
        day = d.day
        if 4 <= day <= 20 or 24 <= day <= 30:
            suffix = "th"
        else:
            suffix = ["st", "nd", "rd"][day % 10 - 1]
        return f"{day}{suffix} {d.strftime('%B %Y')}"

    # Render body_text with placeholders
    from django.utils.html import escape
    skills_html = ''
    if cert.skills:
        items = ''.join(f'<li>{escape(s)}</li>' for s in cert.skills)
        skills_html = f'<ul class="skills-list">{items}</ul>'

    body_raw = cert.body_text or ''
    try:
        body_rendered = body_raw.format(
            salutation=cert.salutation,
            student_name=cert.student_name,
            college_name=cert.college_name or '',
            course_name=cert.course_name or '',
            start_date=format_date(cert.start_date) if cert.start_date else '',
            end_date=format_date(cert.end_date) if cert.end_date else '',
            duration_days=cert.duration_days or '',
            mode=cert.get_mode_display() if cert.mode else '',
            skills=skills_html,
            pronoun=cert.pronoun,
            pronoun_cap=cert.pronoun_cap,
            possessive=cert.possessive,
            object_pronoun=cert.object_pronoun,
        )
    except (KeyError, IndexError):
        body_rendered = body_raw

    # Convert **bold** to <strong> tags
    import re
    body_rendered = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', body_rendered)

    paragraphs = body_rendered.split('\n\n')
    rendered_body = ''
    for p in paragraphs:
        p = p.strip()
        if not p:
            continue
        if p.startswith('<ul'):
            rendered_body += p
        else:
            rendered_body += f'<p class="body-text">{p}</p>'

    wish_text = cert.wish_text.format(
        pronoun=cert.object_pronoun,
        possessive=cert.possessive,
    )

    context = {
        'cert': cert,
        'qr_base64': qr_base64,
        'header_logo': header_logo,
        'signature': signature,
        'seal': seal,
        'footer_logo': footer_logo,
        'award_badge': award_badge,
        'bottom_graphics': bottom_graphics,
        'rendered_body': rendered_body,
        'date_of_issuance_fmt': cert.date_of_issuance.strftime('%d/%m/%Y'),
        'wish_text': wish_text,
    }

    html_string = render_to_string('employees/certificate_pdf.html', context)
    pdf = weasyprint.HTML(string=html_string).write_pdf()

    from django.http import HttpResponse as HR
    response = HR(pdf, content_type='application/pdf')
    filename = f"Certificate_{cert.student_name.replace(' ', '_')}_{cert.certificate_number.replace('/', '_')}.pdf"
    if request.GET.get('download'):
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
    else:
        response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


# ============== Feature Request Links ==============

@login_required
def feature_request_list(request):
    """Admin: list all feature-request links with status filter."""
    status = request.GET.get('status', '')
    qs = FeatureRequestLink.objects.select_related('client', 'created_by', 'selected_project_type').all()
    if status == 'pending':
        qs = qs.filter(submitted_at__isnull=True)
    elif status == 'submitted':
        qs = qs.filter(submitted_at__isnull=False)
    return render(request, 'feature_requests/list.html', {
        'links': qs,
        'status_filter': status,
    })


@login_required
def feature_request_create(request):
    """Admin: create a new shareable link for a client."""
    if request.method == 'POST':
        client_id = request.POST.get('client_id')
        client = get_object_or_404(Client, pk=client_id)
        link = FeatureRequestLink.objects.create(client=client, created_by=request.user)
        messages.success(request, f'Link created for {client}. OTP: {link.otp}')
        return redirect('feature_request_detail', pk=link.pk)
    clients = Client.objects.filter(is_active=True).order_by('name')
    return render(request, 'feature_requests/create.html', {'clients': clients})


@login_required
def feature_request_detail(request, pk):
    """Admin: view a link (show URL + OTP, or submission details)."""
    link = get_object_or_404(
        FeatureRequestLink.objects.select_related('client', 'selected_project_type', 'created_by')
                                   .prefetch_related('selected_features__project_type'),
        pk=pk,
    )
    share_url = request.build_absolute_uri(
        reverse('public_feature_request', kwargs={'token': link.token})
    )
    return render(request, 'feature_requests/detail.html', {
        'link': link,
        'share_url': share_url,
    })


@login_required
def feature_request_regenerate(request, pk):
    """Admin: regenerate OTP (and optionally the token) for a link."""
    import secrets
    if request.method != 'POST':
        return redirect('feature_request_detail', pk=pk)
    link = get_object_or_404(FeatureRequestLink, pk=pk)
    link.otp = f"{secrets.randbelow(1000000):06d}"
    if request.POST.get('reset_submission'):
        link.submitted_at = None
        link.selected_project_type = None
        link.client_notes = ''
        link.selected_features.clear()
        link.token = uuid.uuid4()
    link.save()
    messages.success(request, f'Regenerated. New OTP: {link.otp}')
    return redirect('feature_request_detail', pk=link.pk)


@login_required
def feature_request_delete(request, pk):
    if request.method != 'POST':
        return redirect('feature_request_list')
    link = get_object_or_404(FeatureRequestLink, pk=pk)
    link.delete()
    messages.success(request, 'Feature request link deleted.')
    return redirect('feature_request_list')


# ---- Project Types & Features admin ----

@login_required
def project_type_list(request):
    types = ProjectType.objects.prefetch_related('features').all()
    return render(request, 'feature_requests/types_list.html', {'types': types})


@login_required
def project_type_save(request):
    """Create or update a ProjectType via form POST."""
    if request.method != 'POST':
        return redirect('project_type_list')
    pk = request.POST.get('pk')
    name = (request.POST.get('name') or '').strip()
    description = (request.POST.get('description') or '').strip()
    sort_order = int(request.POST.get('sort_order') or 0)
    is_active = request.POST.get('is_active') == 'on'
    if not name:
        messages.error(request, 'Name is required.')
        return redirect('project_type_list')
    if pk:
        pt = get_object_or_404(ProjectType, pk=pk)
        pt.name = name
        pt.description = description
        pt.sort_order = sort_order
        pt.is_active = is_active
        pt.save()
        messages.success(request, 'Project type updated.')
    else:
        ProjectType.objects.create(name=name, description=description,
                                   sort_order=sort_order, is_active=is_active)
        messages.success(request, 'Project type created.')
    return redirect('project_type_list')


@login_required
def project_type_delete(request, pk):
    if request.method != 'POST':
        return redirect('project_type_list')
    pt = get_object_or_404(ProjectType, pk=pk)
    pt.delete()
    messages.success(request, 'Project type deleted.')
    return redirect('project_type_list')


@login_required
def project_feature_save(request):
    """Create or update a ProjectFeature via form POST."""
    if request.method != 'POST':
        return redirect('project_type_list')
    pk = request.POST.get('pk')
    project_type_id = request.POST.get('project_type_id')
    label = (request.POST.get('label') or '').strip()
    description = (request.POST.get('description') or '').strip()
    sort_order = int(request.POST.get('sort_order') or 0)
    is_active = request.POST.get('is_active') == 'on'
    if not label or not project_type_id:
        messages.error(request, 'Label and project type are required.')
        return redirect('project_type_list')
    pt = get_object_or_404(ProjectType, pk=project_type_id)
    if pk:
        feat = get_object_or_404(ProjectFeature, pk=pk)
        feat.project_type = pt
        feat.label = label
        feat.description = description
        feat.sort_order = sort_order
        feat.is_active = is_active
        feat.save()
        messages.success(request, 'Feature updated.')
    else:
        ProjectFeature.objects.create(project_type=pt, label=label, description=description,
                                      sort_order=sort_order, is_active=is_active)
        messages.success(request, 'Feature added.')
    return redirect('project_type_list')


@login_required
def project_feature_delete(request, pk):
    if request.method != 'POST':
        return redirect('project_type_list')
    feat = get_object_or_404(ProjectFeature, pk=pk)
    feat.delete()
    messages.success(request, 'Feature deleted.')
    return redirect('project_type_list')


# ---- Public (no login) ----

def public_feature_request(request, token):
    """Client-facing page. OTP gate, then feature selection form."""
    link = FeatureRequestLink.objects.filter(token=token).select_related('client').first()
    if not link:
        return render(request, 'feature_requests/public_invalid.html', status=404)
    if link.is_submitted:
        return render(request, 'feature_requests/public_already_submitted.html', {'link': link})

    verified = request.session.get(f'fr_verified_{link.pk}') is True

    if request.method == 'POST' and not verified:
        otp = (request.POST.get('otp') or '').strip()
        if otp == link.otp:
            request.session[f'fr_verified_{link.pk}'] = True
            return redirect('public_feature_request', token=token)
        return render(request, 'feature_requests/public_otp.html', {
            'link': link, 'error': 'Invalid OTP. Please try again.',
        })

    if request.method == 'POST' and verified:
        project_type_id = request.POST.get('project_type')
        feature_ids = request.POST.getlist('features')
        notes = (request.POST.get('notes') or '').strip()
        if not project_type_id:
            return render(request, 'feature_requests/public_form.html', {
                'link': link,
                'project_types': ProjectType.objects.filter(is_active=True).prefetch_related('features'),
                'error': 'Please select a project type.',
            })
        pt = get_object_or_404(ProjectType, pk=project_type_id)
        link.selected_project_type = pt
        link.client_notes = notes
        link.submitted_at = timezone.now()
        link.save()
        link.selected_features.set(
            ProjectFeature.objects.filter(pk__in=feature_ids, project_type=pt)
        )
        request.session.pop(f'fr_verified_{link.pk}', None)
        return render(request, 'feature_requests/public_success.html', {'link': link})

    if not verified:
        return render(request, 'feature_requests/public_otp.html', {'link': link})

    return render(request, 'feature_requests/public_form.html', {
        'link': link,
        'project_types': ProjectType.objects.filter(is_active=True).prefetch_related('features'),
    })


# ============== Bank Accounts & Internal Transfers ==============

@login_required
def bank_account_list(request):
    from .cash_position import cash_position, pending_transfers
    cp = cash_position(include_inactive=True)
    pending = pending_transfers()
    return render(request, 'accounts/list.html', {
        'rows': cp['accounts'],
        'total_assets': cp['total'],
        'pending_transfers': pending,
        'pending_transfer_count': pending.count(),
    })


@login_required
def bank_account_create(request):
    if request.method == 'POST':
        acct = BankAccount(
            name=request.POST.get('name', '').strip(),
            account_type=request.POST.get('account_type', 'bank'),
            bank_name=request.POST.get('bank_name', '').strip(),
            account_number=request.POST.get('account_number', '').strip(),
            ifsc=request.POST.get('ifsc', '').strip(),
            branch=request.POST.get('branch', '').strip(),
            upi_id=request.POST.get('upi_id', '').strip(),
            opening_balance=request.POST.get('opening_balance') or 0,
            opening_date=request.POST.get('opening_date') or timezone.now().date(),
            is_active=request.POST.get('is_active') == 'on',
            is_primary_bank=request.POST.get('is_primary_bank') == 'on',
            is_cash=request.POST.get('is_cash') == 'on',
            display_order=request.POST.get('display_order') or 0,
            notes=request.POST.get('notes', ''),
        )
        acct.save()
        log_activity(request, 'created', acct)
        messages.success(request, f'Account "{acct.name}" created.')
        return redirect('bank_account_list')
    return render(request, 'accounts/form.html', {
        'account': None,
        'type_choices': BankAccount.ACCOUNT_TYPE_CHOICES,
    })


@login_required
def bank_account_update(request, pk):
    acct = get_object_or_404(BankAccount, pk=pk)
    if request.method == 'POST':
        acct.name = request.POST.get('name', '').strip()
        acct.account_type = request.POST.get('account_type', 'bank')
        acct.bank_name = request.POST.get('bank_name', '').strip()
        acct.account_number = request.POST.get('account_number', '').strip()
        acct.ifsc = request.POST.get('ifsc', '').strip()
        acct.branch = request.POST.get('branch', '').strip()
        acct.upi_id = request.POST.get('upi_id', '').strip()
        acct.opening_balance = request.POST.get('opening_balance') or 0
        acct.opening_date = request.POST.get('opening_date') or acct.opening_date
        acct.is_active = request.POST.get('is_active') == 'on'
        acct.is_primary_bank = request.POST.get('is_primary_bank') == 'on'
        acct.is_cash = request.POST.get('is_cash') == 'on'
        acct.display_order = request.POST.get('display_order') or 0
        acct.notes = request.POST.get('notes', '')
        acct.save()
        log_activity(request, 'updated', acct)
        messages.success(request, f'Account "{acct.name}" updated.')
        return redirect('bank_account_list')
    return render(request, 'accounts/form.html', {
        'account': acct,
        'type_choices': BankAccount.ACCOUNT_TYPE_CHOICES,
    })


@login_required
def bank_account_detail(request, pk):
    """Ledger view: payments + expenses + transfers for one account, chronological."""
    from .cash_position import compute_account_balance
    acct = get_object_or_404(BankAccount, pk=pk)
    balance = compute_account_balance(acct)

    entries = []
    pay_methods = acct.resolved_payment_methods()
    exp_methods = acct.resolved_expense_methods()
    if pay_methods:
        for p in Payment.objects.filter(
            payment_date__gte=acct.opening_date,
            payment_method__in=pay_methods,
        ).select_related('invoice', 'invoice__client').order_by('-payment_date'):
            entries.append({
                'date': p.payment_date, 'kind': 'Payment in',
                'description': f"Invoice #{p.invoice.invoice_number} - {p.invoice.client.name}",
                'method': p.get_payment_method_display(),
                'amount': p.amount, 'direction': 'in',
                'url': reverse('invoice_detail', args=[p.invoice.pk]),
            })
    if exp_methods:
        for e in Expense.objects.filter(
            date__gte=acct.opening_date,
            payment_method__in=exp_methods,
        ).order_by('-date'):
            entries.append({
                'date': e.date, 'kind': 'Expense',
                'description': f"{e.vendor} - {e.get_category_display()}",
                'method': e.get_payment_method_display(),
                'amount': e.amount, 'direction': 'out',
                'url': reverse('expense_list'),
            })
    for t in InternalTransfer.objects.filter(to_account=acct).select_related('from_account'):
        entries.append({
            'date': t.date, 'kind': 'Transfer in',
            'description': f"From {t.from_account.name}" + (f" (ref: {t.reference})" if t.reference else ''),
            'method': 'Internal', 'amount': t.amount, 'direction': 'in',
            'url': reverse('transfer_list'),
        })
    for t in InternalTransfer.objects.filter(from_account=acct).select_related('to_account'):
        entries.append({
            'date': t.date, 'kind': 'Transfer out',
            'description': f"To {t.to_account.name}" + (f" (ref: {t.reference})" if t.reference else ''),
            'method': 'Internal', 'amount': t.amount, 'direction': 'out',
            'url': reverse('transfer_list'),
        })
    entries.sort(key=lambda x: x['date'], reverse=True)

    return render(request, 'accounts/detail.html', {
        'account': acct,
        'balance': balance,
        'entries': entries,
    })


@login_required
def transfer_list(request):
    transfers = InternalTransfer.objects.select_related('from_account', 'to_account', 'created_by').all()
    today = timezone.now().date()
    from_account_id = request.GET.get('account')
    if from_account_id:
        transfers = transfers.filter(
            Q(from_account_id=from_account_id) | Q(to_account_id=from_account_id)
        )
    return render(request, 'transfers/list.html', {
        'transfers': transfers,
        'today': today,
        'accounts': BankAccount.objects.filter(is_active=True),
        'selected_account': from_account_id,
    })


def _save_transfer(transfer, request, is_new):
    transfer.from_account_id = request.POST.get('from_account')
    transfer.to_account_id = request.POST.get('to_account')
    transfer.amount = request.POST.get('amount')
    transfer.date = request.POST.get('date') or timezone.now().date()
    transfer.reference = request.POST.get('reference', '').strip()
    transfer.notes = request.POST.get('notes', '')
    if is_new and request.user.is_authenticated:
        transfer.created_by = request.user
    transfer.full_clean()
    transfer.save()


_TRANSFER_FORM_KEYS = ['from_account', 'to_account', 'amount', 'date', 'reference', 'notes']


def _empty_transfer_form():
    return {k: '' for k in _TRANSFER_FORM_KEYS}


def _form_data_from_post(post):
    return {k: post.get(k, '') for k in _TRANSFER_FORM_KEYS}


@login_required
def transfer_create(request):
    accounts = BankAccount.objects.filter(is_active=True)
    if request.method == 'POST':
        transfer = InternalTransfer()
        try:
            _save_transfer(transfer, request, is_new=True)
        except Exception as e:
            messages.error(request, f'Could not save transfer: {e}')
            return render(request, 'transfers/form.html', {
                'transfer': None, 'accounts': accounts,
                'form_data': _form_data_from_post(request.POST),
            })
        log_activity(request, 'created', transfer)
        messages.success(request, 'Transfer recorded.')
        return redirect('transfer_list')
    return render(request, 'transfers/form.html', {
        'transfer': None, 'accounts': accounts, 'form_data': _empty_transfer_form(),
    })


@login_required
def transfer_update(request, pk):
    transfer = get_object_or_404(InternalTransfer, pk=pk)
    accounts = BankAccount.objects.filter(is_active=True)
    if request.method == 'POST':
        try:
            _save_transfer(transfer, request, is_new=False)
        except Exception as e:
            messages.error(request, f'Could not save transfer: {e}')
            return render(request, 'transfers/form.html', {
                'transfer': transfer, 'accounts': accounts,
                'form_data': _form_data_from_post(request.POST),
            })
        log_activity(request, 'updated', transfer)
        messages.success(request, 'Transfer updated.')
        return redirect('transfer_list')
    return render(request, 'transfers/form.html', {
        'transfer': transfer, 'accounts': accounts, 'form_data': _empty_transfer_form(),
    })


@login_required
def transfer_delete(request, pk):
    transfer = get_object_or_404(InternalTransfer, pk=pk)
    if request.method == 'POST':
        log_activity(request, 'deleted', transfer)
        transfer.delete()
        messages.success(request, 'Transfer deleted.')
        return redirect('transfer_list')
    return render(request, 'transfers/confirm_delete.html', {'transfer': transfer})


# ============== Company Documents ==============

_COMPANY_DOC_FORM_KEYS = ['title', 'document_type', 'issuer', 'reference_number', 'issue_date', 'expiry_date', 'notes']


def _company_doc_form_data(source=None, doc=None):
    """Return a dict with every form key populated (template never hits a missing key)."""
    if doc is not None:
        return {
            'title': doc.title or '',
            'document_type': doc.document_type or 'other',
            'issuer': doc.issuer or '',
            'reference_number': doc.reference_number or '',
            'issue_date': doc.issue_date.strftime('%Y-%m-%d') if doc.issue_date else '',
            'expiry_date': doc.expiry_date.strftime('%Y-%m-%d') if doc.expiry_date else '',
            'notes': doc.notes or '',
        }
    if source is not None:
        return {k: source.get(k, '') for k in _COMPANY_DOC_FORM_KEYS}
    return {k: '' for k in _COMPANY_DOC_FORM_KEYS}


def _save_company_document(doc, request, is_new):
    doc.title = request.POST.get('title', '').strip()
    doc.document_type = request.POST.get('document_type', 'other')
    doc.issuer = request.POST.get('issuer', '').strip()
    doc.reference_number = request.POST.get('reference_number', '').strip()
    doc.issue_date = request.POST.get('issue_date') or None
    doc.expiry_date = request.POST.get('expiry_date') or None
    doc.notes = request.POST.get('notes', '')
    if 'file' in request.FILES:
        doc.file = request.FILES['file']
    if is_new and request.user.is_authenticated:
        doc.uploaded_by = request.user
    doc.save()


@login_required
def company_document_list(request):
    docs = CompanyDocument.objects.all()
    doc_type = request.GET.get('type')
    if doc_type:
        docs = docs.filter(document_type=doc_type)
    today = timezone.now().date()
    soon = today + timedelta(days=30)
    expiring_soon = CompanyDocument.objects.filter(expiry_date__gte=today, expiry_date__lte=soon).count()
    expired = CompanyDocument.objects.filter(expiry_date__lt=today).count()
    return render(request, 'company_documents/list.html', {
        'documents': docs,
        'selected_type': doc_type,
        'type_choices': CompanyDocument.DOCUMENT_TYPE_CHOICES,
        'expiring_soon_count': expiring_soon,
        'expired_count': expired,
    })


@login_required
def company_document_create(request):
    if request.method == 'POST':
        if 'file' not in request.FILES:
            messages.error(request, 'Please attach a file.')
            return render(request, 'company_documents/form.html', {
                'document': None,
                'type_choices': CompanyDocument.DOCUMENT_TYPE_CHOICES,
                'form_data': _company_doc_form_data(source=request.POST),
            })
        doc = CompanyDocument()
        _save_company_document(doc, request, is_new=True)
        log_activity(request, 'created', doc)
        messages.success(request, f'Document "{doc.title}" uploaded.')
        return redirect('company_document_list')
    return render(request, 'company_documents/form.html', {
        'document': None,
        'type_choices': CompanyDocument.DOCUMENT_TYPE_CHOICES,
        'form_data': _company_doc_form_data(),
    })


@login_required
def company_document_update(request, pk):
    doc = get_object_or_404(CompanyDocument, pk=pk)
    if request.method == 'POST':
        _save_company_document(doc, request, is_new=False)
        log_activity(request, 'updated', doc)
        messages.success(request, f'Document "{doc.title}" updated.')
        return redirect('company_document_list')
    return render(request, 'company_documents/form.html', {
        'document': doc,
        'type_choices': CompanyDocument.DOCUMENT_TYPE_CHOICES,
        'form_data': _company_doc_form_data(doc=doc),
    })


@login_required
def company_document_delete(request, pk):
    doc = get_object_or_404(CompanyDocument, pk=pk)
    if request.method == 'POST':
        log_activity(request, 'deleted', doc)
        doc.delete()
        messages.success(request, 'Document deleted.')
        return redirect('company_document_list')
    return render(request, 'company_documents/confirm_delete.html', {'document': doc})


@login_required
def company_document_download(request, pk):
    from django.http import FileResponse, Http404
    doc = get_object_or_404(CompanyDocument, pk=pk)
    if not doc.file:
        raise Http404
    return FileResponse(doc.file.open('rb'), as_attachment=True, filename=doc.file.name.split('/')[-1])


# ============== Partners & Capital Contributions ==============

_PARTNER_FORM_KEYS = ['name', 'title', 'email', 'phone', 'join_date', 'notes']


def _partner_form_data(source=None, partner=None):
    if partner is not None:
        return {
            'name': partner.name or '',
            'title': partner.title or '',
            'email': partner.email or '',
            'phone': partner.phone or '',
            'join_date': partner.join_date.strftime('%Y-%m-%d') if partner.join_date else '',
            'notes': partner.notes or '',
            'is_active': partner.is_active,
        }
    if source is not None:
        d = {k: source.get(k, '') for k in _PARTNER_FORM_KEYS}
        d['is_active'] = source.get('is_active') == 'on'
        return d
    return {**{k: '' for k in _PARTNER_FORM_KEYS}, 'is_active': True}


def _save_partner(partner, request, is_new):
    partner.name = request.POST.get('name', '').strip()
    partner.title = request.POST.get('title', '').strip()
    partner.email = request.POST.get('email', '').strip()
    partner.phone = request.POST.get('phone', '').strip()
    partner.join_date = request.POST.get('join_date') or timezone.now().date()
    partner.is_active = request.POST.get('is_active') == 'on'
    partner.notes = request.POST.get('notes', '')
    if 'photo' in request.FILES:
        partner.photo = request.FILES['photo']
    partner.save()


@login_required
def partner_list(request):
    from decimal import Decimal
    partners = list(Partner.objects.all())
    total_capital = sum((p.total_contribution for p in partners), Decimal('0')) if partners else Decimal('0')
    return render(request, 'partners/list.html', {
        'partners': partners,
        'total_capital': total_capital,
    })


@login_required
def partner_create(request):
    if request.method == 'POST':
        partner = Partner()
        _save_partner(partner, request, is_new=True)
        log_activity(request, 'created', partner)
        messages.success(request, f'Partner "{partner.name}" added.')
        return redirect('partner_detail', pk=partner.pk)
    return render(request, 'partners/form.html', {
        'partner': None,
        'form_data': _partner_form_data(),
    })


@login_required
def partner_update(request, pk):
    partner = get_object_or_404(Partner, pk=pk)
    if request.method == 'POST':
        _save_partner(partner, request, is_new=False)
        log_activity(request, 'updated', partner)
        messages.success(request, f'Partner "{partner.name}" updated.')
        return redirect('partner_detail', pk=partner.pk)
    return render(request, 'partners/form.html', {
        'partner': partner,
        'form_data': _partner_form_data(partner=partner),
    })


@login_required
def partner_detail(request, pk):
    partner = get_object_or_404(Partner, pk=pk)
    contributions = partner.contributions.select_related('bank_account').all()
    return render(request, 'partners/detail.html', {
        'partner': partner,
        'contributions': contributions,
        'total': partner.total_contribution,
    })


@login_required
def partner_delete(request, pk):
    partner = get_object_or_404(Partner, pk=pk)
    if request.method == 'POST':
        log_activity(request, 'deleted', partner)
        partner.delete()
        messages.success(request, 'Partner deleted.')
        return redirect('partner_list')
    return render(request, 'partners/confirm_delete.html', {
        'partner': partner,
        'contribution_count': partner.contributions.count(),
    })


_CONTRIB_FORM_KEYS = ['date', 'amount', 'contribution_type', 'bank_account', 'description']


def _contribution_form_data(source=None, contribution=None):
    if contribution is not None:
        return {
            'date': contribution.date.strftime('%Y-%m-%d') if contribution.date else '',
            'amount': str(contribution.amount or ''),
            'contribution_type': contribution.contribution_type or 'bank_transfer',
            'bank_account': str(contribution.bank_account_id or ''),
            'description': contribution.description or '',
        }
    if source is not None:
        return {k: source.get(k, '') for k in _CONTRIB_FORM_KEYS}
    return {k: '' for k in _CONTRIB_FORM_KEYS}


def _save_contribution(contribution, request, is_new):
    contribution.date = request.POST.get('date') or timezone.now().date()
    contribution.amount = request.POST.get('amount') or 0
    contribution.contribution_type = request.POST.get('contribution_type', 'bank_transfer')
    bank_id = request.POST.get('bank_account') or None
    contribution.bank_account_id = bank_id if bank_id else None
    contribution.description = request.POST.get('description', '').strip()
    if 'receipt' in request.FILES:
        contribution.receipt = request.FILES['receipt']
    contribution.save()


@login_required
def contribution_create(request, partner_pk):
    partner = get_object_or_404(Partner, pk=partner_pk)
    accounts = BankAccount.objects.filter(is_active=True)
    if request.method == 'POST':
        contribution = CapitalContribution(partner=partner)
        try:
            _save_contribution(contribution, request, is_new=True)
        except Exception as e:
            messages.error(request, f'Could not save contribution: {e}')
            return render(request, 'partners/contribution_form.html', {
                'partner': partner,
                'contribution': None,
                'accounts': accounts,
                'type_choices': CapitalContribution.CONTRIBUTION_TYPE_CHOICES,
                'form_data': _contribution_form_data(source=request.POST),
            })
        log_activity(request, 'created', contribution)
        messages.success(request, f'Contribution of ₹{contribution.amount} recorded.')
        return redirect('partner_detail', pk=partner.pk)
    return render(request, 'partners/contribution_form.html', {
        'partner': partner,
        'contribution': None,
        'accounts': accounts,
        'type_choices': CapitalContribution.CONTRIBUTION_TYPE_CHOICES,
        'form_data': _contribution_form_data(),
    })


@login_required
def contribution_update(request, pk):
    contribution = get_object_or_404(CapitalContribution, pk=pk)
    accounts = BankAccount.objects.filter(is_active=True)
    if request.method == 'POST':
        try:
            _save_contribution(contribution, request, is_new=False)
        except Exception as e:
            messages.error(request, f'Could not save contribution: {e}')
            return render(request, 'partners/contribution_form.html', {
                'partner': contribution.partner,
                'contribution': contribution,
                'accounts': accounts,
                'type_choices': CapitalContribution.CONTRIBUTION_TYPE_CHOICES,
                'form_data': _contribution_form_data(source=request.POST),
            })
        log_activity(request, 'updated', contribution)
        messages.success(request, 'Contribution updated.')
        return redirect('partner_detail', pk=contribution.partner.pk)
    return render(request, 'partners/contribution_form.html', {
        'partner': contribution.partner,
        'contribution': contribution,
        'accounts': accounts,
        'type_choices': CapitalContribution.CONTRIBUTION_TYPE_CHOICES,
        'form_data': _contribution_form_data(contribution=contribution),
    })


@login_required
def contribution_delete(request, pk):
    contribution = get_object_or_404(CapitalContribution, pk=pk)
    partner_pk = contribution.partner.pk
    if request.method == 'POST':
        log_activity(request, 'deleted', contribution)
        contribution.delete()
        messages.success(request, 'Contribution deleted.')
        return redirect('partner_detail', pk=partner_pk)
    return render(request, 'partners/contribution_confirm_delete.html', {'contribution': contribution})


# ============== Company Assets (deposits / advances / equipment) ==============

_ASSET_FORM_KEYS = ['name', 'asset_type', 'amount', 'acquired_date', 'counterparty', 'expected_return_date', 'notes']


def _asset_form_data(source=None, asset=None):
    if asset is not None:
        return {
            'name': asset.name or '',
            'asset_type': asset.asset_type or 'rent_deposit',
            'amount': str(asset.amount or ''),
            'acquired_date': asset.acquired_date.strftime('%Y-%m-%d') if asset.acquired_date else '',
            'counterparty': asset.counterparty or '',
            'expected_return_date': asset.expected_return_date.strftime('%Y-%m-%d') if asset.expected_return_date else '',
            'notes': asset.notes or '',
            'is_refundable': asset.is_refundable,
            'is_active': asset.is_active,
        }
    if source is not None:
        d = {k: source.get(k, '') for k in _ASSET_FORM_KEYS}
        d['is_refundable'] = source.get('is_refundable') == 'on'
        d['is_active'] = source.get('is_active') == 'on'
        return d
    return {**{k: '' for k in _ASSET_FORM_KEYS}, 'is_refundable': True, 'is_active': True}


def _save_asset(asset, request, is_new):
    asset.name = request.POST.get('name', '').strip()
    asset.asset_type = request.POST.get('asset_type', 'rent_deposit')
    asset.amount = request.POST.get('amount') or 0
    asset.acquired_date = request.POST.get('acquired_date') or timezone.now().date()
    asset.counterparty = request.POST.get('counterparty', '').strip()
    asset.expected_return_date = request.POST.get('expected_return_date') or None
    asset.is_refundable = request.POST.get('is_refundable') == 'on'
    asset.is_active = request.POST.get('is_active') == 'on'
    asset.notes = request.POST.get('notes', '')
    asset.save()


@login_required
def asset_list(request):
    from decimal import Decimal
    assets = CompanyAsset.objects.all()
    active_assets = assets.filter(is_active=True)
    total_active = active_assets.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_refundable = active_assets.filter(is_refundable=True).aggregate(t=Sum('amount'))['t'] or Decimal('0')
    return render(request, 'assets/list.html', {
        'assets': assets,
        'total_active': total_active,
        'total_refundable': total_refundable,
        'type_choices': CompanyAsset.ASSET_TYPE_CHOICES,
    })


@login_required
def asset_create(request):
    if request.method == 'POST':
        asset = CompanyAsset()
        _save_asset(asset, request, is_new=True)
        log_activity(request, 'created', asset)
        messages.success(request, f'Asset "{asset.name}" recorded.')
        return redirect('asset_list')
    return render(request, 'assets/form.html', {
        'asset': None,
        'type_choices': CompanyAsset.ASSET_TYPE_CHOICES,
        'form_data': _asset_form_data(),
    })


@login_required
def asset_update(request, pk):
    asset = get_object_or_404(CompanyAsset, pk=pk)
    if request.method == 'POST':
        _save_asset(asset, request, is_new=False)
        log_activity(request, 'updated', asset)
        messages.success(request, f'Asset "{asset.name}" updated.')
        return redirect('asset_list')
    return render(request, 'assets/form.html', {
        'asset': asset,
        'type_choices': CompanyAsset.ASSET_TYPE_CHOICES,
        'form_data': _asset_form_data(asset=asset),
    })


@login_required
def asset_delete(request, pk):
    asset = get_object_or_404(CompanyAsset, pk=pk)
    if request.method == 'POST':
        log_activity(request, 'deleted', asset)
        asset.delete()
        messages.success(request, 'Asset deleted.')
        return redirect('asset_list')
    return render(request, 'assets/confirm_delete.html', {'asset': asset})
